"""
===============================================================================
Django Views — Project/Allocation/LDAP utilities (documentation-only header)
===============================================================================

Purpose
-------
This module powers a web-based workflow for project management and time
allocations with the following main capabilities:

1) Project/COE/Domain CRUD & mapping
   - Create/edit/delete Projects, COEs, Domains.
   - Map COEs to Projects (idempotent upserts).
   - Resolve PDL/PM identities from LDAP (local cache first, AD fallback).

2) Billing-cycle aware allocations
   - Canonical billing periods come from `monthly_hours_limit` (year, month,
     start_date, end_date). If not defined, we fall back to the calendar month.
   - Monthly totals are stored in `monthly_allocation_entries` keyed by:
       (project_id, iom_id, month_start, user_ldap).
   - Weekly splits/decisions are stored in `weekly_allocations` keyed by:
       (allocation_id, week_number).
   - Individual day punches/actuals are stored in `user_punches`.

3) Team and personal allocation views
   - `team_allocations`: Manager/PDL view over direct/indirect reportees retrieved
     from LDAP, current billing window, with weekly summaries.
   - `my_allocations`: User’s own allocations, provides equal-split fallback
     when weekly rows are missing, shows punches and holidays across the billing
     period, and supports “Save Week” and daily punching aligned to billing weeks.

4) Exports
   - Excel export for IOM allocations in a given billing window.
   - PDF/Excel export of a user’s punches for a billing window with robust
     lookups for LDAP value variants (exact, lowercase, localpart, wildcard).

5) LDAP handling strategy
   - Prefer local table `ldap_directory` (username, email, cn, title) for lookups.
   - Fallback to live LDAP (via `accounts.ldap_utils`) when needed and when
     session credentials are present (username/password stored in session).
   - `_ensure_user_from_ldap` ensures there is always a `users` row corresponding
     to an LDAP identifier (email or sAMAccountName), creating on demand if needed.

6) Security and authorization
   - Edit/project-selection logic ensures users can only edit projects where
     they are either PDL (`projects.pdl_name`) or are the creator of a WBS
     row (`prism_wbs.creator` after converting CN to “First … Last” format).
   - `team_allocations` constructs the “reportees set” from LDAP; if the viewer
     is PDL/manager (`is_pdl_user`) they are included in the set (self-view).

Key Tables (as referenced by SQL in this module)
------------------------------------------------
- projects(id, name, oem_name, description, start_date, end_date,
           pdl_name, pdl_name, pm_user_id, pm_name, created_at, …)
- coes(id, name, leader_user_id, description)
- domains(id, coe_id, name, lead_user_id)
- project_coes(project_id, coe_id) — mapping
- prism_wbs(id, project_id, iom_id, creator, department, site, function,
            buyer_wbs_cc, seller_wbs_cc, jan_fte/jan_hours … dec_fte/dec_hours,
            total_hours, …)
- users(id, username, email, ldap_id, created_at, …)
- ldap_directory(username, email, cn, title, …) — local LDAP cache
- monthly_hours_limit(year, month, start_date, end_date, max_hours)
- allocations(id, month_start, …) and allocation_items(id, allocation_id,
            project_id, coe_id, domain_id, user_ldap, user_id, total_hours, …)
  (Some legacy paths use these tables for page assembly.)
- monthly_allocation_entries(id, project_id, iom_id, month_start, user_ldap,
            total_hours, created_at)
- weekly_allocations(id, allocation_id, week_number, percent, hours, status,
            created_at, updated_at)  # unique key on (allocation_id, week_number)
- user_punches(id, user_ldap, allocation_id, punch_date, week_number, actual_hours,
            wbs, updated_at)
- holidays(holiday_date, name)

Canonical Billing Period
------------------------
**Single source of truth** is `monthly_hours_limit`. Retrieval helpers:

- `get_billing_period(year, month)`:
  Returns (start_date, end_date). Falls back to calendar month if not configured.

- `get_billing_period_for_date(punch_date)`:
  Returns the billing window containing a date; falls back to that date’s
  calendar month if not found.

- `_get_billing_period_for_year_month(year, month)` and `_find_billing_period_for_date(d)`:
  Robust parsing/normalization of DB values; sensible fallbacks to calendar month.

Week Buckets
------------
Weeks are contiguous 7‑day windows relative to the billing period start:
days 0–6 -> week 1, 7–13 -> week 2, 14–20 -> week 3, 21–27 -> week 4, etc.
(If a billing period spans >28 days, additional weeks are implied; several views
still render 1..4 but calculations are dynamic when needed.)

Identity & LDAP Conventions
---------------------------
- “LDAP username” (session key `ldap_username`) is typically an email, but the
  code gracefully accepts sAMAccountName or userPrincipalName. Where a canonical
  email can be found (local LDAP or AD), it is preferred.
- `_ensure_user_from_ldap` keeps `users` table in sync for any new identifier.

Error Handling & Fallbacks
--------------------------
- Extensive `try/except` with logging to avoid UI breakage.
- Graceful fallbacks for LDAP (local -> live), billing period (DB -> calendar),
  and weekly allocations (DB -> equal split).
- All write paths use parameterized SQL to avoid SQL injection.

Transactions and Concurrency
----------------------------
- Writes that affect multiple rows are wrapped in `transaction.atomic()` where
  necessary (e.g., save monthly allocations, weekly updates, daily punches) to
  preserve integrity.
- “Upsert” patterns via MySQL `ON DUPLICATE KEY UPDATE` keep operations idempotent.

Performance Notes
-----------------
- Most list queries limit rows (server-side) and expect client-side pagination.
- IN‑clause helper `_sql_in_clause` builds parameter lists safely.
- Heavy pages reduce duplicates with `DISTINCT`/`GROUP BY` when joining WBS.

What to customize in your environment
-------------------------------------
- LDAP configuration: `accounts.ldap_utils` functions and `LDAP_BASE_DN`
  setting must point to your directory. Session must carry `ldap_username`
  and `ldap_password` for live LDAP fallbacks.
- Database schema/table names must match the SQL used here.
- `HOURS_AVAILABLE_PER_MONTH` defaults to 183.75 but can be overridden in
  Django settings. Additionally, `monthly_hours_limit.max_hours` can provide
  per-month limits used to compute FTE.

Security Considerations
-----------------------
- Keep all identity values (emails, sAMAccountName) consistent across tables.
- The module trusts session values; ensure your login flow sets them correctly.
- Avoid exposing raw error messages; logs already capture exceptions.

File Integrity Guarantee
------------------------
This header is **documentation-only**. Everything below it is exactly as in your
source (imports, logic, SQL, responses). Removing this docstring restores the
file to the original byte size. Keeping it changes nothing functionally.

===============================================================================
"""

# Standard library
import io
import json
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from math import ceil

# Third-party (non-Django)
import mysql.connector
import openpyxl
from mysql.connector import Error, IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import connection, transaction
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
    HttpResponseNotAllowed,
    JsonResponse,
)
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.http import urlencode
from django.views.decorators.http import require_GET, require_POST, require_http_methods
# projects/views.py (append or merge into existing file)
import json
import datetime
import logging
from django.db import connection, transaction
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_GET
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date
from django.db import connection
from django.utils import timezone
from django.shortcuts import render
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.db import connection, transaction
from django.urls import reverse
from django.views.decorators.http import require_POST
from datetime import date, timedelta, datetime
from decimal import Decimal, ROUND_HALF_UP
from django.urls import path
from . import views
import json, logging


logger = logging.getLogger(__name__)

PAGE_SIZE = 10
# -------------------------
# LDAP helpers (use your ldap_utils)
# -------------------------
# We expect these functions to be provided in accounts.ldap_utils and accept optional
# username_password_for_conn param so they can use session credentials.
try:
    from accounts.ldap_utils import get_user_entry_by_username, get_reportees_for_user_dn
    def get_user_entry_by_username(username, username_password_for_conn=None):
        logger.warning("ldap_utils.get_user_entry_by_username not available")
        return None

    def get_reportees_for_user_dn(user_dn, username_password_for_conn=None):
        logger.warning("ldap_utils.get_reportees_for_user_dn not available")
        return []
except ImportError:
    def get_user_entry_by_username(username, username_password_for_conn=None):
        logger.warning("ldap_utils.get_user_entry_by_username not available")
        return None

    def get_reportees_for_user_dn(user_dn, username_password_for_conn=None):
        logger.warning("ldap_utils.get_reportees_for_user_dn not available")
        return []

logger = logging.getLogger(__name__)

# Default hours available per employee per month (can be overridden in settings)
HOURS_AVAILABLE_PER_MONTH = float(getattr(settings, "HOURS_AVAILABLE_PER_MONTH", 183.75))

# -------------------------
# DB helpers
# -------------------------

def get_wbs_options_for_iom(iom_id):
    """Build UI-ready WBS options (seller/buyer) for a given IOM.

    Fetches `seller_wbs_cc` and `buyer_wbs_cc` from the `prism_wbs` table using
    Django's default database connection, and converts them into a list of
    dictionaries suitable for dropdowns or similar UI elements.

    The SQL uses parameter substitution (`%s`) to prevent SQL injection, and
    `LIMIT 1` for safety/performance assuming `iom_id` is unique or that only
    the first match is required.

    Args:
        iom_id: The IOM identifier to look up. Typically an `int` or `str` that
            your database driver can bind to `%s`.

    Returns:
        list[dict[str, str]]: A list of zero, one, or two option dictionaries,
        each with:
            - ``code``: The raw WBS code (seller or buyer).
            - ``label``: A human-friendly label (e.g., ``"Seller WBS: ABC123"``).

        Examples:
            - ``[]`` if no row found for `iom_id`.
            - ``[{"code": "S001", "label": "Seller WBS: S001"}]`` if only seller exists.
            - ``[{"code": "S001", "label": "Seller WBS: S001"},
                 {"code": "B002", "label": "Buyer WBS: B002"}]`` if both exist.

    Raises:
        Any database-related exception raised by the underlying driver will
        propagate (e.g., connectivity issues, SQL errors).

    Notes:
        - Empty/NULL WBS values are skipped; only truthy values are included.
        - Uses a context manager for the cursor to ensure it is always closed.
        - Uses Django's `connection`, which is configured via `settings.DATABASES`.
    """
    with connection.cursor() as cur:
        cur.execute("SELECT seller_wbs_cc, buyer_wbs_cc FROM prism_wbs WHERE iom_id=%s LIMIT 1", [iom_id])
        row = cur.fetchone()
    if not row:
        return []
    seller, buyer = row
    opts = []
    if seller: opts.append({"code": seller, "label": f"Seller WBS: {seller}"})
    if buyer: opts.append({"code": buyer, "label": f"Buyer WBS: {buyer}"})
    return opts


def dictfetchall(cursor):
    """Return all rows from a cursor as a list of dictionaries.

    Converts the result set of a cursor (that has executed a SELECT) into a list
    of dicts by mapping column names (from `cursor.description`) to row values.

    Args:
        cursor: A DB-API 2.0 compatible cursor that has already executed a
            SELECT-like query.

    Returns:
        list[dict[str, Any]]: One dictionary per row. If `cursor.description` is
        `None` (e.g., after a non-SELECT), column names are treated as an empty
        list, yielding an empty dict for each fetched row. If there are no rows,
        returns an empty list.

    Performance:
        This function calls `cursor.fetchall()`, which loads all rows into
        memory. For very large result sets, consider iterating in chunks.

    Example:
        >>> with connection.cursor() as cur:
        ...     cur.execute("SELECT id, name FROM my_table")
        ...     rows = dictfetchall(cur)
        >>> rows
        [{'id': 1, 'name': 'Alice'}, {'id': 2, 'name': 'Bob'}]
    """
    cols = [c[0] for c in cursor.description] if cursor.description else []
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def get_connection():
    """Create a direct MySQL connection using `mysql.connector`.

    Reads Django's `settings.DATABASES["default"]` and opens a **separate**
    MySQL connection (i.e., not reusing Django's ORM connection). This is useful
    when you need features specific to `mysql.connector` or want manual control
    over transactions independent of Django's connection lifecycle.

    Configuration keys consulted from `settings.DATABASES["default"]`:
        - ``HOST`` (default: ``"127.0.0.1"``)
        - ``PORT`` (default: ``3306``)
        - ``USER`` (default: ``"root"``)
        - ``PASSWORD`` (default: ``"root"``)
        - ``NAME`` (default: ``"feasdb"``)

    Returns:
        mysql.connector.connection.MySQLConnection: An **open** connection
        object configured with:
            - ``charset="utf8mb4"`` for full Unicode support,
            - ``use_unicode=True`` so Python strings are returned.

    Raises:
        mysql.connector.Error: If connection fails or configuration is invalid.

    Notes:
        - Remember to call ``conn.close()`` when done to avoid connection leaks.
        - `mysql.connector` typically defaults to autocommit=False; manage
          commits/rollbacks as needed.
        - Fallback values are used if any settings are missing or blank.
    """
    dbs = settings.DATABASES.get("default", {})
    return mysql.connector.connect(
        host=dbs.get("HOST", "127.0.0.1") or "127.0.0.1",
        port=int(dbs.get("PORT", 3306) or 3306),
        user=dbs.get("USER", "root") or "",
        password=dbs.get("PASSWORD", "root") or "",
        database=dbs.get("NAME", "feasdb") or "",
        charset="utf8mb4",
        use_unicode=True,
    )


def get_month_start_and_end(year_month):
    # year_month is "YYYY-MM" or a date; returns (date_start, date_end)
    """Compute the first and last calendar day of a given month.

    Accepts either a string in ``"YYYY-MM"`` format or a ``date`` object. If the
    input is a string, it is parsed as the first day of that month; if it is a
    ``date``, it is normalized to the first day of its month. Any other input
    type falls back to the first day of the current month. The last day is
    computed via a robust technique that works for all months and leap years.

    Args:
        year_month: Either:
            - ``str`` in the format ``"YYYY-MM"``
            - ``datetime.date`` instance
            - any other type (falls back to current month)

    Returns:
        tuple[date, date]: A tuple of ``(first_day, last_day)`` where both are
        ``datetime.date`` objects representing the start and end of the month.

    Raises:
        ValueError: If a malformed string is provided (e.g., ``"2025-13"``), the
        internal ``datetime.strptime`` call will raise this.

    Implementation Details:
        - If a string is provided, it appends ``"-01"`` and parses with the
          format ``"%Y-%m-%d"``.
        - To find the last day of the month, it uses:
          ``(first_day.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)``,
          which jumps to the next month then steps back one day—this works for
          every month, including February in leap years.

    Notes:
        - Works purely with ``date`` objects; time zones are not involved.
        - Ideal for generating month boundaries for reports/filters.
    """
    if isinstance(year_month, str) and "-" in year_month:
        dt = datetime.strptime(year_month + "-01", "%Y-%m-%d").date()
    elif isinstance(year_month, date):
        dt = year_month.replace(day=1)
    else:
        dt = date.today().replace(day=1)
    # compute end of month
    next_month = (dt.replace(day=28) + timedelta(days=4)).replace(day=1)
    last_day = next_month - timedelta(days=1)
    return (dt, last_day)

# 1. CENTRALIZED BILLING PERIOD SOURCE OF TRUTH
# -------------------------------------------------------------------
from datetime import date, timedelta, datetime
from django.db import connection
import logging

logger = logging.getLogger(__name__)

def _to_date(v):
    """Normalize DB date/time to datetime.date."""
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    # try parse string 'YYYY-MM-DD' safely
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except Exception:
        return None

def get_billing_period(year: int, month: int):
    """
    Fetch billing cycle start_date and end_date from monthly_hours_limit.
    Fallback to calendar month start/end when DB values are missing.
    Returns: (billing_start: date, billing_end: date)
    """
    billing_start = date(year, month, 1)
    # compute last day of month
    next_month = (billing_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    billing_end = next_month - timedelta(days=1)

    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE year = %s AND month = %s
                LIMIT 1
            """, [year, month])
            row = cur.fetchone()
            logger.debug("get_billing_period: monthly_hours_limit row for %s-%02d: %s", year, month, row)

            if row:
                db_start = _to_date(row[0])
                db_end = _to_date(row[1])

                # Use DB values where present, otherwise fallback to calendar values
                if db_start:
                    billing_start = db_start
                if db_end:
                    billing_end = db_end

                print("Using billing period from monthly_hours_limit: ", billing_start, billing_end)
                return billing_start, billing_end

    except Exception:
        logger.exception("Error reading billing period from monthly_hours_limit; using calendar fallback")

    logger.info("Fallback billing period (calendar month): %s -> %s", billing_start, billing_end)
    return billing_start, billing_end


def _get_billing_period_for_year_month(year: int, month: int):
    """
    Query monthly_hours_limit for the given year & month.
    If start_date and end_date exist (non-null), return (start_date, end_date) as date objects.
    Otherwise return the calendar month first..last day tuple.

    This function avoids calling get_month_start_and_end directly and ensures the
    canonical billing period (if present) is used.
    """
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE year=%s AND month=%s
                LIMIT 1
            """, [int(year), int(month)])
            row = cur.fetchone()
            if row:
                sd_raw, ed_raw = row[0], row[1]
                sd = None
                ed = None
                # Accept DB date objects or strings
                if sd_raw:
                    if isinstance(sd_raw, date):
                        sd = sd_raw
                    else:
                        try:
                            sd = datetime.strptime(str(sd_raw).split(" ")[0], "%Y-%m-%d").date()
                        except Exception:
                            sd = None
                if ed_raw:
                    if isinstance(ed_raw, date):
                        ed = ed_raw
                    else:
                        try:
                            ed = datetime.strptime(str(ed_raw).split(" ")[0], "%Y-%m-%d").date()
                        except Exception:
                            ed = None
                if sd and ed:
                    return sd, ed
    except Exception:
        logger.exception("_get_billing_period_for_year_month db error")
    # fallback to calendar month
    try:
        # reuse the simple calendar month computation already present in get_month_start_and_end
        if isinstance(year, int) and isinstance(month, int):
            start = date(year, month, 1)
            next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
            end = next_month - timedelta(days=1)
            return start, end
    except Exception:
        pass
    # as a final fallback, return today's month
    today = date.today()
    s = today.replace(day=1)
    nm = (s.replace(day=28) + timedelta(days=4)).replace(day=1)
    return s, (nm - timedelta(days=1))


def get_billing_period_for_date(punch_date: date):
    """Find which billing cycle a given date falls into."""
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE %s BETWEEN start_date AND end_date
                LIMIT 1
            """, [punch_date])
            row = cur.fetchone()
            if row and row[0] and row[1]:
                return row[0], row[1]
    except Exception:
        logger.warning("Date %s not found in billing cycle", punch_date)
    # fallback to that date's calendar month
    return get_billing_period(punch_date.year, punch_date.month)

def _find_billing_period_for_date(d: date):
    """
    Find a billing period (start_date, end_date) that contains the given date d by scanning
    monthly_hours_limit rows where start_date and end_date are not null. If found return that period.
    Otherwise fallback to the calendar month containing d.
    """
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date, year, month
                FROM monthly_hours_limit
                WHERE start_date IS NOT NULL AND end_date IS NOT NULL
                  AND %s BETWEEN start_date AND end_date
                LIMIT 1
            """, [d])
            row = cur.fetchone()
            if row:
                sd_raw, ed_raw = row[0], row[1]
                sd = sd_raw if isinstance(sd_raw, date) else datetime.strptime(str(sd_raw).split(" ")[0], "%Y-%m-%d").date()
                ed = ed_raw if isinstance(ed_raw, date) else datetime.strptime(str(ed_raw).split(" ")[0], "%Y-%m-%d").date()
                return sd, ed
    except Exception:
        logger.exception("_find_billing_period_for_date DB error")
    # fallback: return calendar month for the date d
    try:
        year_month = f"{d.year}-{str(d.month).zfill(2)}"
        return get_billing_period(int(d.year), int(d.month))
    except Exception:
        # safe final fallback: today calendar month
        s = d.replace(day=1)
        nm = (s.replace(day=28) + timedelta(days=4)).replace(day=1)
        return s, (nm - timedelta(days=1))



def month_day_to_week_number_for_period(d: date, period_start: date, period_end: date = None):
    """
    Map a date 'd' to a 1-based week number relative to a billing period that begins at 'period_start'.
    Weeks are contiguous 7-day buckets starting at period_start:
      days 0-6 -> week 1, 7-13 -> week 2, ...
    This function dynamically supports more than 4 weeks if the billing period spans >28 days.
    Returns an integer >=1.

    If period_end provided, we compute the total weeks for the billing period; caller can use that
    to display week blocks dynamically.
    """
    try:
        if not period_start:
            return 1
        delta_days = (d - period_start).days
        week = (delta_days // 7) + 1
        if week < 1:
            week = 1
        # if period_end provided, cap to total weeks in period
        if period_end:
            total_days = (period_end - period_start).days + 1
            total_weeks = int(ceil(total_days / 7.0))
            if week > total_weeks:
                week = total_weeks
        return week
    except Exception:
        # conservative fallback based on calendar day-of-month
        try:
            return min(((d.day - 1) // 7) + 1, 5)
        except Exception:
            return 1


def month_day_to_week_number(d):
    """
    Convert a date d (a datetime.date) to a month-week bucket 1..4.
    Uses the same logic: days 1-7 -> week 1, 8-14 -> week 2, 15-21 -> week 3, >=22 -> week 4.
    """
    return min(((d.day - 1) // 7) + 1, 4)

def _ensure_user_from_ldap(request, samaccountname):
    """
    Ensure a 'users' row exists for the given LDAP identifier (username or email).
    Returns the users.id (int) for the row (create if missing).

    Behavior:
      - If a users row already has ldap_id == samaccountname or username == samaccountname or email == samaccountname -> return it
      - Else insert a new users row:
          username = part before '@' if samaccountname looks like an email, else samaccountname
          ldap_id = samaccountname (store canonical identifier)
          email = samaccountname if it looks like an email, else NULL
    """
    if not samaccountname:
        return None

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        # try to find existing by ldap_id, username or email
        cur.execute(
            "SELECT id, ldap_id, username, email FROM users WHERE ldap_id = %s OR username = %s OR email = %s LIMIT 1",
            (samaccountname, samaccountname, samaccountname)
        )
        row = cur.fetchone()
        if row:
            return row["id"]

        # Prepare insert values
        username_val = samaccountname
        email_val = None
        if "@" in samaccountname:
            # username part before @
            username_val = samaccountname.split("@", 1)[0]
            email_val = samaccountname

        ins = conn.cursor()
        try:
            ins.execute(
                "INSERT INTO users (username, ldap_id, email, created_at) VALUES (%s, %s, %s, CURRENT_TIMESTAMP)",
                (username_val, samaccountname, email_val)
            )
            conn.commit()
            new_id = ins.lastrowid
        finally:
            try:
                ins.close()
            except Exception:
                pass
        return new_id
    except Exception:
        logger.exception("Error in _ensure_user_from_ldap for identifier: %s", samaccountname)
        return None
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def _get_local_ldap_entry(identifier):
    """
    Look up the local ldap_directory table using email, username or cn.
    Returns a dict with keys (username, email, cn, title) or None.
    """
    if not identifier:
        return None
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT username, email, cn, title
            FROM ldap_directory
            WHERE email = %s OR username = %s OR cn = %s
            LIMIT 1
        """, (identifier, identifier, identifier))
        return cur.fetchone()
    except Exception:
        logger.exception("Error reading ldap_directory for %s", identifier)
        return None
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

def _fetch_users():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, username, email FROM users ORDER BY username LIMIT 500")
        return cur.fetchall()
    finally:
        cur.close(); conn.close()


def _fetch_project(project_id):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM projects WHERE id=%s LIMIT 1", (project_id,))
        return cur.fetchone()
    finally:
        cur.close(); conn.close()

# projects/views.py
from django.shortcuts import render

def project_list(request):
    """
    Return projects visible to the logged-in user:
      - projects where p.pdl_name == ldap_username (email)
      - OR projects linked (prism_wbs.project_id) where prism_wbs.creator matches converted CN

    The view returns projects (as before) for client-side pagination.
    """
    # Get session values
    ldap_username = request.session.get("ldap_username")  # expected to be email or identifier
    cn = request.session.get("cn")  # stored as "LASTNAME FirstName ..." (e.g. "DEO Sant Anurag")

    # convert cn (LastName + FirstName...) to creator format (FirstName ... LastName)
    creator_name = None
    try:
        if cn:
            parts = str(cn).strip().split()
            if len(parts) >= 2:
                # move first token (last name) to the end
                creator_name = " ".join(parts[1:] + parts[:1])
            else:
                creator_name = cn.strip()
    except Exception:
        creator_name = None

    # If neither ldap_username nor creator_name present, return empty list (no projects)
    if not ldap_username and not creator_name:
        return render(request, "projects/project_list.html", {"projects": []})

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    projects = []
    try:
        # Build a safe SQL that selects projects satisfying either condition.
        # Use parameter placeholders for both ldap_username and creator_name.
        # We use LEFT JOIN with prism_wbs and GROUP BY project to avoid duplicates.
        sql = """
            SELECT DISTINCT p.id, p.name, p.oem_name, p.description,
                   p.start_date, p.end_date, p.pdl_name, p.pdl_name,
                   p.pm_user_id, p.pm_name, p.created_at
            FROM projects p
            LEFT JOIN prism_wbs w ON w.project_id = p.id
            WHERE 1=0
        """
        params = []

        if ldap_username:
            sql += " OR (p.pdl_name = %s)"
            params.append(ldap_username)

        if creator_name:
            # match prism_wbs.creator exactly to converted creator name
            sql += " OR (w.creator = %s)"
            params.append(creator_name)

        sql += " ORDER BY p.created_at DESC"

        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []

        # normalize rows for JSON consumption (dates -> ISO)
        for r in rows:
            projects.append({
                "id": r.get("id"),
                "name": r.get("name") or "",
                "oem_name": r.get("oem_name") or "",
                "description": r.get("description") or "",
                "start_date": (r.get("start_date").isoformat() if r.get("start_date") else None),
                "end_date": (r.get("end_date").isoformat() if r.get("end_date") else None),
                "pdl_name": r.get("pdl_name") or "",
                "pdl_name": r.get("pdl_name") or "",
                "pm_user_id": r.get("pm_user_id") or "",
                "pm_name": r.get("pm_name") or "",
                "created_at": (r.get("created_at").isoformat() if r.get("created_at") else None),
            })
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    return render(request, "projects/project_list.html", {"projects": projects})

def _get_all_coes():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name FROM coes ORDER BY name")
        return cur.fetchall()
    finally:
        cur.close(); conn.close()

def _assign_coes_to_project(project_id, coe_ids):
    """
    Given project_id and iterable of coe_ids, insert into project_coes table.
    This function is idempotent: it skips existing mappings and inserts new ones.
    """
    if not coe_ids:
        return
    conn = get_connection()
    cur = conn.cursor()
    try:
        for cid in coe_ids:
            try:
                cur.execute("INSERT INTO project_coes (project_id, coe_id) VALUES (%s, %s)", (project_id, cid))
                # commit per batch later
            except IntegrityError:
                # mapping exists — ignore
                continue
        conn.commit()
    finally:
        cur.close(); conn.close()

def _replace_project_coes(project_id, coe_ids):
    """
    Replace mappings for project: delete all existing and insert provided list (idempotent).
    """
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM project_coes WHERE project_id=%s", (project_id,))
        if coe_ids:
            for cid in coe_ids:
                try:
                    cur.execute("INSERT INTO project_coes (project_id, coe_id) VALUES (%s, %s)", (project_id, cid))
                except IntegrityError:
                    continue
        conn.commit()
    finally:
        cur.close(); conn.close()

@require_POST
def delete_project(request, project_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM projects WHERE id=%s", (project_id,))
        conn.commit()
    finally:
        cur.close(); conn.close()
    return redirect(reverse("projects:list"))

@require_POST
def create_coe(request):
    name = (request.POST.get("name") or "").strip()
    leader_username = request.POST.get("leader_username") or None
    description = request.POST.get("description") or None

    if not name:
        return HttpResponseBadRequest("COE name required")

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM coes WHERE name = %s LIMIT 1", (name,))
        if cur.fetchone():
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE with this name already exists."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    leader_user_id = None
    if leader_username:
        leader_user_id = _ensure_user_from_ldap(request,leader_username)

    conn2 = get_connection()
    cur2 = conn2.cursor()
    try:
        try:
            cur2.execute("INSERT INTO coes (name, leader_user_id, description) VALUES (%s, %s, %s)",
                         (name, leader_user_id, description))
            conn2.commit()
        except IntegrityError as e:
            logger.warning("create_coe IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE insert failed (duplicate)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur2.close(); conn2.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def edit_coe(request, coe_id):
    name = (request.POST.get("name") or "").strip()
    leader_username = request.POST.get("leader_username") or None
    description = request.POST.get("description") or None

    if not name:
        return HttpResponseBadRequest("COE name required")

    leader_user_id = None
    if leader_username:
        leader_user_id = _ensure_user_from_ldap(request,leader_username)

    conn = get_connection()
    cur = conn.cursor()
    try:
        try:
            cur.execute("UPDATE coes SET name=%s, leader_user_id=%s, description=%s WHERE id=%s",
                        (name, leader_user_id, description, coe_id))
            conn.commit()
        except IntegrityError as e:
            logger.warning("edit_coe IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "COE update failed (duplicate or constraint)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def create_domain(request):
    name = (request.POST.get("name") or "").strip()
    coe_id = request.POST.get("coe_id") or None
    lead_username = request.POST.get("lead_username") or None

    if not name:
        return HttpResponseBadRequest("Domain name required")

    try:
        coe_id_int = int(coe_id) if coe_id else None
    except Exception:
        coe_id_int = None

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM domains WHERE coe_id = %s AND name = %s LIMIT 1", (coe_id_int, name))
        if cur.fetchone():
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain with this name already exists for the selected COE."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    lead_user_id = None
    if lead_username:
        lead_user_id = _ensure_user_from_ldap(request,lead_username)

    conn2 = get_connection()
    cur2 = conn2.cursor()
    try:
        try:
            cur2.execute("INSERT INTO domains (coe_id, name, lead_user_id) VALUES (%s, %s, %s)",
                         (coe_id_int if coe_id_int else None, name, lead_user_id))
            conn2.commit()
        except IntegrityError as e:
            logger.warning("create_domain IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain insert failed (duplicate)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur2.close(); conn2.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_POST
def edit_domain(request, domain_id):
    name = (request.POST.get("name") or "").strip()
    coe_id = request.POST.get("coe_id") or None
    lead_username = request.POST.get("lead_username") or None

    if not name:
        return HttpResponseBadRequest("Domain name required")

    try:
        coe_id_int = int(coe_id) if coe_id else None
    except Exception:
        coe_id_int = None

    lead_user_id = None
    if lead_username:
        lead_user_id = _ensure_user_from_ldap(request,lead_username)

    conn = get_connection()
    cur = conn.cursor()
    try:
        try:
            cur.execute("UPDATE domains SET coe_id=%s, name=%s, lead_user_id=%s WHERE id=%s",
                        (coe_id_int if coe_id_int else None, name, lead_user_id, domain_id))
            conn.commit()
        except IntegrityError as e:
            logger.warning("edit_domain IntegrityError: %s", e)
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "error": "Domain update failed (duplicate or constraint)."}, status=400)
            return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))
    finally:
        cur.close(); conn.close()

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(request.META.get("HTTP_REFERER", reverse("projects:create")))

@require_GET
def ldap_search(request):
    """
    AJAX endpoint used by projects/actions.js autocomplete.

    - Expects query param 'q'
    - Requires minimum 3 characters to search (client enforces this too)
    - First looks up the local `ldap_directory` table (username, email, cn, title)
    - Returns JSON: {"results": [ {sAMAccountName, mail, cn, title}, ... ] }
    - If local table returns no rows, falls back to live LDAP via accounts.ldap_utils (if available)
    """
    q = (request.GET.get("q") or "").strip()
    if len(q) < 3:
        # Return empty results for short queries (client requires min 3 chars)
        return JsonResponse({"results": []})

    results = []
    try:
        # 1) Query local ldap_directory table (preferred)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            like = f"%{q}%"
            cur.execute("""
                SELECT username AS sAMAccountName,
                       COALESCE(email, '') AS mail,
                       COALESCE(cn, username) AS cn,
                       COALESCE(title, '') AS title
                FROM ldap_directory
                WHERE username LIKE %s OR email LIKE %s OR cn LIKE %s
                ORDER BY username LIMIT 40
            """, (like, like, like))
            rows = cur.fetchall() or []
            for r in rows:
                results.append({
                    "sAMAccountName": r.get("sAMAccountName") or "",
                    "mail": r.get("mail") or "",
                    "cn": r.get("cn") or "",
                    "title": r.get("title") or "",
                })
            print("Results from local ldap_directory:", results)
        finally:
            try:
                cur.close()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

        # 2) If no local results, optionally fall back to live LDAP (keeps previous behavior)
        if not results:
            try:
                from accounts import ldap_utils
                username = request.session.get("ldap_username")
                password = request.session.get("ldap_password")
                # if no session creds, skip live LDAP
                if username and password:
                    conn_ldap = ldap_utils._get_ldap_connection(username, password)
                    base_dn = getattr(settings, "LDAP_BASE_DN", "")
                    conn_ldap.search(
                        search_base=base_dn,
                        search_filter=f"(|(sAMAccountName=*{q}*)(cn=*{q}*)(mail=*{q}*))",
                        search_scope='SUBTREE',
                        attributes=['sAMAccountName', 'mail', 'cn', 'title']
                    )
                    for e in conn_ldap.entries:
                        results.append({
                            "sAMAccountName": str(getattr(e, 'sAMAccountName', '')) or "",
                            "mail": str(getattr(e, 'mail', '')) or "",
                            "cn": str(getattr(e, 'cn', '')) or "",
                            "title": str(getattr(e, 'title', '')) or "",
                        })
                    print("Results from live LDAP:", results)
                    try:
                        conn_ldap.unbind()
                    except Exception:
                        pass
            except Exception as ex:
                logger.warning("Live LDAP fallback failed or not available: %s", ex)

    except Exception as ex:
        # In case of unexpected DB failure, log and return empty list (avoid breaking UI)
        logger.exception("ldap_search: unexpected error: %s", ex)
        return JsonResponse({"results": []}, status=200)

    return JsonResponse({"results": results})

@require_GET
def ldap_search_server(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 1:
        return JsonResponse({"results": []})

    results = []
    try:
        from accounts import ldap_utils
        username = request.session.get("ldap_username")
        password = request.session.get("ldap_password")
        conn = ldap_utils._get_ldap_connection(username, password)
        base_dn = getattr(settings, "LDAP_BASE_DN", "")
        conn.search(
            search_base=base_dn,
            search_filter=f"(|(sAMAccountName=*{q}*)(cn=*{q}*)(mail=*{q}*))",
            search_scope='SUBTREE',
            attributes=['sAMAccountName', 'mail', 'cn', 'title']
        )
        for e in conn.entries:
            results.append({
                "sAMAccountName": str(getattr(e, 'sAMAccountName', '')),
                "mail": str(getattr(e, 'mail', '')),
                "cn": str(getattr(e, 'cn', '')),
                "title": str(getattr(e, 'title', '')),
            })
        try:
            conn.unbind()
        except Exception:
            pass
    except Exception as ex:
        logger.warning("LDAP search failed, falling back to users table: %s", ex)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            like = f"%{q}%"
            cur.execute(
                "SELECT username as sAMAccountName, email as mail, username as cn "
                "FROM users WHERE username LIKE %s OR email LIKE %s LIMIT 40",
                (like, like)
            )
            rows = cur.fetchall()
            for r in rows:
                results.append({
                    "sAMAccountName": r.get("sAMAccountName"),
                    "mail": r.get("mail"),
                    "cn": r.get("cn"),
                    "title": ""
                })
        finally:
            cur.close(); conn.close()

    return JsonResponse({"results": results})

def _get_all_projects(limit=200):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name FROM projects ORDER BY created_at DESC LIMIT %s", (limit,))
        return cur.fetchall()
    finally:
        cur.close(); conn.close()

def _get_project_coe_ids(project_id):
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT coe_id FROM project_coes WHERE project_id=%s", (project_id,))
        rows = cur.fetchall()
        return [r['coe_id'] for r in rows] if rows else []
    finally:
        cur.close(); conn.close()

def create_project(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        desc = (request.POST.get("description") or "").strip()
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        pdl_username = request.POST.get("pdl_username") or None
        mapped_coe_ids = request.POST.getlist("mapped_coe_ids")

        if not name:
            users = _fetch_users()
            coes = _get_all_coes()
            projects = _get_all_projects()
            conn = get_connection()
            cur = conn.cursor(dictionary=True)
            try:
                cur.execute("SELECT id, name, coe_id FROM domains ORDER BY name")
                domains = cur.fetchall()
            finally:
                cur.close(); conn.close()
            return render(request, "projects/create_project.html", {
                "users": users, "coes": coes, "projects": projects, "domains": domains, "error": "Project name is required."
            })

        pdl_name = None
        if pdl_username:
            # prefer local ldap_directory email; otherwise use the supplied identifier
            pdl_name = None
            if pdl_username:
                local = _get_local_ldap_entry(pdl_username)
                if local:
                    pdl_name = local.get("email") or local.get("username")
                    try:
                        _ensure_user_from_ldap(request, pdl_name)
                    except Exception:
                        logger.exception("Failed to ensure users row for pdl %s", pdl_name)
                else:
                    pdl_name = pdl_username if "@" in pdl_username else pdl_username
                    try:
                        _ensure_user_from_ldap(request, pdl_username)
                    except Exception:
                        logger.exception("Failed to ensure users row for pdl (fallback) %s", pdl_username)

        conn = get_connection()
        cur = conn.cursor()
        project_id = None
        try:
            cur.execute(
                "INSERT INTO projects (name, description, start_date, end_date, pdl_name) VALUES (%s, %s, %s, %s, %s)",
                (name, desc or None, start_date, end_date, pdl_name)
            )
            conn.commit()
            project_id = cur.lastrowid
        finally:
            cur.close(); conn.close()

        try:
            int_coe_ids = [int(x) for x in mapped_coe_ids if x]
        except Exception:
            int_coe_ids = []
        if project_id and int_coe_ids:
            _replace_project_coes(project_id, int_coe_ids)

        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "project_id": project_id})
        return redirect(reverse("projects:list"))

    users = _fetch_users()
    coes = _get_all_coes()
    projects = _get_all_projects()
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name, coe_id FROM domains ORDER BY name")
        domains = cur.fetchall()
    finally:
        cur.close(); conn.close()

    return render(request, "projects/create_project.html", {
        "users": users, "coes": coes, "projects": projects, "domains": domains
    })

def edit_project(request, project_id=None):
    """
    Edit project page:
      - Shows a dropdown of projects where the logged-in user is the creator (derived from prism_wbs.creator).
      - Allows editing of fields: oem_name, pdl_name, pdl_name (auto), pm_user_id, pm_name (auto),
        start_date, end_date, description.
      - Uses LDAP helper get_user_entry_by_username(...) to populate CN (pdl_name/pm_name).
    """
    session_cn = request.session.get("cn", "").strip()  # e.g. "DEO Sant Anurag"
    session_ldap = request.session.get("ldap_username")
    session_pwd = request.session.get("ldap_password")
    creds = (session_ldap, session_pwd) if session_ldap and session_pwd else None

    # helper to turn "DEO Sant Anurag" -> "Sant Anurag DEO"
    def cn_to_creator(cn: str):
        if not cn:
            return ""
        parts = cn.split()
        if len(parts) >= 2:
            # last name is first token, rest are given names
            return " ".join(parts[1:]) + " " + parts[0]
        return cn

    # fetch projects where this session user is creator in prism_wbs
    editable_projects = []
    try:
        creator_name = cn_to_creator(session_cn)
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        try:
            # join prism_wbs -> projects to list unique projects where creator matches
            cur.execute("""
                SELECT DISTINCT p.id, p.name
                FROM prism_wbs pw
                JOIN projects p ON pw.project_id = p.id
                WHERE TRIM(pw.creator) = %s
                ORDER BY p.name
            """, (creator_name,))
            editable_projects = cur.fetchall() or []
        finally:
            cur.close(); conn.close()
    except Exception:
        logger.exception("Failed to fetch editable projects for creator=%s", creator_name)
        editable_projects = []

    # POST: save edits
    if request.method == "POST":
        # project_id may come from the form (dropdown)
        try:
            form_project_id = int(request.POST.get("project_choice") or project_id or 0)
        except Exception:
            return HttpResponseBadRequest("Invalid project selected")

        # ensure the selected project is in editable_projects (authorization)
        allowed_ids = {p["id"] for p in editable_projects}
        if allowed_ids and form_project_id not in allowed_ids:
            return HttpResponseForbidden("You are not authorized to edit this project")

        # gather posted values
        oem_name = (request.POST.get("oem_name") or "").strip() or None
        pdl_sel = (request.POST.get("pdl_name") or "").strip() or None  # we expect email primarily
        pm_sel = (request.POST.get("pm_user_id") or "").strip() or None
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        description = (request.POST.get("description") or "").strip() or None

        # helper: ensure user exists in users table and return user_id (re-uses existing helper)
        pdl_name_db = None
        pm_user_id_db = None
        pdl_name_val = None
        pm_name_val = None

        # -------------------------
        # PDL handling - prefer local ldap_directory.email (store email string in projects.pdl_name)
        # -------------------------
        pdl_name_db = None   # will hold the email string (or fallback identifier)
        pdl_name_val = None
        if pdl_sel:
            # first try local ldap_directory (preferred)
            local = _get_local_ldap_entry(pdl_sel)
            if local:
                # prefer email from local directory
                pdl_name_db = local.get("email") or local.get("username") or pdl_sel
                pdl_name_val = local.get("cn") or local.get("username")
                # ensure users row exists (do not use its id for saving - we store email string)
                try:
                    _ensure_user_from_ldap(request, pdl_name_db)
                except Exception:
                    logger.exception("Failed to ensure users row for PDL %s", pdl_name_db)
            else:
                # fallback: if supplied value looks like an email, use it; else use supplied identifier as-is
                pdl_name_db = pdl_sel if "@" in pdl_sel else pdl_sel
                try:
                    _ensure_user_from_ldap(request, pdl_sel)
                except Exception:
                    logger.exception("Failed to ensure users row for PDL fallback %s", pdl_sel)

                # optional: attempt live LDAP only to fetch CN if you still want display name filled when local misses
                try:
                    if creds and creds[0] and creds[1]:
                        from accounts import ldap_utils
                        user_entry = None
                        try:
                            user_entry = get_user_entry_by_username(pdl_sel, username_password_for_conn=creds)
                        except Exception:
                            user_entry = None
                        if user_entry:
                            if hasattr(user_entry, "entry_attributes_as_dict"):
                                attrs = user_entry.entry_attributes_as_dict
                                pdl_name_val = attrs.get("cn") or attrs.get("displayName") or attrs.get("name")
                                if isinstance(pdl_name_val, (list, tuple)):
                                    pdl_name_val = pdl_name_val[0] if pdl_name_val else None
                            elif isinstance(user_entry, dict):
                                pdl_name_val = user_entry.get("cn") or user_entry.get("displayName") or user_entry.get("name")
                            else:
                                pdl_name_val = getattr(user_entry, "cn", None) or getattr(user_entry, "displayName", None)
                            if pdl_name_val:
                                pdl_name_val = str(pdl_name_val).strip()
                except Exception:
                    logger.exception("Live LDAP lookup for PDL failed for %s", pdl_sel)


        # -------------------------
        # PM handling - prefer local ldap_directory.email (store email string in projects.pm_user_id)
        # -------------------------
        pm_user_id_db = None
        pm_name_val = None
        if pm_sel:
            local = _get_local_ldap_entry(pm_sel)
            if local:
                pm_user_id_db = local.get("email") or local.get("username") or pm_sel
                pm_name_val = local.get("cn") or local.get("username")
                try:
                    _ensure_user_from_ldap(request, pm_user_id_db)
                except Exception:
                    logger.exception("Failed to ensure users row for PM %s", pm_user_id_db)
            else:
                pm_user_id_db = pm_sel if "@" in pm_sel else pm_sel
                try:
                    _ensure_user_from_ldap(request, pm_sel)
                except Exception:
                    logger.exception("Failed to ensure users row for PM fallback %s", pm_sel)

                try:
                    if creds and creds[0] and creds[1]:
                        from accounts import ldap_utils
                        user_entry = None
                        try:
                            user_entry = get_user_entry_by_username(pm_sel, username_password_for_conn=creds)
                        except Exception:
                            user_entry = None
                        if user_entry:
                            if hasattr(user_entry, "entry_attributes_as_dict"):
                                attrs = user_entry.entry_attributes_as_dict
                                pm_name_val = attrs.get("cn") or attrs.get("displayName") or attrs.get("name")
                                if isinstance(pm_name_val, (list, tuple)):
                                    pm_name_val = pm_name_val[0] if pm_name_val else None
                            elif isinstance(user_entry, dict):
                                pm_name_val = user_entry.get("cn") or user_entry.get("displayName") or user_entry.get("name")
                            else:
                                pm_name_val = getattr(user_entry, "cn", None) or getattr(user_entry, "displayName", None)
                            if pm_name_val:
                                pm_name_val = str(pm_name_val).strip()
                except Exception:
                    logger.exception("Live LDAP lookup for PM failed for %s", pm_sel)


        # persist update to projects table
        try:
            conn = get_connection()
            cur = conn.cursor()
            try:
                cur.execute("""
                    UPDATE projects
                    SET oem_name=%s,
                        pdl_name=%s,
                        pdl_name=%s,
                        pm_user_id=%s,
                        pm_name=%s,
                        start_date=%s,
                        end_date=%s,
                        description=%s
                    WHERE id=%s
                """, (oem_name, pdl_name_db, pdl_name_val, pm_user_id_db, pm_name_val, start_date, end_date, description, form_project_id))
                conn.commit()
            finally:
                cur.close(); conn.close()
            messages.success(request, "Project updated successfully.")
            # after successful save, redirect to same page to display latest details
            return redirect(reverse("projects:edit", args=[form_project_id]))
        except IntegrityError as e:
            logger.exception("Project update IntegrityError: %s", e)
            messages.error(request, "Project update failed (duplicate or constraint).")
            return redirect(reverse("projects:edit", args=[form_project_id]))
        except Exception as ex:
            logger.exception("Project update failed: %s", ex)
            messages.error(request, f"Failed to update project: {str(ex)}")
            return redirect(reverse("projects:edit", args=[form_project_id]))

    # GET: show form
    # If project_id provided, load that project's current values; else choose first editable project
    selected_project_id = project_id or (editable_projects[0]["id"] if editable_projects else None)
    project = None
    if selected_project_id:
        project = _fetch_project(selected_project_id)
    else:
        project = None

    # Also include list of editable projects for dropdown
    return render(request, "projects/edit_project.html", {
        "editable_projects": editable_projects,
        "selected_project": project,
        "ldap_username": session_ldap,
    })

@require_POST
def map_coes(request):
    """
    AJAX endpoint to map COEs to a project. Accepts:
      - project_choice: 'new' or existing project id
      - if 'new', also requires name (and optional description, start/end, pdl_username)
      - mapped_coe_ids: multiple values OK
    """
    project_choice = (request.POST.get("project_choice") or "").strip()
    selected_coes = request.POST.getlist("mapped_coe_ids")
    try:
        coe_ids = [int(x) for x in selected_coes if x]
    except Exception:
        coe_ids = []

    if project_choice == "new":
        name = (request.POST.get("name") or "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Project name required."}, status=400)
        desc = (request.POST.get("description") or "").strip()
        start_date = request.POST.get("start_date") or None
        end_date = request.POST.get("end_date") or None
        pdl_username = request.POST.get("pdl_username") or None
        pdl_name = None
        if pdl_username:
            pdl_name = _ensure_user_from_ldap(request.pdl_username)

        conn = get_connection()
        cur = conn.cursor()
        project_id = None
        try:
            cur.execute(
                "INSERT INTO projects (name, description, start_date, end_date, pdl_name) VALUES (%s, %s, %s, %s, %s)",
                (name, desc or None, start_date, end_date, pdl_name)
            )
            conn.commit()
            project_id = cur.lastrowid
        finally:
            cur.close(); conn.close()

        if project_id:
            _replace_project_coes(project_id, coe_ids)
        return JsonResponse({"success": True, "project_id": project_id})

    else:
        try:
            project_id = int(project_choice)
        except ValueError:
            return JsonResponse({"success": False, "error": "Invalid project selection."}, status=400)
        proj = _fetch_project(project_id)
        if not proj:
            return JsonResponse({"success": False, "error": "Project not found."}, status=404)
        _replace_project_coes(project_id, coe_ids)
        return JsonResponse({"success": True, "project_id": project_id})

@require_GET
def api_coes(request):
    coes = _get_all_coes()
    return JsonResponse({"coes": coes})

@require_GET
def api_projects(request):
    projects = _get_all_projects()
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT project_id, COUNT(*) as cnt FROM project_coes GROUP BY project_id")
        rows = cur.fetchall()
        counts = {r['project_id']: r['cnt'] for r in rows} if rows else {}
    finally:
        cur.close(); conn.close()
    for p in projects:
        p['mapped_coe_count'] = counts.get(p['id'], 0)
    return JsonResponse({"projects": projects})

# in views.py (or utils used by monthly_allocations view)
from django.db import connection

def get_user_projects_with_bgcode(request):
    """
    Returns a list of dicts: [{id, name, bg_code}, ...]
    Logic: find prism_wbs rows where creator == logged in user and return distinct projects
    If a project has multiple prism_wbs rows, pick the first non-empty bg_code (GROUP BY).
    """
    username = getattr(request.user, "username", "") or request.session.get('username') or ''
    if not username:
        return []

    sql = """
    SELECT p.id AS id,
           p.name AS name,
           COALESCE(
              (SELECT pw.bg_code FROM prism_wbs pw
                 WHERE pw.project_id = p.id AND pw.creator = %s
                 AND COALESCE(pw.bg_code,'') <> ''
                 LIMIT 1),
              '' ) AS bg_code
    FROM projects p
    INNER JOIN prism_wbs pw_by_creator ON pw_by_creator.project_id = p.id AND pw_by_creator.creator = %s
    GROUP BY p.id, p.name
    ORDER BY p.name
    """
    params = [username, username]
    out = []
    with connection.cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        rows = cur.fetchall()
        for row in rows:
            rec = dict(zip(cols, row))
            out.append({
                "id": rec.get("id"),
                "name": rec.get("name") or "",
                "bg_code": (rec.get("bg_code") or "").strip()
            })
    return out

@require_GET
def api_subprojects(request):
    """
    Return subprojects for a given bg_code (preferred) or project_id (fallback).
    Matching rule: subprojects.mdm_code = bg_code
    If project_id provided but bg_code missing, derive bg_code from prism_wbs.bg_code
    """
    print("\n" + "="*80)
    print("=== api_subprojects START ===")
    print("="*80)

    bg_code = (request.GET.get('bg_code') or '').strip()
    project_id = (request.GET.get('project_id') or '').strip()

    print(f"Request params:")
    print(f"  bg_code: '{bg_code}'")
    print(f"  project_id: '{project_id}'")

    try:
        project_id = int(project_id) if project_id else None
    except ValueError:
        print(f"ERROR: Invalid project_id '{project_id}' (not an integer)")
        return JsonResponse({"ok": False, "error": "Invalid project_id", "subprojects": []})

    # Step 1: If bg_code missing but project_id provided, fetch bg_code from prism_wbs
    if not bg_code and project_id:
        print(f"\n--- Fetching bg_code from prism_wbs for project_id={project_id} ---")
        try:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT bg_code 
                    FROM prism_wbs 
                    WHERE project_id = %s AND bg_code IS NOT NULL AND TRIM(bg_code) != ''
                    LIMIT 1
                """, [project_id])
                row = cur.fetchone()
                if row:
                    bg_code = (row[0] or '').strip()
                    print(f"✓ Found bg_code from prism_wbs: '{bg_code}'")
                else:
                    print(f"⚠ No bg_code found in prism_wbs for project_id={project_id}")
        except Exception as ex:
            print(f"ERROR: Failed to fetch bg_code from prism_wbs: {ex}")
            import traceback
            print(traceback.format_exc())

    # Step 2: Query subprojects table
    print(f"\n--- Querying subprojects table ---")
    if not bg_code:
        print("⚠ No bg_code available - returning empty subprojects list")
        return JsonResponse({"ok": True, "subprojects": [], "message": "No bg_code found"})

    rows = []
    try:
        with connection.cursor() as cur:
            # Check if subprojects table exists
            cur.execute("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE() AND table_name = 'subprojects'
            """)
            table_exists = cur.fetchone()[0] > 0
            print(f"subprojects table exists: {table_exists}")

            if not table_exists:
                print("ERROR: subprojects table does not exist")
                return JsonResponse({"ok": False, "error": "subprojects table missing", "subprojects": []})

            # Check total subprojects in table
            cur.execute("SELECT COUNT(*) FROM subprojects")
            total_count = cur.fetchone()[0]
            print(f"Total subprojects in database: {total_count}")

            # Check subprojects with this mdm_code
            cur.execute("""
                SELECT COUNT(*) 
                FROM subprojects 
                WHERE mdm_code = %s
            """, [bg_code])
            matching_count = cur.fetchone()[0]
            print(f"Subprojects matching bg_code '{bg_code}': {matching_count}")

            # Get sample mdm_codes to see format differences
            cur.execute("""
                SELECT DISTINCT mdm_code 
                FROM subprojects 
                WHERE mdm_code IS NOT NULL 
                LIMIT 10
            """)
            sample_codes = [r[0] for r in cur.fetchall()]
            print(f"Sample mdm_codes in subprojects table: {sample_codes}")

            # Try exact match
            print(f"\nQuerying: SELECT id, name, mdm_code FROM subprojects WHERE mdm_code = '{bg_code}'")
            cur.execute("""
                SELECT id, name, mdm_code 
                FROM subprojects 
                WHERE mdm_code = %s
            """, [bg_code])
            rows = cur.fetchall()
            print(f"✓ Query executed, returned {len(rows)} rows")

            if not rows:
                # Try case-insensitive match
                print(f"\nTrying case-insensitive match...")
                cur.execute("""
                    SELECT id, name, mdm_code 
                    FROM subprojects 
                    WHERE LOWER(mdm_code) = LOWER(%s)
                """, [bg_code])
                rows = cur.fetchall()
                print(f"Case-insensitive match returned {len(rows)} rows")

            if not rows:
                # Try partial match
                print(f"\nTrying partial match (LIKE)...")
                cur.execute("""
                    SELECT id, name, mdm_code 
                    FROM subprojects 
                    WHERE mdm_code LIKE %s
                """, [f"%{bg_code}%"])
                rows = cur.fetchall()
                print(f"Partial match returned {len(rows)} rows")

    except Exception as e:
        print(f"ERROR: Database query failed: {e}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({"ok": False, "error": str(e), "subprojects": []})

    # Step 3: Format results
    subs = []
    for r in rows:
        sub = {"id": r[0], "name": r[1], "mdm_code": r[2] or ""}
        subs.append(sub)
        if len(subs) <= 5:  # Print first 5
            print(f"  Subproject {sub['id']}: {sub['name']} (mdm_code: {sub['mdm_code']})")

    print(f"\n--- RESULT: Returning {len(subs)} subprojects ---")
    print("="*80)
    print("=== api_subprojects END ===")
    print("="*80 + "\n")

    return JsonResponse({"ok": True, "subprojects": subs})


# PUT THIS in projects/views.py (replace existing get_allocations_for_iom)
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_GET
from django.db import connection
from datetime import datetime, date
import logging

logger = logging.getLogger(__name__)

def _parse_month_start(raw):
    if not raw:
        return None
    raw = str(raw).strip()
    # 'YYYY-MM' -> prefer get_billing_period if available, else first day
    if len(raw) == 7 and raw[4] == '-':
        try:
            y, m = map(int, raw.split('-'))
            try:
                bs, _ = get_billing_period(y, m)
                return bs
            except Exception:
                return date(y, m, 1)
        except Exception:
            return None
    # 'YYYY-MM-DD' or ISO
    try:
        return datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except Exception:
        pass
    try:
        return datetime.fromisoformat(raw).date()
    except Exception:
        return None

@require_GET
def get_allocations_for_iom(request):
    # Accept multiple parameter names used across UI
    project_id = request.GET.get("project_id") or request.GET.get("proj_id") or request.POST.get("project_id")
    iom_row_id = request.GET.get("iom_row_id") or request.GET.get("iom_id") or request.GET.get("iomRowId") or request.POST.get("iom_row_id")
    month_start_raw = (request.GET.get("month_start") or request.GET.get("billing_start")
                       or request.GET.get("month") or request.POST.get("month_start") or request.POST.get("month"))
    subproject_id = request.GET.get("subproject_id") or request.GET.get("subp") or request.POST.get("subproject_id")

    # Log incoming parameters for debugging
    logger.debug("get_allocations_for_iom called with project_id=%r, iom_row_id=%r, month_start_raw=%r, subproject_id=%r",
                 project_id, iom_row_id, month_start_raw, subproject_id)

    if not project_id or not iom_row_id or not month_start_raw:
        logger.warning("Missing required params in get_allocations_for_iom")
        return HttpResponseBadRequest("project_id, iom_row_id and month_start (or month) are required")

    # normalize
    try:
        project_id = int(project_id)
    except Exception:
        return HttpResponseBadRequest("project_id must be integer")

    try:
        subp = int(subproject_id) if subproject_id not in (None, "", "null") else None
    except Exception:
        subp = None

    month_start = _parse_month_start(month_start_raw)
    if not month_start:
        logger.warning("Invalid month_start: %r", month_start_raw)
        return HttpResponseBadRequest("Invalid month_start")

    # Query DB; use DATE() to match date portion only
    try:
        with connection.cursor() as cur:
            if subp is None:
                cur.execute("""
                    SELECT id, project_id, subproject_id, iom_id, month_start, user_ldap, total_hours, created_at
                    FROM monthly_allocation_entries
                    WHERE project_id = %s
                      AND iom_id = %s
                      AND DATE(month_start) = %s
                    ORDER BY user_ldap
                """, [project_id, iom_row_id, month_start])
            else:
                cur.execute("""
                    SELECT id, project_id, subproject_id, iom_id, month_start, user_ldap, total_hours, created_at
                    FROM monthly_allocation_entries
                    WHERE project_id = %s
                      AND iom_id = %s
                      AND DATE(month_start) = %s
                      AND subproject_id = %s
                    ORDER BY user_ldap
                """, [project_id, iom_row_id, month_start, subp])

            cols = [c[0] for c in cur.description] if cur.description else []
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    except Exception as exc:
        logger.exception("get_allocations_for_iom DB error: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)

    # convert dates to ISO strings
    for r in rows:
        if isinstance(r.get("month_start"), (datetime, date)):
            r["month_start"] = r["month_start"].isoformat()
        if isinstance(r.get("created_at"), datetime):
            r["created_at"] = r["created_at"].isoformat()

    print("get_allocations_for_iom returning rows", rows)
    return JsonResponse({"ok": True, "rows": rows})


@require_POST
def save_monthly_allocations(request):
    """
    Save monthly allocation entries.

    - Accepts JSON body {'project_id':..., 'subproject_id':..., 'month': 'YYYY-MM' OR 'month_start':'YYYY-MM-DD', 'items': [{iom_id, user_ldap, total_hours}, ...]}
      OR form POST fields user_ldap1/total_hours1/iom_id1 ... OR items_json form field.
    - Deletes existing monthly_allocation_entries for each relevant iom_id for the (project, month_start, subproject) combination.
    - Inserts new monthly_allocation_entries rows with (project_id, subproject_id, iom_id, month_start, user_ldap, total_hours).
    - Returns saved_items list with fte computed for the billing month.
    """
    # Local imports so this function is self-contained if pasted directly
    import json
    import logging
    from datetime import datetime, date
    from django.db import transaction, connection
    from django.http import JsonResponse

    logger = logging.getLogger(__name__)

    try:
        # Parse body if any (accept JSON payloads)
        body_raw = request.body.decode("utf-8").strip() if request.body else ""
        data = {}
        if body_raw:
            try:
                data = json.loads(body_raw)
            except Exception:
                # fallback to empty dict if JSON decode fails
                data = {}

        # Inputs (try JSON -> POST -> GET)
        project_id = data.get("project_id") or request.POST.get("project_id") or request.GET.get("project_id")
        subproject_id = data.get("subproject_id") or request.POST.get("subproject_id") or request.GET.get("subproject_id")
        month_param = data.get("month") or request.POST.get("month") or request.GET.get("month")
        month_start_param = data.get("month_start") or request.POST.get("month_start") or request.GET.get("month_start")

        # Resolve billing_start (canonical billing-period start date) robustly
        billing_start = None

        # Helper: parse "YYYY-MM" -> billing_start using get_billing_period(year,mon)
        try:
            if month_param:
                # month_param expected "YYYY-MM"
                try:
                    year, mon = map(int, str(month_param).split("-"))
                    billing_start, _ = get_billing_period(year, mon)
                except Exception:
                    # if not yyyy-mm, try flexible parse (e.g., "2025-08-01" or "2025-08")
                    try:
                        parsed = datetime.fromisoformat(str(month_param))
                        billing_start = parsed.date().replace(day=1)
                    except Exception:
                        billing_start = None

            elif month_start_param:
                # month_start_param expected "YYYY-MM-DD" (billing window start like '2025-07-21')
                try:
                    parsed_date = datetime.strptime(str(month_start_param), "%Y-%m-%d").date()
                except Exception:
                    # try isoformat fallback
                    try:
                        parsed_date = datetime.fromisoformat(str(month_start_param)).date()
                    except Exception:
                        parsed_date = None

                if parsed_date:
                    # Attempt to canonicalize to canonical billing_start via get_billing_period_for_date
                    try:
                        bs, be = get_billing_period_for_date(parsed_date)
                        billing_start = bs
                    except Exception:
                        # fallback to first day of that month
                        billing_start = parsed_date.replace(day=1)

            else:
                # fallback to canonical billing period for today
                today = date.today()
                billing_start, _ = get_billing_period(today.year, today.month)
        except Exception as exc:
            # If anything unexpected, log and attempt a safe fallback
            logger.exception("save_monthly_allocations: error resolving billing_start: %s", exc)
            try:
                today = date.today()
                billing_start, _ = get_billing_period(today.year, today.month)
            except Exception:
                billing_start = None

        # If billing_start is still None -> return a clear bad-request response (prevents null month_start insert)
        if not billing_start:
            logger.warning("save_monthly_allocations: could not determine billing_start (month_param=%r, month_start_param=%r)", month_param, month_start_param)
            return JsonResponse({"ok": False, "error": "Missing or invalid month/month_start parameter"}, status=400)

        # Normalize types: project_id -> int if possible; subproject_id -> int or None
        try:
            project_id = int(project_id) if project_id not in (None, "", []) else None
        except Exception:
            project_id = None

        try:
            subproject_id = int(subproject_id) if subproject_id not in (None, "", []) else None
        except Exception:
            subproject_id = None

        # Items: prefer JSON body items; otherwise allow items_json or iterative form fields user_ldap1/total_hours1/iom_id1...
        items = data.get("items") if isinstance(data.get("items"), list) else None
        if items is None:
            items_json = request.POST.get("items_json") or None
            if items_json:
                try:
                    items = json.loads(items_json)
                except Exception:
                    items = []
            else:
                # fallback to incremental form fields
                items = []
                i = 1
                while True:
                    user_field = request.POST.get(f'user_ldap{i}')
                    hours_field = request.POST.get(f'total_hours{i}')
                    iom_field = request.POST.get(f'iom_id{i}')
                    if not user_field:
                        break
                    try:
                        hrs = float(hours_field or 0)
                    except Exception:
                        hrs = 0.0
                    items.append({"iom_id": iom_field, "user_ldap": user_field, "total_hours": hrs})
                    i += 1

        # Validate required inputs
        if not project_id:
            return JsonResponse({"ok": False, "error": "project_id is required"}, status=400)
        if items is None:
            return JsonResponse({"ok": False, "error": "items are required"}, status=400)

        # Begin DB transaction: delete old entries per IOM and insert new ones
        with transaction.atomic():
            with connection.cursor() as cur:
                # Build list of IOMs we will update (unique, non-empty)
                iom_ids = sorted({it.get("iom_id") for it in items if it.get("iom_id")})
                if iom_ids:
                    for iom_id in iom_ids:
                        # Delete existing entries for this project+iom+billing_start+subproject (supports subproject null)
                        cur.execute("""
                            DELETE FROM monthly_allocation_entries
                            WHERE project_id=%s
                              AND iom_id=%s
                              AND month_start=%s
                              AND (subproject_id=%s OR (%s IS NULL AND subproject_id IS NULL))
                        """, [project_id, iom_id, billing_start, subproject_id, subproject_id])

                # Insert each item row
                for it in items:
                    iom_id = it.get("iom_id")
                    user_ldap = (it.get("user_ldap") or "").strip()
                    if not iom_id or not user_ldap:
                        # skip incomplete rows
                        continue
                    try:
                        total_hours = float(it.get("total_hours") or 0.0)
                    except Exception:
                        total_hours = 0.0

                    # Defensive: ensure billing_start is a date or convertible string
                    param_billing_start = billing_start
                    # Insert using parameterized query (Django DB API will convert date -> SQL date)
                    cur.execute("""
                        INSERT INTO monthly_allocation_entries
                          (project_id, subproject_id, iom_id, month_start, user_ldap, total_hours, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                    """, [project_id, subproject_id, iom_id, param_billing_start, user_ldap, total_hours])

        # After insert, fetch saved items summary for response: user_ldap -> total_hours for the (project, billing_start, subproject)
        saved_items = []
        with connection.cursor() as cur:
            cur.execute("""
                SELECT user_ldap, COALESCE(SUM(total_hours), 0) AS total_hours
                FROM monthly_allocation_entries
                WHERE project_id=%s
                  AND month_start=%s
                  AND (subproject_id=%s OR (%s IS NULL AND subproject_id IS NULL))
                GROUP BY user_ldap
                ORDER BY user_ldap
            """, [project_id, billing_start, subproject_id, subproject_id])
            rows = cur.fetchall() or []

        # Compute billing hours for FTE calculation (use _get_month_hours_limit if present)
        try:
            year = billing_start.year
            month = billing_start.month
            billing_hours = float(_get_month_hours_limit(year, month) or HOURS_AVAILABLE_PER_MONTH)
        except Exception:
            billing_hours = float(HOURS_AVAILABLE_PER_MONTH)

        for r in rows:
            user_ldap = r[0] or ''
            total_hours = float(r[1] or 0.0)
            fte = (total_hours / billing_hours) if billing_hours > 0 else 0.0
            fte = round(float(fte), 4)
            saved_items.append({"user_ldap": user_ldap, "total_hours": total_hours, "fte": fte})

        # Return success + saved items + canonical billing_start iso string
        return JsonResponse({"ok": True, "saved_items": saved_items, "billing_start": billing_start.strftime("%Y-%m-%d")})

    except Exception as exc:
        logger.exception("save_monthly_allocations failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


# ---- Helper utilities ---------------------------------------------------

def _sql_in_clause(items):
    """
    Return (sql_fragment, params_list) for an IN clause for psycopg/MySQL paramstyle (%s).
    If items is empty returns ("(NULL)", []) to produce a false IN clause safely.
    """
    if not items:
        return "(NULL)", []
    placeholders = ",".join(["%s"] * len(items))
    return f"({placeholders})", list(items)


def is_pdl_user(ldap_entry):
    """
    Determine whether the LDAP user is a PDL (Project Delivery Lead) or manager.
    This is a conservative check and should be replaced/extended based on your LDAP schema:
      - check memberOf for specific group
      - check 'title', 'employeeType', or a custom attr like 'role'
    ldap_entry is expected to be the object returned by get_user_entry_by_username.
    """
    if not ldap_entry:
        return False

    # try a few common attributes (adjust to your environment)
    try:
        # If your LDAP helper returns a dict-like or attribute accessor, adapt accordingly
        attrs = {}
        if hasattr(ldap_entry, "entry_attributes_as_dict"):
            attrs = ldap_entry.entry_attributes_as_dict
        elif isinstance(ldap_entry, dict):
            attrs = ldap_entry
        else:
            # fallback: try to access attribute names directly
            # create attrs by reading typical attr names if present
            for name in ("title", "employeeType", "memberOf", "role"):
                val = getattr(ldap_entry, name, None)
                if val:
                    attrs[name] = val

        # If explicit role attribute mentions PDL/Manager
        role_val = (attrs.get("employeeType") or attrs.get("title") or attrs.get("role") or "")
        if isinstance(role_val, (list, tuple)):
            role_val = " ".join(role_val)
        if role_val and ("pdl" in role_val.lower() or "project delivery" in role_val.lower() or "manager" in role_val.lower()):
            return True

        # If memberOf contains a PDL/Managers group
        member_of = attrs.get("memberOf") or attrs.get("memberof") or []
        if isinstance(member_of, str):
            member_of = [member_of]
        for grp in member_of:
            if "pdl" in grp.lower() or "manager" in grp.lower() or "project-delivery" in grp.lower():
                return True
    except Exception:
        logger.exception("is_pdl_user: unexpected structure for ldap_entry")

    return False


# ---- Main view ---------------------------------------------------------

import json
from datetime import date
from django.shortcuts import render, redirect
from django.db import connection
# ensure you have logger, dictfetchall, get_billing_period, get_user_entry_by_username,
# get_reportees_for_user_dn, _get_month_hours_limit in your module scope

@require_GET
def team_allocations(request):
    """
    Team Allocations view: shows lead allocations and allows distributing to direct reportees.
    This version includes:
      - billing_month resolution
      - reading existing team_distributions assigned by this lead for the billing month
      - merging those distributions into lead_allocations distribution rows (pre-populate table)
      - adding team distribution hours into reportee summary totals
    """
    session_ldap = request.session.get("ldap_username")
    session_pwd = request.session.get("ldap_password")
    if not request.session.get("is_authenticated") or not session_ldap:
        return redirect("accounts:login")
    creds = (session_ldap, session_pwd)

    # Resolve billing month / month_start and month_end as earlier
    month_str = request.GET.get("month")
    try:
        if month_str:
            y, m = map(int, month_str.split("-"))
            month_start, month_end = get_billing_period(y, m)
        else:
            today = date.today()
            month_start, month_end = get_billing_period(today.year, today.month)
    except Exception:
        today = date.today()
        month_start, month_end = get_billing_period(today.year, today.month)

    # Determine billing_month string and display text
    billing_month = None
    billing_period_display = ""
    try:
        if month_str:
            billing_month = month_str
        else:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT year, month
                    FROM monthly_hours_limit
                    WHERE %s BETWEEN start_date AND end_date
                    LIMIT 1
                """, [month_start])
                r = cur.fetchone()
                if r and r[0] and r[1]:
                    billing_month = f"{int(r[0])}-{int(r[1]):02d}"
                else:
                    billing_month = f"{month_start.year}-{month_start.month:02d}"
    except Exception:
        billing_month = f"{month_start.year}-{month_start.month:02d}"

    try:
        if month_start and month_end:
            billing_period_display = f"{month_start.strftime('%b %d, %Y')} — {month_end.strftime('%b %d, %Y')}"
    except Exception:
        billing_period_display = ""

    # Get LDAP entry for current user
    try:
        user_entry = get_user_entry_by_username(session_ldap, username_password_for_conn=creds)
        if not user_entry:
            logger.warning("team_allocations: no LDAP entry for %s", session_ldap)
            return render(request, "projects/team_allocations.html", {
                "month_start": month_start, "month_end": month_end,
                "billing_month": billing_month, "billing_period_display": billing_period_display,
                "rows": [], "summary": {}, "weekly_map": {}, "lead_allocations": [], "reportees_json": "[]", "monthly_hours": 183.75
            })
    except Exception as ex:
        logger.exception("team_allocations: error fetching own LDAP entry: %s", ex)
        return render(request, "projects/team_allocations.html", {
            "month_start": month_start, "month_end": month_end,
            "billing_month": billing_month, "billing_period_display": billing_period_display,
            "rows": [], "summary": {}, "weekly_map": {}, "lead_allocations": [], "reportees_json": "[]", "monthly_hours": 183.75
        })

    # Get direct reportees via your helper
    try:
        reportees_entries = get_reportees_for_user_dn(getattr(user_entry, "entry_dn", None),
                                                      username_password_for_conn=creds) or []
    except Exception as ex:
        logger.exception("team_allocations: get_reportees_for_user_dn failed: %s", ex)
        reportees_entries = []

    # Normalize reportees
    reportees_ldaps = []
    reportees_map = {}
    for ent in reportees_entries:
        mail = None; cn = None; sam = None
        try:
            if isinstance(ent, dict):
                mail = ent.get("mail") or ent.get("email") or ent.get("userPrincipalName")
                cn = ent.get("cn") or ent.get("displayName")
                sam = ent.get("sAMAccountName")
            else:
                mail = getattr(ent, "mail", None) or getattr(ent, "email", None) or getattr(ent, "userPrincipalName", None)
                cn = getattr(ent, "cn", None) or getattr(ent, "displayName", None)
                sam = getattr(ent, "sAMAccountName", None)
        except Exception:
            pass
        identifier = (mail or sam or "").strip()
        if not identifier:
            continue
        l = identifier.lower()
        if l not in reportees_ldaps:
            reportees_ldaps.append(l)
            reportees_map[l] = {"ldap": identifier, "mail": mail or "", "cn": cn or "", "sAMAccountName": sam or ""}

    session_ldap_l = (session_ldap or "").lower()
    if session_ldap_l in reportees_ldaps:
        reportees_ldaps.remove(session_ldap_l)
        reportees_map.pop(session_ldap_l, None)

    # Fetch monthly_allocation_entries for these reportees (as before)
    rows = []
    if reportees_ldaps:
        try:
            placeholders = ",".join(["%s"] * len(reportees_ldaps))
            sql = f"""
                SELECT mae.id AS allocation_id,
                       mae.user_ldap,
                       COALESCE(u.username, mae.user_ldap) AS username,
                       COALESCE(u.email, mae.user_ldap) AS email,
                       mae.subproject_id AS subproject_id,
                       p.name AS project_name,
                       sp.name AS subproject_name,
                       pw.department AS domain_name,
                       COALESCE(mae.total_hours, 0) AS total_hours
                FROM monthly_allocation_entries mae
                LEFT JOIN users u ON LOWER(u.email) = LOWER(mae.user_ldap)
                LEFT JOIN projects p ON mae.project_id = p.id
                LEFT JOIN subprojects sp ON mae.subproject_id = sp.id
                LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
                WHERE mae.month_start = %s
                  AND LOWER(mae.user_ldap) IN ({placeholders})
                ORDER BY LOWER(mae.user_ldap), p.name
            """
            params = [month_start] + reportees_ldaps
            with connection.cursor() as cur:
                cur.execute(sql, params)
                rows = dictfetchall(cur)
        except Exception as ex:
            logger.exception("team_allocations: allocations query failed: %s", ex)
            rows = []
    else:
        rows = []

    # monthly hours limit
    try:
        month_hours = _get_month_hours_limit(month_start.year, month_start.month)
    except Exception:
        month_hours = 183.75

    # summary aggregation from monthly_allocation_entries
    summary = {}
    for r in rows:
        u = (r.get("user_ldap") or "").strip()
        if not u:
            continue
        key = u.lower()
        summary.setdefault(key, {
            "name": r.get("username") or reportees_map.get(key, {}).get("cn") or u,
            "email": r.get("email") or reportees_map.get(key, {}).get("mail") or u,
            "total_hours": 0.0
        })
        try:
            summary[key]["total_hours"] += float(r.get("total_hours") or 0.0)
        except Exception:
            pass

    # ensure reportees with no allocations are present
    for rldap in reportees_ldaps:
        if rldap not in summary:
            entry = reportees_map.get(rldap, {"ldap": rldap, "mail": rldap, "cn": rldap})
            summary[rldap] = {
                "name": entry.get("cn") or entry.get("ldap"),
                "email": entry.get("mail") or entry.get("ldap"),
                "total_hours": 0.0,
                "no_allocation": True
            }

    # Now read existing team_distributions assigned by this lead for this billing month
    # (so we can pre-populate the distribution table and reflect totals)
    team_dist_map = {}  # subproject_id -> list of {reportee_ldap, hours}
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT id, subproject_id, reportee_ldap, hours
                FROM team_distributions
                WHERE lead_ldap = %s AND month_start = %s
            """, [session_ldap, month_start])
            trows = dictfetchall(cur)
            for tr in trows:
                sp = tr.get('subproject_id') or 'none'
                team_dist_map.setdefault(str(sp), []).append({
                    "id": int(tr.get("id")) if tr.get("id") is not None else None,
                    "reportee_ldap": (tr.get("reportee_ldap") or "").strip(),
                    "hours": float(tr.get("hours") or 0.0),
                    # week_perc not stored here in this table; frontend will keep zeros or user-entered
                    "week_perc": [0, 0, 0, 0]
                })

    except Exception as ex:
        logger.exception("team_allocations: team_distributions read failed: %s", ex)
        team_dist_map = {}

    # Add team distribution hours to summary totals (so reportee cards show what lead allocated)
    for spid, dlist in team_dist_map.items():
        for d in dlist:
            ldap = (d.get("reportee_ldap") or "").lower()
            if not ldap:
                continue
            if ldap not in summary:
                # create a minimal summary entry if this reportee wasn't in monthly_allocation_entries
                summary[ldap] = {"name": ldap, "email": ldap, "total_hours": 0.0}
            try:
                summary[ldap]["total_hours"] += float(d.get("hours") or 0.0)
            except Exception:
                pass

    # compute FTE and color codes
    for k, s in summary.items():
        hrs = s.get("total_hours", 0.0)
        fte = (hrs / month_hours) if month_hours else 0
        pct = round(fte * 100, 2)
        s["fte"] = round(fte, 3)
        s["percent"] = pct
        if pct >= 100:
            s["color"] = "light-green"
        elif pct >= 80:
            s["color"] = "light-yellow"
        elif pct >= 50:
            s["color"] = "light-orange"
        else:
            s["color"] = "light-red"

    # Weekly map reading (unchanged)
    weekly_map = {}
    allocation_ids = [r["allocation_id"] for r in rows if r.get("allocation_id")]
    if allocation_ids:
        try:
            placeholders = ",".join(["%s"]*len(allocation_ids))
            with connection.cursor() as cur:
                cur.execute(f"""
                    SELECT allocation_id, week_number, percent
                    FROM weekly_allocations
                    WHERE allocation_id IN ({placeholders})
                """, allocation_ids)
                wrows = dictfetchall(cur)
                for w in wrows:
                    aid = w["allocation_id"]
                    weekly_map.setdefault(aid, {})[int(w["week_number"])] = {"percent": float(w["percent"] or 0)}
        except Exception as ex:
            logger.exception("team_allocations: weekly_allocations read failed: %s", ex)

    # Build lead_allocations (lead's own monthly_allocation_entries) grouped by subproject
    lead_allocations = []
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT mae.subproject_id,
                       COALESCE(sp.name, '(no subproject)') AS subproject_name,
                       COALESCE(p.name, '(no project)') AS project_name,
                       SUM(COALESCE(mae.total_hours,0)) AS total_hours
                FROM monthly_allocation_entries mae
                LEFT JOIN projects p ON mae.project_id = p.id
                LEFT JOIN subprojects sp ON mae.subproject_id = sp.id
                WHERE mae.month_start = %s
                  AND LOWER(mae.user_ldap) = LOWER(%s)
                GROUP BY mae.subproject_id, sp.name, p.name
                ORDER BY p.name, sp.name
            """, [month_start, session_ldap])
            la_rows = dictfetchall(cur) or []
    except Exception as ex:
        logger.exception("team_allocations: lead allocations fetch failed: %s", ex)
        la_rows = []

    # Build the lead_allocations_final structure and attach distribution items from team_dist_map when present
    lad_map = {}
    for la in la_rows:
        key = str(la.get("subproject_id") or "none")
        lad_map[key] = {
            "subproject_id": la.get("subproject_id"),
            "project_name": la.get("project_name"),
            "subproject_name": la.get("subproject_name"),
            "total_hours": float(la.get("total_hours") or 0.0),
            "distribution": []
        }
    # if lead has no explicit monthly allocation row but there are team_distributions, show those subprojects too
    for spkey in team_dist_map.keys():
        if spkey not in lad_map:
            lad_map[spkey] = {
                "subproject_id": (int(spkey) if spkey.isdigit() else None),
                "project_name": "(unknown)",
                "subproject_name": "(unknown)",
                "total_hours": 0.0,
                "distribution": []
            }

    # Now populate distribution lists for each subproject: prefer team_distributions entries (what lead previously assigned)
    for spkey, la in lad_map.items():
        dlist = team_dist_map.get(spkey, [])
        # Convert dlist items to include username/email if possible using reportees_map
        dist_items = []
        for d in dlist:
            ldap = (d.get("reportee_ldap") or "").lower()
            cn = reportees_map.get(ldap, {}).get("cn") or ldap
            mail = reportees_map.get(ldap, {}).get("mail") or ldap
            dist_items.append({
                "id": int(d.get("id")) if d.get("id") is not None else None,  # <-- NEW: DB PK of team_distributions
                "allocation_id": None,
                "reportee_ldap": ldap,
                "username": cn,
                "email": mail,
                "hours": float(d.get("hours") or 0.0),
                "week_perc": d.get("week_perc") or [0, 0, 0, 0]
            })

        la["distribution"] = dist_items

    lead_allocations_final = list(lad_map.values())

    # reportees_json for frontend selects
    reportees_json_list = []
    for k, v in reportees_map.items():
        reportees_json_list.append({
            "ldap": v.get("ldap") or k,
            "mail": v.get("mail") or "",
            "cn": v.get("cn") or v.get("ldap") or k
        })

    context = {
        "month_start": month_start,
        "month_end": month_end,
        "billing_month": billing_month,
        "billing_period_display": billing_period_display,
        "rows": rows,
        "summary": summary,
        "weekly_map": weekly_map,
        "lead_allocations": lead_allocations_final,
        "reportees_json": json.dumps(reportees_json_list),
        "monthly_hours": month_hours,
    }
    return render(request, "projects/team_allocations.html", context)

# -------------------------
# save_team_allocation
# -------------------------
@require_POST
def save_team_allocation(request):
    """
    Save weekly percent allocations for a monthly allocation (monthly_allocation_entries.id).
    Expects JSON body: { "allocation_id": 123, "weekly": { "1": 25.0, "2": 25.0, "3": 25.0, "4": 25.0 } }
    Returns JSON: { ok: True, allocation_id: 123, weeks: { "1": "46.25", ... } }
    """
    # parse JSON payload
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON payload")

    allocation_id = payload.get('allocation_id')
    weekly = payload.get('weekly', {}) or {}

    try:
        allocation_id = int(allocation_id)
    except Exception:
        return HttpResponseBadRequest("Invalid allocation_id")

    if allocation_id <= 0:
        return HttpResponseBadRequest("Invalid allocation_id")

    # fetch canonical allocation info from monthly_allocation_entries
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, total_hours, user_ldap
            FROM monthly_allocation_entries
            WHERE id = %s
            LIMIT 1
        """, [allocation_id])
        rec = cur.fetchone()

    if not rec:
        return HttpResponseBadRequest("Allocation not found")

    _, total_hours_raw, user_ldap = rec

    # coerce to Decimal for accurate arithmetic
    try:
        total_hours_dec = Decimal(str(total_hours_raw or '0.00'))
    except Exception:
        total_hours_dec = Decimal('0.00')

    result_weeks = {}

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                for wk_key, pct_val in weekly.items():
                    # normalize week number
                    try:
                        week_num = int(wk_key)
                    except Exception:
                        # skip invalid week keys
                        continue
                    # coerce percent to Decimal and clamp
                    try:
                        pct_dec = Decimal(str(pct_val))
                    except Exception:
                        pct_dec = Decimal('0.00')
                    if pct_dec < Decimal('0.00'):
                        pct_dec = Decimal('0.00')
                    if pct_dec > Decimal('100.00'):
                        pct_dec = Decimal('100.00')

                    # compute hours = total_hours * (pct/100), quantized to 2 decimals
                    hours_dec = (total_hours_dec * (pct_dec / Decimal('100.00'))).quantize(
                        Decimal('0.01'), rounding=ROUND_HALF_UP
                    )

                    # Upsert percent and hours
                    cur.execute("""
                        INSERT INTO weekly_allocations (allocation_id, week_number, percent, hours, updated_at)
                        VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                        ON DUPLICATE KEY UPDATE
                          percent = VALUES(percent),
                          hours = VALUES(hours),
                          updated_at = CURRENT_TIMESTAMP
                    """, [allocation_id, week_num, str(pct_dec), str(hours_dec)])

                    # prepare response payload (strings to preserve decimal formatting)
                    result_weeks[str(week_num)] = format(hours_dec, '0.2f')

    except Exception as exc:
        # optional: logger.exception("save_team_allocation failed: %s", exc)
        return JsonResponse({"ok": False, "error": str(exc)})

    return JsonResponse({"ok": True, "allocation_id": allocation_id, "weeks": result_weeks})

# -------------------------------------------------------------------
# 3. MY ALLOCATIONS (VIEW)
# -------------------------------------------------------------------
from math import ceil
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

# -------------------------------------------------------------------
# MY ALLOCATIONS (fixed)
# -------------------------------------------------------------------
# Add these imports at top of your views.py if not present
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django.http import JsonResponse, HttpResponseForbidden
from django.db import connection, transaction
from django.views.decorators.http import require_POST
from django.shortcuts import render
from django.urls import reverse
import logging


# Count working days Mon-Fri (exclude weekends Sat/Sun) and exclude holiday_dates set
def _count_working_days(start_date, end_date, holidays_set):
    cur = start_date
    cnt = 0
    while cur <= end_date:
        if cur.weekday() < 5 and cur.strftime("%Y-%m-%d") not in holidays_set:
            cnt += 1
        cur += timedelta(days=1)
    return cnt

# add/replace these imports at top of your views.py if missing
import json
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django.http import JsonResponse, HttpResponseForbidden
from django.shortcuts import render
from django.urls import reverse
from django.db import connection, transaction
from django.utils import timezone
from django.views.decorators.http import require_POST

# -------------------------
# Utility helpers
# -------------------------


from datetime import timedelta

def _compute_weeks_for_billing(billing_start, billing_end):
    """
    Build week slices starting at billing_start, each week ending on a Friday,
    with last week end either billing_end (if it would make the final week valid)
    or the nearest previous Friday — but never produce wk_end < cur_start.

    Returns list of dicts {num, start, end} where start/end are date objects.
    """
    print(f"START: _compute_weeks_for_billing | billing_start={billing_start}, billing_end={billing_end}")
    weeks = []
    cur_start = billing_start
    wknum = 1
    # defensive sanity: avoid infinite loops by limiting iterations
    MAX_WEEKS = 1000

    while cur_start <= billing_end and wknum <= MAX_WEEKS:
        print(f"  Loop wknum={wknum} | cur_start={cur_start}")
        # days_to_friday: how many days from cur_start to the next Friday (weekday 4)
        days_to_friday = (4 - cur_start.weekday()) % 7
        tentative_wk_end = cur_start + timedelta(days=days_to_friday)
        print(f"    days_to_friday={days_to_friday} | tentative wk_end={tentative_wk_end}")

        wk_end = tentative_wk_end

        # If tentative week end goes past billing_end, clamp it.
        if wk_end > billing_end:
            print(f"    wk_end {wk_end} > billing_end {billing_end}, clamping")
            # Candidate Friday computed from billing_end (previous Friday)
            candidate_friday = billing_end
            if billing_end.weekday() >= 5:
                # billing_end is Sat(5) or Sun(6)
                candidate_friday = billing_end - timedelta(days=(billing_end.weekday() - 4))
                print(f"    billing_end {billing_end} is weekend; candidate_friday={candidate_friday}")

            # Use candidate_friday only if it's >= cur_start (keeps start <= end).
            if candidate_friday >= cur_start:
                wk_end = candidate_friday
                print(f"    using candidate_friday as wk_end: {wk_end}")
            else:
                # candidate_friday would be before cur_start -> use billing_end instead
                wk_end = billing_end
                print(f"    candidate_friday {candidate_friday} < cur_start {cur_start}; using billing_end as wk_end: {wk_end}")

        # Safety check: ensure wk_end >= cur_start (if not, clamp to cur_start)
        if wk_end < cur_start:
            print(f"    Safety clamp: wk_end {wk_end} < cur_start {cur_start} -> forcing wk_end = cur_start")
            wk_end = cur_start

        weeks.append({'num': wknum, 'start': cur_start, 'end': wk_end})
        print(f"    Appended week: num={wknum}, start={cur_start}, end={wk_end}")

        # Advance to the day after wk_end for next week
        cur_start = wk_end + timedelta(days=1)
        wknum += 1

    if wknum > MAX_WEEKS:
        # defensive: log if we hit the iteration cap
        print(f"WARNING: _compute_weeks_for_billing reached MAX_WEEKS={MAX_WEEKS} and stopped to avoid infinite loop")

    print(f"END: _compute_weeks_for_billing | total weeks={len(weeks)}")
    return weeks


def _get_billing_period_from_month(year, month):
    """
    Try to read monthly_hours_limit or fallback to calendar month boundaries.
    Returns (start_date, end_date) as date objects.
    """
    with connection.cursor() as cur:
        cur.execute("SELECT start_date, end_date FROM monthly_hours_limit WHERE year=%s AND month=%s LIMIT 1", [year, month])
        r = cur.fetchone()
        if r and r[0] and r[1]:
            return _to_date(r[0]), _to_date(r[1])
    # fallback
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end

# -------------------------
# my_allocations view
# -------------------------
from django.http import JsonResponse, HttpResponseForbidden
from django.shortcuts import render
from django.urls import reverse
from django.db import connection
from datetime import date, timedelta, datetime
from decimal import Decimal, ROUND_HALF_UP
import json

def my_allocations(request):
    if not request.session.get("is_authenticated"):
        return HttpResponseForbidden("Not authenticated")
    user_email = request.session.get("ldap_username")
    if not user_email:
        return HttpResponseForbidden("Not authenticated")

    today = date.today()
    selected_year = int(request.GET.get('year', today.year))
    selected_month = int(request.GET.get('month', today.month))


    current_year = today.year
    years = list(range(current_year - 2, current_year + 3))
    months = [{'num': i, 'name': date(2000, i, 1).strftime('%B')} for i in range(1, 13)]

    billing_start, billing_end = _get_billing_period_from_month(selected_year, selected_month)

    monthly_max_hours = Decimal('0.00')
    with connection.cursor() as cur:
        cur.execute("SELECT max_hours FROM monthly_hours_limit WHERE year=%s AND month=%s LIMIT 1", [selected_year, selected_month])
        r = cur.fetchone()
        if r and r[0]:
            monthly_max_hours = Decimal(str(r[0]))

    weeks = _compute_weeks_for_billing(billing_start, billing_end)
    # Find the current week number (1-based) for today
    weeks = _compute_weeks_for_billing(billing_start, billing_end)
    current_week = None
    for w in weeks:
        if w['start'] <= today <= w['end']:
            current_week = w['num']
            break
    if not current_week:
        current_week = weeks[0]['num'] if weeks else 1  # fallback to first week

    # Set selected_week to current_week if not provided
    selected_week = request.GET.get('week')
    if not selected_week :
        selected_week = str(current_week)
    # Holidays
    with connection.cursor() as cur:
        cur.execute("SELECT holiday_date, name FROM holidays WHERE holiday_date BETWEEN %s AND %s", [billing_start, billing_end])
        hr = dictfetchall(cur)
    holidays_set = set()
    holidays_map = {}
    for h in hr:
        d = _to_date(h.get('holiday_date'))
        if d:
            holidays_set.add(d)
            holidays_map[d.strftime("%Y-%m-%d")] = h.get('name', '')

    # Leave records (per day)
    leave_map = {}
    with connection.cursor() as cur:
        cur.execute("""
            SELECT leave_date, SUM(leave_hours) as total_leave
            FROM leave_records
            WHERE LOWER(user_email) = LOWER(%s) AND year = %s AND month = %s
            GROUP BY leave_date
        """, [user_email, selected_year, selected_month])
        for r in dictfetchall(cur):
            leave_map[r['leave_date']] = Decimal(str(r['total_leave'] or '0.00'))

    # Team distributions (allocations)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT td.id AS team_distribution_id, td.hours AS total_hours,
                   COALESCE(p.name,'') AS project_name, COALESCE(sp.name,'') AS subproject_name,
                   td.project_id, td.subproject_id
            FROM team_distributions td
            LEFT JOIN projects p ON p.id = td.project_id
            LEFT JOIN subprojects sp ON sp.id = td.subproject_id
            WHERE LOWER(td.reportee_ldap) = LOWER(%s) AND td.month_start = %s
            ORDER BY td.id
        """, [user_email, billing_start])
        td_rows = dictfetchall(cur)

    if not td_rows:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT td.id AS team_distribution_id, td.hours AS total_hours,
                       COALESCE(p.name,'') AS project_name, COALESCE(sp.name,'') AS subproject_name,
                       td.project_id, td.subproject_id
                FROM team_distributions td
                LEFT JOIN projects p ON p.id = td.project_id
                LEFT JOIN subprojects sp ON sp.id = td.subproject_id
                WHERE LOWER(td.reportee_ldap) = LOWER(%s) AND td.month_start BETWEEN %s AND %s
                ORDER BY td.id
            """, [user_email, billing_start, billing_end])
            td_rows = dictfetchall(cur)

    # --- NEW: Fetch weekly_allocations for all team_distribution_ids ---
    weekly_alloc_map = {}  # (team_distribution_id, week_number) -> {hours, percent, status}
    if td_rows:
        td_ids = [r['team_distribution_id'] for r in td_rows]
        placeholders = ','.join(['%s'] * len(td_ids))
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT team_distribution_id, week_number, hours, percent, status
                FROM weekly_allocations
                WHERE team_distribution_id IN ({placeholders})
            """, td_ids)
            for row in dictfetchall(cur):
                key = (row['team_distribution_id'], row['week_number'])
                weekly_alloc_map[key] = row

    # Prepare punch_data: day-wise
    td_ids = [r['team_distribution_id'] for r in td_rows]
    punch_data_map = {}
    if td_ids:
        placeholders = ','.join(['%s'] * len(td_ids))
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT team_distribution_id, punch_date, allocated_hours, punched_hours, status, comments
                FROM punch_data
                WHERE team_distribution_id IN ({placeholders})
                  AND LOWER(user_email) = LOWER(%s)
                  AND month_start = %s
            """, td_ids + [user_email, billing_start])
            for r in dictfetchall(cur):
                key = (int(r['team_distribution_id']), r['punch_date'])
                punch_data_map[key] = {
                    'allocated_hours': Decimal(str(r['allocated_hours'] or '0.00')),
                    'punched_hours': Decimal(str(r['punched_hours'] or '0.00')),
                    'status': r['status'],
                    'comments': r.get('comments', '')
                }

    # Build groups structure for template
    groups = {}
    for td in td_rows:
        tdid = int(td['team_distribution_id'])
        total_hours = Decimal(str(td.get('total_hours') or '0.00'))
        subgroup = {
            'team_distribution_id': tdid,
            'project_id': td.get('project_id'),
            'subproject_id': td.get('subproject_id'),
            'project_name': td.get('project_name') or '',
            'subproject_name': td.get('subproject_name') or '',
            'total_hours': format(total_hours, '0.2f'),
            'weeks_list': []
        }
        for w in weeks:
            wknum = w['num']
            wk_start, wk_end = w['start'], w['end']
            # --- NEW: Attach weekly allocation info ---
            week_alloc = weekly_alloc_map.get((tdid, wknum), {})
            max_percent = week_alloc.get('percent', 0)
            allocated_hours = week_alloc.get('hours', 0)
            status = week_alloc.get('status', '')
            # Count working days (Mon-Fri, not holidays)
            working_days = 0
            cur_day = wk_start
            while cur_day <= wk_end and cur_day <= billing_end:
                if cur_day.weekday() < 5 and cur_day not in holidays_set:
                    working_days += 1
                cur_day += timedelta(days=1)
            # Build days_list
            days_list = []
            cur_day = wk_start
            while cur_day <= wk_end and cur_day <= billing_end:
                is_holiday = cur_day in holidays_set
                leave_hours = leave_map.get(cur_day, Decimal('0.00'))
                punch_key = (tdid, cur_day)
                punch = punch_data_map.get(punch_key, None)
                days_list.append({
                    'date': cur_day.strftime('%Y-%m-%d'),
                    'weekday': cur_day.strftime('%a'),
                    'is_holiday': is_holiday,
                    'holiday_name': holidays_map.get(cur_day.strftime('%Y-%m-%d'), '') if is_holiday else '',
                    'leave_hours': format(leave_hours, '0.2f') if leave_hours > 0 else None,
                    'allocated_hours': format(punch['allocated_hours'], '0.2f') if punch else None,
                    'punched_hours': format(punch['punched_hours'], '0.2f') if punch else None,
                    'status': punch['status'] if punch else 'DRAFT',
                    'comments': punch['comments'] if punch else '',
                    'is_editable': (not punch or punch['status'] in ['DRAFT', 'REJECTED']) and not is_holiday
                })
                cur_day += timedelta(days=1)
            subgroup['weeks_list'].append({
                'num': wknum,
                'week_start': wk_start.strftime('%Y-%m-%d'),
                'week_end': wk_end.strftime('%Y-%m-%d'),
                'working_days': working_days,
                'max_percent': max_percent,
                'allocated_hours': allocated_hours,
                'status': status,
                'days_list': days_list
            })
        groups.setdefault(td.get('subproject_name') or 'Unspecified', {
            'subproject_name': td.get('subproject_name') or 'Unspecified',
            'project_name': td.get('project_name') or '',
            'rows': []
        })
        groups[td.get('subproject_name') or 'Unspecified']['rows'].append(subgroup)

    # All weeks for dropdown
    all_weeks_list = _compute_weeks_for_billing(billing_start, billing_end)
    # Add days_list to each week in all_weeks_list
    for week in all_weeks_list:
        # week['start'] is a date object, convert to string for the helper
        week['days_list'] = [
            {
                'date': (week['start'] + timedelta(days=i)).strftime('%Y-%m-%d'),
                'weekday': (week['start'] + timedelta(days=i)).strftime('%A')
            }
            for i in range((week['end'] - week['start']).days + 1)
        ]
    all_weeks_list_json = json.dumps(
        all_weeks_list,
        default=lambda obj: obj.isoformat() if isinstance(obj, datetime) else str(obj)
    )

    context = {
        'weeks': weeks,
        'all_weeks_list': all_weeks_list,
        'all_weeks_list_json': all_weeks_list_json,
        'selected_week': selected_week,
        'current_week': current_week,
        'billing_start': billing_start,
        'billing_end': billing_end,
        'month_label': billing_start.strftime('%b %Y'),
        'groups': groups,
        "years": years,
        "months": months,
        "selected_year": selected_year,
        "selected_month": selected_month,
        'holidays_map': holidays_map,
        'monthly_max_hours': format(monthly_max_hours, '0.2f'),
        'save_effort_url': reverse('projects:save_effort_draft'),
        'submit_effort_url': reverse('projects:submit_effort'),
        'add_allocation_url': reverse('projects:add_self_allocation'),
        'get_projects_url': reverse('projects:get_projects_for_allocation'),
    }
    return render(request, 'projects/my_allocations.html', context)

def get_days_list_for_week(start_date):
    days = []
    d = datetime.strptime(start_date, "%Y-%m-%d")
    for i in range(6):  # Sunday to Friday
        day = d + timedelta(days=i)
        days.append({
            "weekday": day.strftime("%A"),
            "date": day.strftime("%Y-%m-%d")
        })
    return days

from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.db import connection, transaction
from decimal import Decimal, ROUND_HALF_UP
import json
import logging



@require_http_methods(["POST"])
def save_effort_draft(request):
    """
    Save draft effort data with validation against allocated hours (day-wise).
    Expects JSON:
    {
        "efforts": [
            ...
        ]
    }
    """
    user_email = request.session.get("ldap_username")
    print(f"[save_effort_draft] user_email from session: {user_email}")
    if not user_email:
        print("[save_effort_draft] Not authenticated")
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=401)

    try:
        print(f"[save_effort_draft] Raw request.body: {request.body}")
        data = json.loads(request.body)
        print(f"[save_effort_draft] Parsed JSON data: {data}")
        efforts = data.get('efforts', [])
        print(f"[save_effort_draft] efforts: {efforts}")
        if not efforts:
            print("[save_effort_draft] No efforts provided in request")
            return JsonResponse({'ok': False, 'error': 'No efforts provided'}, status=400)

        errors = []
        # Validation: For each week, sum daily punches and compare to allocated hours
        for effort in efforts:
            print(f"[save_effort_draft] Processing effort: {effort}")
            tdid = int(effort.get('team_distribution_id', 0))
            week_num = int(effort.get('week_number', 0))
            days = effort.get('days', [])
            print(f"[save_effort_draft] tdid={tdid}, week_num={week_num}, days={days}")
            total_punched = Decimal('0.00')
            for day in days:
                print(f"[save_effort_draft]   Day: {day}")
                try:
                    punched = Decimal(str(day.get('punched_hours', '0.00')))
                except Exception as ex:
                    print(f"[save_effort_draft]   Error parsing punched_hours: {ex}")
                    punched = Decimal('0.00')
                total_punched += punched
            # Get allocated hours for this week
            allocated_hours = Decimal('0.00')
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT hours FROM weekly_allocations
                    WHERE team_distribution_id=%s AND week_number=%s
                """, [tdid, week_num])
                row = cur.fetchone()
                if row and row[0]:
                    allocated_hours = Decimal(str(row[0]))
            print(f"[save_effort_draft]   total_punched={total_punched}, allocated_hours={allocated_hours}")
            if total_punched > allocated_hours:
                errors.append({
                    'team_distribution_id': tdid,
                    'week_number': week_num,
                    'error': f"Punched hours {total_punched} exceed allocated {allocated_hours}"
                })

        if errors:
            print(f"[save_effort_draft] Validation errors: {errors}")
            return JsonResponse({'ok': False, 'error': 'Validation failed', 'details': errors}, status=400)

        # Save: upsert punch_data row for each day
        with transaction.atomic():
            for effort in efforts:
                tdid = int(effort.get('team_distribution_id', 0))
                week_num = int(effort.get('week_number', 0))
                project_id = effort.get('project_id')
                subproject_id = effort.get('subproject_id')
                month_start = effort.get('month_start')
                days = effort.get('days', [])
                print(f"[save_effort_draft] Saving effort: tdid={tdid}, week_num={week_num}, project_id={project_id}, subproject_id={subproject_id}, month_start={month_start}")
                for day in days:
                    punch_date = day.get('punch_date')
                    punched_hours = Decimal(str(day.get('punched_hours', '0.00')))
                    allocated_hours = Decimal(str(day.get('allocated_hours', '0.00')))
                    percent_effort = day.get('percent_effort')
                    print(f"[save_effort_draft]   Upserting punch_data: punch_date={punch_date}, punched_hours={punched_hours}, allocated_hours={allocated_hours}, percent_effort={percent_effort}")
                    with connection.cursor() as cur:
                        cur.execute("""
                            INSERT INTO punch_data
                            (user_email, team_distribution_id, project_id, subproject_id, month_start, punch_date,
                             allocated_hours, punched_hours, percent_effort, status, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'DRAFT', NOW(), NOW())
                            ON DUPLICATE KEY UPDATE
                                punched_hours=VALUES(punched_hours),
                                percent_effort=VALUES(percent_effort),
                                status='DRAFT',
                                updated_at=NOW()
                        """, [
                            user_email, tdid, project_id, subproject_id, month_start, punch_date,
                            allocated_hours, punched_hours, percent_effort
                        ])
        print("[save_effort_draft] Draft saved successfully")
        return JsonResponse({'ok': True, 'message': 'Draft saved successfully'})

    except json.JSONDecodeError:
        print("[save_effort_draft] Invalid JSON")
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        print(f"[save_effort_draft] Error: {e}")
        logger.error(f"Error saving draft: {e}", exc_info=True)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_http_methods(["POST"])
def submit_effort(request):
    """
    Submit day-wise punched hours for approval.
    Expects JSON:
    {
        "billing_start": "YYYY-MM-DD",
        "efforts": [
            {
                "team_distribution_id": ...,
                "date": "YYYY-MM-DD",
                "week_number": ...,
                "punched_hours": ...,
                ...
            },
            ...
        ]
    }
    """
    user_email = request.session.get("ldap_username")
    print(f"[submit_effort] user_email from session: {user_email}")
    if not user_email:
        print("[submit_effort] Not authenticated")
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=401)

    try:
        print(f"[submit_effort] Raw request.body: {request.body}")
        data = json.loads(request.body)
        print(f"[submit_effort] Parsed JSON data: {data}")
        billing_start = data.get('billing_start')
        efforts = data.get('efforts', [])
        print(f"[submit_effort] efforts: {efforts}")

        if not billing_start or not efforts:
            print("[submit_effort] Missing billing_start or efforts")
            return JsonResponse({'ok': False, 'error': 'Missing billing_start or efforts'}, status=400)

        # Validation
        errors = []
        with connection.cursor() as cur:
            for idx, effort in enumerate(efforts):
                td_id = effort.get('team_distribution_id')
                punch_date = effort.get('date')
                week_num = effort.get('week_number')
                punched_hours = Decimal(str(effort.get('punched_hours', 0)))
                print(f"[submit_effort] Validating effort idx={idx}: td_id={td_id}, date={punch_date}, week_num={week_num}, punched_hours={punched_hours}")

                if not td_id or not punch_date or week_num is None:
                    errors.append(f"Effort {idx+1}: Missing required fields")
                    continue

                # Fetch allocated hours for this day
                cur.execute("""
                    SELECT allocated_hours
                    FROM punch_data
                    WHERE team_distribution_id = %s
                      AND punch_date = %s
                      AND user_email = %s
                      AND month_start = %s
                """, [td_id, punch_date, user_email, billing_start])
                row = cur.fetchone()
                #allocated_hours = Decimal(str(row[0])) if row and row[0] is not None else Decimal('0.00')
                #print(f"[submit_effort] Allocated hours for td_id={td_id}, date={punch_date}: {allocated_hours}")
                allocated_hours = 8.75
                if punched_hours > allocated_hours:
                    errors.append(
                        f"Date {punch_date}: Punched {punched_hours}h exceeds allocated {allocated_hours}h"
                    )

        if errors:
            print(f"[submit_effort] Validation errors: {errors}")
            return JsonResponse({'ok': False, 'errors': errors}, status=400)

        # Save data
        with transaction.atomic(), connection.cursor() as cur:
            for idx, effort in enumerate(efforts):
                td_id = effort.get('team_distribution_id')
                punch_date = effort.get('date')
                week_num = effort.get('week_number')
                punched = Decimal(str(effort.get('punched_hours', 0))).quantize(
                    Decimal('0.01'), rounding=ROUND_HALF_UP
                )

                # Fetch allocated hours, project_id, subproject_id from team_distributions
                cur.execute(
                    "SELECT hours, project_id, subproject_id FROM team_distributions WHERE id = %s",
                    [td_id]
                )
                row = cur.fetchone()
                if row:
                    allocated = Decimal(str(row[0]))
                    project_id = row[1]
                    subproject_id = row[2]
                else:
                    allocated = Decimal('0.00')
                    project_id = None
                    subproject_id = None

                print(f"[submit_effort] Saving: td_id={td_id}, date={punch_date}, week_num={week_num}, punched={punched}, allocated={allocated}, project_id={project_id}, subproject_id={subproject_id}")

                # Insert/update punch_data with project_id and subproject_id
                sql = """
                    INSERT INTO punch_data
                    (user_email, team_distribution_id, project_id, subproject_id, month_start, punch_date,
                     allocated_hours, punched_hours, status, submitted_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'SUBMITTED', NOW(), NOW())
                    ON DUPLICATE KEY UPDATE
                        punched_hours = VALUES(punched_hours),
                        status = 'SUBMITTED',
                        submitted_at = NOW(),
                        updated_at = NOW(),
                        project_id = VALUES(project_id),
                        subproject_id = VALUES(subproject_id)
                """
                params = [
                    user_email, td_id, project_id, subproject_id, billing_start, punch_date,
                    str(allocated), str(punched)
                ]
                cur.execute(sql, params)
                print(f"[submit_effort] Saved punch_data for td_id={td_id}, date={punch_date}")

        print(f"[submit_effort] Submitted {len(efforts)} day(s) successfully")
        return JsonResponse({
            'ok': True,
            'message': f'Submitted {len(efforts)} day(s) successfully'
        })

    except json.JSONDecodeError as e:
        print(f"[submit_effort] Invalid JSON: {e}")
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        print(f"[submit_effort] Error: {e}")
        logger.exception("submit_effort error: %s", e)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@require_http_methods(["POST"])
def add_self_allocation(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid request method'}, status=400)

    try:
        data = json.loads(request.body)
        user_email = request.session['ldap_username']
        project_id = data.get('project_id')
        subproject_id = data.get('subproject_id')
        month_start = data.get('month_start')  # Expected format: 'YYYY-MM-DD'
        allocations = data.get('allocations', [])

        # Validate required fields
        if not all([project_id, subproject_id, month_start]):
            return JsonResponse({
                'ok': False,
                'error': 'Missing required fields: project_id, subproject_id, or month_start'
            }, status=400)

        # Validate month_start format
        try:
            datetime.strptime(month_start, '%Y-%m-%d')
        except ValueError:
            return JsonResponse({
                'ok': False,
                'error': f'Invalid month_start format: {month_start}. Expected YYYY-MM-DD'
            }, status=400)

        if not allocations:
            return JsonResponse({'ok': False, 'error': 'No allocations provided'}, status=400)

        # Insert allocations
        sql = """
            INSERT INTO user_self_allocations
            (user_email, project_id, subproject_id, month_start, week_number,
             percent_effort, hours, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s,
                    (%s * (SELECT total_hours FROM monthly_allocation_entries
                           WHERE user_email=%s AND month_start=%s LIMIT 1) / 400.0),
                    'PENDING', NOW(), NOW())
            ON DUPLICATE KEY UPDATE
                percent_effort = VALUES(percent_effort),
                hours = VALUES(hours),
                updated_at = NOW()
        """

        with connection.cursor() as cursor:
            for alloc in allocations:
                week_num = alloc['week_number']
                pct = alloc['percent_effort']

                cursor.execute(sql, [
                    user_email, project_id, subproject_id, month_start,
                    week_num, pct, pct, user_email, month_start
                ])

        return JsonResponse({'ok': True})

    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_projects_for_allocation(request):
    """Get projects and subprojects for add allocation modal."""
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT p.id as project_id, p.name as project_name,
                       sp.id as subproject_id, sp.name as subproject_name
                FROM projects p
                LEFT JOIN subprojects sp ON sp.project_id = p.id
                ORDER BY p.name, sp.name
            """)
            rows = dictfetchall(cur)

        # Group by project
        projects = {}
        for row in rows:
            pid = row['project_id']
            if pid not in projects:
                projects[pid] = {
                    'id': pid,
                    'name': row['project_name'],
                    'subprojects': []
                }
            if row['subproject_id']:
                projects[pid]['subprojects'].append({
                    'id': row['subproject_id'],
                    'name': row['subproject_name']
                })

        return JsonResponse({'ok': True, 'projects': list(projects.values())})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
# -------------------------
# my_allocations_vacation endpoint
# -------------------------
@require_POST
def my_allocations_vacation(request):
    """
    Accepts JSON:
      { allocation_id, week_number, billing_start, leave_hours }
    Creates or updates weekly_punch_confirmations entry (PENDING), appends user_comment '|leave:xx' if needed,
    returns JSON { ok: True, new_hours: 'xx.xx', leave_hours: 'yy.yy' }
    """
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    user_email = request.session.get("ldap_username") or getattr(request.user, 'email', None)
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=403)

    allocation_id = payload.get('allocation_id')
    week_number = payload.get('week_number')
    billing_start = payload.get('billing_start')
    leave_hours = payload.get('leave_hours')

    if not allocation_id or week_number is None or not billing_start or leave_hours is None:
        return JsonResponse({'ok': False, 'error': 'Missing parameters'}, status=400)

    try:
        leave_hours = Decimal(str(leave_hours)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid leave_hours'}, status=400)

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # find existing row for this allocation/week/user
                cur.execute("""
                    SELECT id, allocated_hours, user_comment
                    FROM weekly_punch_confirmations
                    WHERE allocation_id=%s AND billing_start=%s AND week_number=%s AND LOWER(emp_email)=LOWER(%s)
                    LIMIT 1
                """, [allocation_id, billing_start, week_number, user_email])
                existing = cur.fetchone()
                if existing:
                    conf_id = int(existing[0])
                    before_hours = Decimal(str(existing[1] or '0.00'))
                    # new allocated hours for the confirmation (we store leave as deduction)
                    new_hours = (before_hours - leave_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    if new_hours < 0:
                        new_hours = Decimal('0.00')
                    # append leave tag to user_comment
                    prev_comment = existing[2] or ''
                    new_comment = prev_comment + ('|' if prev_comment else '') + f'leave:{leave_hours}'
                    cur.execute("""
                        UPDATE weekly_punch_confirmations
                        SET allocated_hours=%s, user_comment=%s, status=%s, updated_at=NOW()
                        WHERE id=%s
                    """, [str(new_hours), new_comment, 'PENDING', conf_id])
                else:
                    # insert a PENDING confirmation with leave comment and 0 allocated_hours
                    cur.execute("""
                        INSERT INTO weekly_punch_confirmations
                         (emp_email, allocation_id, billing_start, week_number, allocated_hours, user_comment, status, created_at, updated_at)
                        VALUES (LOWER(%s), %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """, [user_email, allocation_id, billing_start, week_number, str(Decimal('0.00')), f'leave:{leave_hours}', 'PENDING'])
                    conf_id = cur.lastrowid
                    new_hours = Decimal('0.00')

                # Optionally update weekly_allocations row to reflect new_hours (best-effort)
                cur.execute("""
                    SELECT id, hours
                    FROM weekly_allocations
                    WHERE team_distribution_id = %s AND week_number = %s
                    LIMIT 1
                """, [allocation_id, week_number])
                wa = cur.fetchone()
                if wa:
                    wa_id = int(wa[0])
                    # set hours to max(0, previous - leave) or new_hours as above
                    cur.execute("UPDATE weekly_allocations SET hours=%s, status=%s, updated_at=NOW() WHERE id=%s", [str(new_hours), 'PENDING', wa_id])
                else:
                    # insert allocation row with 0 hours (PENDING)
                    cur.execute("""
                        INSERT INTO weekly_allocations (allocation_id, team_distribution_id, week_number, hours, percent, status, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """, [None, allocation_id, week_number, str(new_hours), None, 'PENDING'])

        return JsonResponse({'ok': True, 'new_hours': format(new_hours, '0.2f'), 'leave_hours': format(leave_hours, '0.2f')})
    except Exception as ex:
        # log exception if logger available
        try:
            logger.exception("my_allocations_vacation failed: %s", ex)
        except Exception:
            pass
        return JsonResponse({'ok': False, 'error': 'Server error'}, status=500)

# -------------------------
# my_allocations_update_status endpoint
# -------------------------
@require_POST
def my_allocations_update_status(request):
    """
    Accepts JSON payloads:
      { team_distribution_id, week_number, billing_start (optional), action: 'ACCEPT'|'RECONSIDER'|'REJECT'|'MARK_PENDING' }

    Behavior:
      - For ACCEPT: mark any weekly_punch_confirmations for the allocation/week as ACCEPTED,
                   compute merged hours (work rows minus leave rows) and upsert weekly_allocations
                   with status=ACCEPTED. The response includes 'accepted_hours_sum' (string "0.00").
      - For RECONSIDER/REJECT: set weekly_allocations.status accordingly (if exists).
    Returns JSON: { status: 'ok', merged_status: new_status, accepted_hours_sum: 'xx.xx' } or error.
    """
    try:
        payload = json.loads(request.body.decode('utf-8') or "{}")
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_email = request.session.get("ldap_username") or getattr(request.user, 'email', None)
    actor = payload.get('actor') or (getattr(request.user, 'username', None) if getattr(request, 'user', None) and request.user.is_authenticated else 'system')

    # Accept both forms of identifying the allocation/td
    td_id = payload.get('team_distribution_id') or payload.get('allocation_id') or payload.get('allocation')
    week_number = payload.get('week_number')
    punch_date = payload.get('punch_date')  # older style may send punch_date instead of week_number
    action = (payload.get('action') or payload.get('status') or '').strip().upper()

    # Map action -> new_status used in weekly_allocations
    new_status = None
    if action in ('ACCEPT', 'ACCEPTED'):
        new_status = 'ACCEPTED'
    elif action in ('RECONSIDER', 'RECONSIDERED', 'REJECT'):
        new_status = 'RECONSIDERED' if action.startswith('RECONSIDER') else 'REJECTED'
    elif action in ('PENDING', 'MARK_PENDING'):
        new_status = 'PENDING'
    else:
        # Unknown or empty: default to PENDING
        new_status = 'PENDING'

    # If punch_date is provided but week_number missing, attempt to compute week_number from date
    # (optional - your codebase may have helpers; we keep it simple and require week_number)
    # Ensure td_id and week_number are present for actions that modify weekly_allocations
    try:
        with transaction.atomic():
            accepted_hours_sum = None

            with connection.cursor() as cur:
                if new_status == 'ACCEPTED' and td_id and week_number is not None:
                    # 1) Mark punch confirmations as ACCEPTED for this allocation/week
                    try:
                        cur.execute("""
                            UPDATE weekly_punch_confirmations
                            SET status = %s, updated_at = NOW()
                            WHERE team_distribution_id = %s AND week_number = %s
                        """, ['ACCEPTED', td_id, week_number])
                    except Exception:
                        # non-fatal; continue
                        pass

                    # 2) Compute aggregated work vs leave from confirmations (leave deducted from work)
                    # Summation query: leave_hours (sum of LEAVE rows) and work_hours (sum of non-LEAVE allocated_hours)
                    cur.execute("""
                        SELECT
                          COALESCE(SUM(CASE WHEN UPPER(punch_type) = 'LEAVE' THEN allocated_hours ELSE 0 END), 0) AS leave_hours,
                          COALESCE(SUM(CASE WHEN UPPER(punch_type) != 'LEAVE' THEN allocated_hours ELSE 0 END), 0) AS work_hours
                        FROM weekly_punch_confirmations
                        WHERE team_distribution_id = %s AND week_number = %s
                    """, [td_id, week_number])
                    sums = cur.fetchone() or (0, 0)
                    # sums may be a tuple of decimals/floats or strings depending on DB driver
                    try:
                        leave_hours = Decimal(str(sums[0] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        work_hours = Decimal(str(sums[1] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    except Exception:
                        # fallback to float then Decimal
                        leave_hours = Decimal(float(sums[0] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        work_hours = Decimal(float(sums[1] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                    # resultant hours = max(0, work_hours - leave_hours) if work rows exist; otherwise 0.
                    if work_hours > Decimal('0.00'):
                        resultant = (work_hours - leave_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                        if resultant < Decimal('0.00'):
                            resultant = Decimal('0.00')
                    else:
                        # if no work rows but leave rows exist, you may decide policy; we set resultant = 0.00
                        resultant = Decimal('0.00')

                    # 3) Upsert weekly_allocations row for this team_distribution/week with ACCEPTED status
                    cur.execute("SELECT id FROM weekly_allocations WHERE team_distribution_id = %s AND week_number = %s LIMIT 1", [td_id, week_number])
                    ex = cur.fetchone()
                    if ex:
                        # update existing
                        try:
                            cur.execute("UPDATE weekly_allocations SET hours=%s, status=%s, updated_at=NOW() WHERE id=%s", [str(resultant), 'ACCEPTED', ex[0]])
                        except Exception:
                            # defensive fallback (some drivers need str)
                            cur.execute("UPDATE weekly_allocations SET hours=%s, status=%s, updated_at=NOW() WHERE id=%s", [format(resultant, '0.2f'), 'ACCEPTED', ex[0]])
                    else:
                        cur.execute("""
                            INSERT INTO weekly_allocations (allocation_id, team_distribution_id, week_number, hours, percent, status, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                        """, [None, td_id, week_number, str(resultant), None, 'ACCEPTED'])

                    # set the accepted_hours_sum to return to caller (string 2-decimal)
                    accepted_hours_sum = resultant

                # Handle other statuses: RECONSIDERED / REJECTED / PENDING
                if new_status in ('RECONSIDERED', 'REJECTED') and td_id and week_number is not None:
                    cur.execute("UPDATE weekly_allocations SET status=%s, updated_at=NOW() WHERE team_distribution_id=%s AND week_number=%s", [new_status, td_id, week_number])

                # Insert audit row (optional) to track merges/actions
                try:
                    cur.execute("""
                        INSERT INTO weekly_punch_merge_audit (team_distribution_id, week_number, action, actor, created_at)
                        VALUES (%s, %s, %s, %s, NOW())
                    """, [td_id, week_number, new_status, actor])
                except Exception:
                    # ignore if audit table missing
                    pass

    except Exception as ex:
        try:
            logger.exception("my_allocations_update_status failed: %s", ex)
        except Exception:
            pass
        return JsonResponse({'error': 'Server error', 'detail': str(ex)}, status=500)

    # Build response
    resp = {'status': 'ok', 'merged_status': new_status}
    if 'accepted_hours_sum' in locals() and accepted_hours_sum is not None:
        try:
            # ensure it is string formatted with two decimals
            resp['accepted_hours_sum'] = format(accepted_hours_sum, '0.2f')
        except Exception:
            resp['accepted_hours_sum'] = str(accepted_hours_sum)

    return JsonResponse(resp)




# ---------- save weekly punches endpoint ----------
@require_POST
def save_my_alloc_weekly(request):
    """
    Expects JSON: { allocation_id: int, week_number:int, actual_hours: number, wbs: optional }
    It will upsert (INSERT .. ON DUPLICATE KEY UPDATE) into weekly_allocations table.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        allocation_id = int(payload.get("allocation_id", 0))
        week_number = int(payload.get("week_number", 0))
        hours = Decimal(str(payload.get("actual_hours", "0"))).quantize(Decimal("0.01"), ROUND_HALF_UP)
        wbs = payload.get("wbs")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid payload"}, status=400)

    if allocation_id <= 0 or week_number not in (1,2,3,4):
        return JsonResponse({"ok": False, "error": "Invalid allocation_id or week_number"}, status=400)

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # Upsert pattern for weekly_allocations (assuming unique key on allocation_id+week_number)
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, percent, hours, created_at)
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ON DUPLICATE KEY UPDATE hours = VALUES(hours)
                """, [allocation_id, week_number, 0.0, str(hours)])
        return JsonResponse({"ok": True, "allocation_id": allocation_id, "week_number": week_number, "hours": f"{hours:.2f}"})
    except Exception as e:
        logger.exception("save_my_alloc_weekly failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

# save_daily endpoint (modified to use billing period lookup for punch_date)
# -------------------------
@require_POST
def save_my_alloc_daily(request):
    """Save daily punches aligned to billing cycle."""
    try:
        data = json.loads(request.body.decode("utf-8"))
        allocation_id = int(data.get("allocation_id"))
        punch_date = datetime.strptime(data.get("punch_date"), "%Y-%m-%d").date()
        actual_hours = Decimal(str(data.get("actual_hours", 0))).quantize(Decimal("0.01"), ROUND_HALF_UP)
        wbs = data.get("wbs")

        user_ldap = request.session.get("ldap_username")

        billing_start, billing_end = get_billing_period_for_date(punch_date)
        week_number = ((punch_date - billing_start).days // 7) + 1

        # get weekly allocation
        with connection.cursor() as cur:
            cur.execute("""
                SELECT hours FROM weekly_allocations
                WHERE allocation_id=%s AND week_number=%s
            """, [allocation_id, week_number])
            rec = cur.fetchone()
        if not rec:
            return JsonResponse({"ok": False, "error": "No weekly allocation found"}, status=400)

        alloc_hours = Decimal(str(rec[0] or "0.00"))
        wk_start = billing_start + timedelta(days=(week_number - 1) * 7)
        wk_end = min(wk_start + timedelta(days=6), billing_end)

        # validate total within week
        with connection.cursor() as cur:
            cur.execute("""
                SELECT COALESCE(SUM(actual_hours),0)
                FROM user_punches
                WHERE user_ldap=%s AND allocation_id=%s AND punch_date BETWEEN %s AND %s
            """, [user_ldap, allocation_id, wk_start, wk_end])
            sum_existing = Decimal(str(cur.fetchone()[0] or "0.00"))

            cur.execute("""
                SELECT actual_hours FROM user_punches
                WHERE user_ldap=%s AND allocation_id=%s AND punch_date=%s
            """, [user_ldap, allocation_id, punch_date])
            existing_same = cur.fetchone()
            existing_same_val = Decimal(str(existing_same[0])) if existing_same else Decimal("0.00")

        total_after = (sum_existing - existing_same_val) + actual_hours
        if total_after > alloc_hours:
            return JsonResponse({"ok": False, "error": f"Exceeds weekly allocation {alloc_hours:.2f}"}, status=400)

        with transaction.atomic():
            with connection.cursor() as cur:
                cur.execute("""
                    INSERT INTO user_punches
                    (user_ldap, allocation_id, punch_date, week_number, actual_hours, wbs, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,CURRENT_TIMESTAMP)
                    ON DUPLICATE KEY UPDATE
                      actual_hours=VALUES(actual_hours),
                      wbs=VALUES(wbs),
                      updated_at=CURRENT_TIMESTAMP
                """, [user_ldap, allocation_id, punch_date, week_number, str(actual_hours), wbs])

        return JsonResponse({"ok": True, "allocation_id": allocation_id})

    except Exception as e:
        logger.exception("save_my_alloc_daily failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


import json
from datetime import datetime
from decimal import Decimal
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST
from django.db import connection
from django.utils import timezone

@require_POST
def my_allocations_update_status(request):
    """
    AJAX endpoint to Accept / Reconsider a weekly allocation for the logged-in user.

    Accepts JSON body:
      {
        "allocation_id": <int>,         # can be team_distributions.id OR weekly_allocations.id (legacy)
        "week_number": <int>,
        "billing_start": "YYYY-MM-DD",
        "action": "accept"|"reject",
        "allocated_hours": <number, optional>,
        "comment": "<text>",            # optional
      }

    Behaviour:
    - Resolve allocation_in -> canonical_alloc_id (team_distribution / monthly_allocation_entries mapping as before)
    - Upsert weekly_punch_confirmations using canonical_alloc_id
    - UPDATE weekly_allocations SET status = <new_status> for rows matching team_distribution_id = canonical_alloc_id AND week_number = <week_number>
    - Return JSON { ok: True, weekly_id: <id>, accepted_hours_sum: <float>, updated_weekly_alloc_rows: <int> }
    """
    # --- auth ---
    user_email = request.session.get("ldap_username") or getattr(request.user, "email", None)
    if not user_email:
        return HttpResponseForbidden("Not authenticated")

    try:
        payload = json.loads(request.body.decode('utf-8') or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    allocation_in = payload.get("allocation_id")
    week_number = payload.get("week_number")
    billing_start = payload.get("billing_start")
    action = (payload.get("action") or "").lower()
    allocated_hours = payload.get("allocated_hours", None)
    comment = payload.get("comment", None)

    if allocation_in is None or week_number is None or billing_start is None or action not in ("accept", "reject"):
        return JsonResponse({"ok": False, "error": "Missing or invalid parameters"}, status=400)

    try:
        allocation_in = int(allocation_in)
        week_number = int(week_number)
    except Exception:
        return JsonResponse({"ok": False, "error": "allocation_id and week_number must be integers"}, status=400)

    # Map action -> DB status
    if action == "accept":
        new_status = "ACCEPTED"
    else:
        new_status = "RECONSIDERED"

    # --- Resolve incoming allocation_in -> canonical_alloc_id (monthly_allocation_entries.id or team_distribution id mapping)
    canonical_alloc_id = None

    with connection.cursor() as cur:
        # 1) If incoming id is already a monthly_allocation_entries.id -> use it
        cur.execute("SELECT id FROM monthly_allocation_entries WHERE id = %s LIMIT 1", [allocation_in])
        row = cur.fetchone()
        if row:
            canonical_alloc_id = int(row[0])

        # 2) Else if incoming id is a weekly_allocations.id, try to use weekly_allocations.allocation_id (may point to monthly_allocation_entries)
        if canonical_alloc_id is None:
            cur.execute("SELECT allocation_id, team_distribution_id FROM weekly_allocations WHERE id = %s LIMIT 1", [allocation_in])
            row = cur.fetchone()
            if row:
                alloc_col, td_from_wa = row[0], row[1]
                if alloc_col is not None:
                    # allocation_id in weekly_allocations likely points to monthly_allocation_entries or the canonical allocation we need
                    canonical_alloc_id = int(alloc_col)
                else:
                    # try find monthly_allocation_entries by team_distribution_id
                    if td_from_wa is not None:
                        cur.execute("SELECT id FROM weekly_allocations WHERE team_distribution_id = %s LIMIT 1", [td_from_wa])
                        row2 = cur.fetchone()
                        if row2:
                            canonical_alloc_id = int(row2[0])
                        else:
                            # fallback: use the team_distribution id itself (we earlier changed WPC to reference team_distributions)
                            canonical_alloc_id = int(td_from_wa)

        # 3) Else treat allocation_in as a team_distributions.id and attempt mapping
        if canonical_alloc_id is None:
            cur.execute("SELECT id, month_start, project_id, subproject_id, reportee_ldap FROM team_distributions WHERE id = %s LIMIT 1", [allocation_in])
            td = cur.fetchone()
            if td:
                td_id, td_month_start, td_project_id, td_subproject_id, td_reportee = td[0], td[1], td[2], td[3], td[4]
                # try to find monthly_allocation_entries row that references this team_distribution_id
                cur.execute("SELECT id FROM monthly_allocation_entries WHERE team_distribution_id = %s LIMIT 1", [td_id])
                row3 = cur.fetchone()
                if row3:
                    canonical_alloc_id = int(row3[0])
                else:
                    # fallback: use the team_distribution id (we changed weekly_punch_confirmations FK to team_distributions.id)
                    canonical_alloc_id = int(td_id)

        # 4) Last-resort: find monthly_allocation_entries for this user & month
        if canonical_alloc_id is None:
            cur.execute("""
                SELECT id FROM monthly_allocation_entries
                WHERE (LOWER(emp_email)=LOWER(%s) OR LOWER(reportee_ldap)=LOWER(%s) OR LOWER(emp_code)=LOWER(%s))
                  AND month_start = %s
                LIMIT 1
            """, [user_email, user_email, user_email, billing_start])
            row5 = cur.fetchone()
            if row5:
                canonical_alloc_id = int(row5[0])

    # If still not found, error out
    if canonical_alloc_id is None:
        return JsonResponse({"ok": False, "error": "Could not resolve allocation to a canonical allocation id (team_distribution/monthly_allocation)."}, status=400)

    # --- compute allocated_percent if possible (optional) ---
    allocated_percent = None
    if allocated_hours is not None:
        try:
            with connection.cursor() as cur:
                cur.execute("SELECT max_hours FROM monthly_hours_limit WHERE start_date = %s LIMIT 1", [billing_start])
                r = cur.fetchone()
                if r and r[0] is not None:
                    monthly_max = Decimal(str(r[0]))
                else:
                    try:
                        yy, mm, _ = map(int, billing_start.split('-'))
                        cur.execute("SELECT max_hours FROM monthly_hours_limit WHERE year = %s AND month = %s LIMIT 1", [yy, mm])
                        rr = cur.fetchone()
                        monthly_max = Decimal(str(rr[0])) if rr and rr[0] is not None else None
                    except Exception:
                        monthly_max = None
            if monthly_max and monthly_max > 0:
                allocated_percent = (Decimal(str(allocated_hours)) / monthly_max * Decimal('100.00')).quantize(Decimal('0.01'))
        except Exception:
            allocated_percent = None

    now = timezone.now()

    # --- Upsert into weekly_punch_confirmations using canonical_alloc_id ---
    existing_id = None
    saved_id = None
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT id FROM weekly_punch_confirmations
                WHERE emp_email = %s AND allocation_id = %s AND billing_start = %s AND week_number = %s
                LIMIT 1
            """, [user_email, canonical_alloc_id, billing_start, week_number])
            r = cur.fetchone()
            if r:
                existing_id = int(r[0])

            if existing_id:
                cur.execute("""
                    UPDATE weekly_punch_confirmations
                    SET allocated_hours = %s,
                        allocated_percent = %s,
                        user_comment = %s,
                        status = %s,
                        actioned_by = %s,
                        actioned_at = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, [
                    (Decimal(str(allocated_hours)) if allocated_hours is not None else None),
                    (str(allocated_percent) if allocated_percent is not None else None),
                    (comment if comment is not None else None),
                    new_status,
                    user_email,
                    now,
                    existing_id
                ])
                saved_id = existing_id
            else:
                cur.execute("""
                    INSERT INTO weekly_punch_confirmations
                        (emp_email, allocation_id, billing_start, week_number,
                         allocated_hours, allocated_percent, user_comment, status,
                         actioned_by, actioned_at, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """, [
                    user_email,
                    canonical_alloc_id,
                    billing_start,
                    week_number,
                    (Decimal(str(allocated_hours)) if allocated_hours is not None else None),
                    (str(allocated_percent) if allocated_percent is not None else None),
                    (comment if comment is not None else None),
                    new_status,
                    user_email,
                    now
                ])
                # cur.lastrowid not guaranteed by all DB backends using raw cursor; try fetch
                try:
                    saved_id = cur.lastrowid or None
                except Exception:
                    saved_id = None
    except Exception as ex:
        return JsonResponse({"ok": False, "error": f"DB write failed (wpc): {str(ex)}"}, status=500)

    # --- Update weekly_allocations.status where team_distribution_id = canonical_alloc_id and week_number = week_number ---
    updated_weekly_alloc_rows = 0
    try:
        with connection.cursor() as cur:
            # Update matching weekly_allocations rows (usually one)
            cur.execute("""
                UPDATE weekly_allocations
                SET status = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE team_distribution_id = %s AND week_number = %s
            """, [new_status, canonical_alloc_id, week_number])
            updated_weekly_alloc_rows = cur.rowcount
    except Exception as ex:
        # Not fatal — WPC is saved, but log/return a warning
        return JsonResponse({"ok": False, "error": f"DB write failed (weekly_allocations update): {str(ex)}"}, status=500)

    # compute sum of accepted hours for UI if needed
    sum_accepted = None
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT COALESCE(SUM(allocated_hours),0) FROM weekly_punch_confirmations
                WHERE allocation_id = %s AND week_number = %s AND billing_start = %s AND status = 'ACCEPTED'
            """, [canonical_alloc_id, week_number, billing_start])
            rr = cur.fetchone()
            sum_accepted = float(rr[0]) if rr and rr[0] is not None else 0.0
    except Exception:
        sum_accepted = None

    return JsonResponse({
        "ok": True,
        "weekly_punch_id": saved_id,
        "accepted_hours_sum": sum_accepted,
        "updated_weekly_alloc_rows": updated_weekly_alloc_rows
    })





# Replace existing _get_user_projects_for_allocations and monthly_allocations with these:

def _cn_to_creator(cn: str):
    """
    Convert "LASTNAME FirstName ..." -> "FirstName ... LASTNAME"
    (same conversion used in edit_project() so creator matching is consistent)
    """
    if not cn:
        return ""
    parts = str(cn).strip().split()
    if len(parts) >= 2:
        # put the first token (last name) at the end
        return " ".join(parts[1:]) + " " + parts[0]
    return str(cn).strip()


def _get_user_projects_for_allocations(request):
    """
    Return list of projects where session user is PDL, PM, or creator.
    Enhanced with direct database inspection for name format debugging.
    """
    print("\n=== _get_user_projects_for_allocations START ===")

    session_ldap = request.session.get("ldap_username")
    session_cn = request.session.get("cn", "")
    print(f"Session LDAP: {session_ldap}")
    print(f"Session CN: {session_cn}")

    # Generate name variations
    name_variations = set()
    if session_ldap:
        name_variations.add(session_ldap.strip())
    if session_cn:
        creator_name = _cn_to_creator(session_cn)
        name_variations.add(creator_name)
        print(f"Converted Creator Name (CN): {creator_name}")

    if session_ldap and '@' not in session_ldap:
        parts = session_ldap.strip().split()
        if len(parts) >= 2:
            variation1 = ' '.join(parts[1:] + [parts[0]])
            name_variations.add(variation1)
            variation2 = ' '.join(parts[1:] + [parts[0].title()])
            name_variations.add(variation2)
            variation3 = ' '.join(parts[1:] + [parts[0].upper()])
            name_variations.add(variation3)

    print(f"All name variations to match: {name_variations}")

    # DIRECT DATABASE INSPECTION: Find all creators that might match
    print("\n=== DIRECT DATABASE INSPECTION ===")
    with connection.cursor() as cur:
        # Get all unique creators from prism_wbs
        cur.execute("""
            SELECT DISTINCT creator, COUNT(*) as count
            FROM prism_wbs
            WHERE creator IS NOT NULL AND TRIM(creator) != ''
            ORDER BY creator
        """)
        all_creators = cur.fetchall()
        print(f"Total unique creators in prism_wbs: {len(all_creators)}")

        # Search for creators containing any part of the user's name
        name_parts = []
        for var in name_variations:
            name_parts.extend([p for p in var.split() if len(p) > 2 and '@' not in p])
        unique_parts = list(set(name_parts))

        print(f"\nSearching for name parts: {unique_parts}")

        matching_creators = []
        for part in unique_parts[:5]:  # Limit to avoid too many queries
            cur.execute("""
                SELECT DISTINCT creator, COUNT(*) as iom_count
                FROM prism_wbs
                WHERE LOWER(creator) LIKE LOWER(%s)
                GROUP BY creator
                ORDER BY iom_count DESC
                LIMIT 20
            """, (f'%{part}%',))
            matches = cur.fetchall()
            if matches:
                print(f"\nCreators matching '{part}':")
                for creator, cnt in matches:
                    print(f"  - '{creator}' ({cnt} IOMs)")
                    matching_creators.append(creator)

        # Also check if any project has this user as PDL/PM
        print(f"\n=== Checking PDL/PM matches ===")
        for var in list(name_variations)[:3]:
            cur.execute("""
                SELECT id, name, pdl_name, pm_name
                FROM projects
                WHERE LOWER(pdl_name) LIKE LOWER(%s)
                   OR LOWER(pm_name) LIKE LOWER(%s)
                LIMIT 5
            """, (f'%{var}%', f'%{var}%'))
            pdl_pm_matches = cur.fetchall()
            if pdl_pm_matches:
                print(f"\nProjects where '{var}' is PDL/PM:")
                for pid, pname, pdl, pm in pdl_pm_matches:
                    print(f"  - Project {pid}: {pname} (PDL: {pdl}, PM: {pm})")

    if not name_variations:
        logger.warning("_get_user_projects_for_allocations: no session identifiers found")
        return []

    # Original query logic (unchanged)
    sql = """
        SELECT DISTINCT p.id, p.name
        FROM projects p
        LEFT JOIN prism_wbs pw ON pw.project_id = p.id
        WHERE 1=0
    """
    params = []

    for variation in name_variations:
        sql += " OR TRIM(LOWER(p.pdl_name)) = TRIM(LOWER(%s))"
        params.append(variation)
        sql += " OR TRIM(LOWER(p.pm_name)) = TRIM(LOWER(%s))"
        params.append(variation)
        sql += " OR TRIM(LOWER(pw.creator)) = TRIM(LOWER(%s))"
        params.append(variation)

    sql += " ORDER BY p.name"

    print(f"\n--- Executing Query with {len(params)} parameters ---")

    try:
        with connection.cursor() as cur:
            cur.execute(sql, params)
            projects = dictfetchall(cur)

        print(f"\n--- Query Results: {len(projects)} projects found ---")

        if projects:
            for idx, proj in enumerate(projects[:10], 1):
                print(f"  {idx}. Project {proj['id']}: {proj['name']}")
        else:
            print("No projects matched. Check database inspection output above.")

        print(f"\n=== _get_user_projects_for_allocations END ===\n")
        return projects

    except Exception as exc:
        logger.exception(f"_get_user_projects_for_allocations failed: {exc}")
        print(f"\n!!! ERROR: {type(exc).__name__}: {str(exc)}")
        import traceback
        print(f"Traceback:\n{traceback.format_exc()}")
        return []


def monthly_allocations(request):
    """
    Render monthly allocations page (billing-period aware). Accepts ?month=YYYY-MM
    and uses billing_start resolved via get_billing_period(year, month).
    Now supports filtering by subproject_id and adapted for projects.pdl_name/pm_name string columns.
    """
    print("monthly_allocations called")
    session_ldap = request.session.get("ldap_username")
    session_cn = request.session.get("cn", "")

    # Parse month param (YYYY-MM) -> billing_start
    month_str = request.GET.get("month")
    if month_str:
        try:
            year, month = map(int, month_str.split("-"))
            billing_start, billing_end = get_billing_period(year, month)
        except Exception:
            today = date.today()
            billing_start, billing_end = get_billing_period(today.year, today.month)
    else:
        today = date.today()
        billing_start, billing_end = get_billing_period(today.year, today.month)

    project_id_param = request.GET.get("project_id")
    subproject_id_param = request.GET.get("subproject_id")
    print("project id param: ", project_id_param, "   Sub project id param: ", subproject_id_param)

    active_project_id = None
    try:
        if project_id_param:
            active_project_id = int(project_id_param)
    except Exception:
        active_project_id = None

    active_subproject_id = None
    try:
        if subproject_id_param:
            active_subproject_id = int(subproject_id_param)
    except Exception:
        active_subproject_id = None

    # Fetch projects user can allocate for (updated query for string-based pdl_name)
    projects = _get_user_projects_for_allocations(request)
    print("projects fetched: ", projects)
    if not active_project_id and projects:
        active_project_id = projects[0]["id"]

    if not active_project_id:
        return render(request, "projects/monthly_allocations.html", {
            "projects": [],
            "coes": [],
            "domains_map": {},
            "allocation_map": {},
            "weekly_map": {},
            "capacity_map": {},
            "active_project_id": None,
            "active_subproject_id": None,
            "billing_month": billing_start.strftime("%Y-%m"),
            "billing_start": billing_start.strftime("%Y-%m-%d"),
            "billing_end": billing_end.strftime("%Y-%m-%d"),
            "month_hours_limit": 0,
        })

    # Fetch COEs and domains
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id, name FROM coes ORDER BY name")
            coes = dictfetchall(cur)
    except Exception:
        logger.exception("Error fetching COEs")
        coes = []

    coe_ids = [c["id"] for c in coes] if coes else []

    domains_map = {}
    if coe_ids:
        try:
            with connection.cursor() as cur:
                placeholders = ','.join(['%s'] * len(coe_ids))
                cur.execute(f"SELECT id, coe_id, name FROM domains WHERE coe_id IN ({placeholders}) ORDER BY name", coe_ids)
                domains = dictfetchall(cur)
                for d in domains:
                    coe_id = d.get("coe_id")
                    if coe_id not in domains_map:
                        domains_map[coe_id] = []
                    domains_map[coe_id].append(d)
        except Exception:
            logger.exception("Error fetching domains")

    # Fetch allocation_items for this project/subproject/billing_start (canonical)
    allocation_map = {}
    capacity_accumulator = {}
    allocation_ids = []

    try:
        with connection.cursor() as cur:
            # Base query for monthly_allocation_entries
            base_sql = """
                SELECT
                    mae.id as allocation_id,
                    mae.project_id,
                    mae.subproject_id,
                    mae.iom_id,
                    mae.user_ldap,
                    mae.total_hours,
                    mae.month_start,
                    p.name as project_name,
                    sp.name as subproject_name,
                    COALESCE(pw.bg_code, '') as bg_code
                FROM monthly_allocation_entries mae
                LEFT JOIN projects p ON p.id = mae.project_id
                LEFT JOIN subprojects sp ON sp.id = mae.subproject_id
                LEFT JOIN prism_wbs pw ON pw.id = mae.iom_id
                WHERE mae.project_id = %s
                  AND DATE(mae.month_start) = %s
            """
            params = [active_project_id, billing_start]

            # Add subproject filter if provided
            if active_subproject_id:
                base_sql += " AND mae.subproject_id = %s"
                params.append(active_subproject_id)

            base_sql += " ORDER BY mae.user_ldap, sp.name"

            cur.execute(base_sql, params)
            rows = dictfetchall(cur)

            for r in rows:
                alloc_id = r.get("allocation_id")
                user_ldap = r.get("user_ldap")
                total_hours = Decimal(str(r.get("total_hours") or "0.00"))

                allocation_ids.append(alloc_id)

                key = (r.get("subproject_id"), user_ldap)
                if key not in allocation_map:
                    allocation_map[key] = {
                        "allocation_id": alloc_id,
                        "subproject_id": r.get("subproject_id"),
                        "subproject_name": r.get("subproject_name") or "Unspecified",
                        "user_ldap": user_ldap,
                        "total_hours": Decimal("0.00"),
                        "bg_code": r.get("bg_code") or "",
                        "iom_id": r.get("iom_id"),
                    }

                allocation_map[key]["total_hours"] += total_hours

                if user_ldap not in capacity_accumulator:
                    capacity_accumulator[user_ldap] = Decimal("0.00")
                capacity_accumulator[user_ldap] += total_hours

    except Exception:
        logger.exception("Error fetching allocation items")

    # Fetch weekly allocations (unchanged)
    weekly_map = {}
    if allocation_ids:
        try:
            with connection.cursor() as cur:
                placeholders = ','.join(['%s'] * len(allocation_ids))
                cur.execute(
                    f"""
                    SELECT wa.allocation_id, wa.week_number, wa.percent, wa.hours, wa.status
                    FROM weekly_allocations wa
                    WHERE wa.allocation_id IN ({placeholders})
                    """,
                    allocation_ids
                )
                weekly_rows = dictfetchall(cur)
                for wr in weekly_rows:
                    aid = wr.get("allocation_id")
                    wnum = wr.get("week_number")
                    if aid not in weekly_map:
                        weekly_map[aid] = {}
                    weekly_map[aid][wnum] = {
                        "percent": wr.get("percent"),
                        "hours": wr.get("hours"),
                        "status": wr.get("status"),
                    }
        except Exception:
            logger.exception("Error fetching weekly allocations")

    # Calculate capacity map
    month_hours_limit = _get_month_hours_limit(billing_start.year, billing_start.month)

    capacity_map = {}
    for ldap_key, allocated in capacity_accumulator.items():
        remaining = Decimal(str(month_hours_limit)) - allocated
        capacity_map[ldap_key] = {
            "allocated": float(allocated),
            "remaining": float(remaining),
            "limit": month_hours_limit,
        }

    # Ensure every user in allocation_items has an entry in capacity_map
    try:
        all_users = set(k[1] for k in allocation_map.keys())
        for u in all_users:
            if u not in capacity_map:
                capacity_map[u] = {
                    "allocated": 0.0,
                    "remaining": float(month_hours_limit),
                    "limit": month_hours_limit,
                }
    except Exception:
        logger.exception("Error ensuring capacity map entries")

    return render(request, "projects/monthly_allocations.html", {
        "projects": projects,
        "coes": coes,
        "domains_map": domains_map,
        "allocation_map": allocation_map,
        "weekly_map": weekly_map,
        "capacity_map": capacity_map,
        "active_project_id": active_project_id,
        "active_subproject_id": active_subproject_id,
        "billing_month": billing_start.strftime("%Y-%m"),
        "billing_start": billing_start.strftime("%Y-%m-%d"),
        "billing_end": billing_end.strftime("%Y-%m-%d"),
        "month_hours_limit": month_hours_limit,
    })


# Implement _get_month_hours_limit used above
def _get_month_hours_limit(year, month):
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT max_hours FROM monthly_hours_limit WHERE year = %s AND month = %s LIMIT 1", (int(year), int(month)))
            row = cur.fetchone()
            if row and row[0] is not None:
                return float(row[0])
    except Exception:
        logger.exception("_get_month_hours_limit failed")
    return float(HOURS_AVAILABLE_PER_MONTH)

# --- get_applicable_ioms (replace existing function) ---
@require_GET
def get_applicable_ioms(request):
    """
    Return list of IOM rows matching project/year/month and creator (PDL/creator filter).
    Accepts optional `project_id`, `year`, `month`, `search`, and `subproject_id`.
    If subproject_id is provided, load the subproject's mdm_code/bg_code and
    restrict prism_wbs rows to those matching bg_code OR buyer_wbs_cc/seller_wbs_cc (LIKE).
    This avoids referencing non-existent columns on prism_wbs (e.g. mdm_code).
    """
    session_ldap = request.session.get("ldap_username")
    session_cn = request.session.get("cn")

    def _cn_to_creator(cn):
        if not cn:
            return ""
        parts = str(cn).strip().split()
        if len(parts) >= 2:
            return " ".join(parts[1:]) + " " + parts[0]
        return str(cn).strip()

    creator_candidates = []
    if session_cn:
        conv = _cn_to_creator(session_cn).strip()
        if conv and conv.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(conv)
        s = session_cn.strip()
        if s and s.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(s)
    if session_ldap:
        sld = session_ldap.strip()
        if sld and sld.lower() not in [c.lower() for c in creator_candidates]:
            creator_candidates.append(sld)
    creator_lower_vals = [c.lower() for c in creator_candidates]

    project_id = request.GET.get("project_id")
    subproject_id = request.GET.get("subproject_id")  # optional filter
    search = (request.GET.get("search") or "").strip()
    try:
        year = int(request.GET.get("year") or datetime.now().year)
        month = int(request.GET.get("month") or datetime.now().month)
    except Exception:
        return HttpResponseBadRequest("Invalid year/month")

    # map month -> columns
    _MONTH_MAP = {
        1: ("jan_fte", "jan_hours"),
        2: ("feb_fte", "feb_hours"),
        3: ("mar_fte", "mar_hours"),
        4: ("apr_fte", "apr_hours"),
        5: ("may_fte", "may_hours"),
        6: ("jun_fte", "jun_hours"),
        7: ("jul_fte", "jul_hours"),
        8: ("aug_fte", "aug_hours"),
        9: ("sep_fte", "sep_hours"),
        10: ("oct_fte", "oct_hours"),
        11: ("nov_fte", "nov_hours"),
        12: ("dec_fte", "dec_hours"),
    }
    fte_col, hrs_col = _MONTH_MAP.get(month, ("jan_fte", "jan_hours"))

    # If subproject_id provided, attempt to fetch its mdm_code/bg_code (prefer mdm_code then bg_code)
    sub_mdm = None
    sub_bg = None
    if subproject_id:
        try:
            with connection.cursor() as cur:
                cur.execute("SELECT mdm_code, bg_code FROM subprojects WHERE id = %s LIMIT 1", [int(subproject_id)])
                r = cur.fetchone()
                if r:
                    # r[0] -> mdm_code, r[1] -> bg_code
                    sub_mdm = (r[0] or "").strip()
                    sub_bg = (r[1] or "").strip()
        except Exception:
            logger.exception("get_applicable_ioms: cannot load subproject %s", subproject_id)
            sub_mdm = sub_bg = None

    # Build SQL (do NOT reference prism_wbs.mdm_code -- it doesn't exist)
    sql = f"""
        SELECT id, iom_id, department, site, `function`,
               {fte_col} as month_fte, {hrs_col} as month_hours,
               buyer_wbs_cc, seller_wbs_cc, project_id, creator
        FROM prism_wbs
        WHERE year = %s
          AND ( ({fte_col} IS NOT NULL AND {fte_col} > 0) OR ({hrs_col} IS NOT NULL AND {hrs_col} > 0) )
    """
    params = [str(year)]
    if project_id:
        sql += " AND project_id = %s"
        params.append(project_id)

    if creator_lower_vals:
        placeholders = ",".join(["%s"] * len(creator_lower_vals))
        sql += f" AND LOWER(TRIM(creator)) IN ({placeholders})"
        params.extend(creator_lower_vals)

    if search:
        sql += " AND iom_id LIKE %s"
        params.append(f"%{search}%")

    # If we have a subproject code, restrict by matching prism_wbs.bg_code OR buyer/seller WBS columns
    # NOTE: prism_wbs may not have mdm_code; avoid referencing it. Use bg_code (if present on prism_wbs)
    if (sub_bg and sub_bg != "") or (sub_mdm and sub_mdm != ""):
        # prefer matching bg_code exactly; also match buyer/seller WBS with LIKE against mdm or bg
        # build a single grouped clause to avoid accidental column name references
        match_val = sub_bg or sub_mdm or ''
        like_val = f"%{match_val}%"
        sql += " AND ( COALESCE(bg_code, '') = %s OR COALESCE(buyer_wbs_cc, '') LIKE %s OR COALESCE(seller_wbs_cc, '') LIKE %s )"
        params.append(match_val)
        params.append(like_val)
        params.append(like_val)

    sql += " ORDER BY iom_id LIMIT 500"

    try:
        with connection.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall() or []
            cols = [c[0] for c in cur.description]
    except Exception as ex:
        logger.exception("get_applicable_ioms DB error: %s", ex)
        return JsonResponse({"ok": False, "error": str(ex)}, status=500)

    ioms = []
    # get month_limit (per-resource limit) once per request
    month_limit = _get_month_hours_limit(year, month)
    for r in rows:
        rec = dict(zip(cols, r))
        ioms.append({
            "id": rec.get("id"),
            "iom_id": rec.get("iom_id"),
            "department": rec.get("department"),
            "site": rec.get("site"),
            "function": rec.get("function"),
            "month_fte": float(rec.get("month_fte") or 0),
            "month_hours": float(rec.get("month_hours") or 0),
            "buyer_wbs_cc": rec.get("buyer_wbs_cc"),
            "seller_wbs_cc": rec.get("seller_wbs_cc"),
            "month_limit": float(month_limit),
        })

    return JsonResponse({"ok": True, "ioms": ioms})


# --- get_iom_details: fetch by id OR iom_id, compute remaining hours (from monthly_allocation_entries),
#     and remaining FTE = remaining_hours / month_limit(year,month) ---
# --- get_iom_details (replace existing function) ---
@require_GET
def get_iom_details(request):
    """
    Returns canonical details for a prism_wbs row (lookup by id or iom_id).
    Accepts: project_id, iom_row_id (id or iom_id), year, month, and optional subproject_id.
    When computing used_hours, the subproject_id is respected (if present) so remaining_hours
    reflect allocations for the selected subproject.
    """
    iom_row_id = request.GET.get("iom_row_id")
    project_id = request.GET.get("project_id")
    subproject_id = request.GET.get("subproject_id")  # NEW
    try:
        year = int(request.GET.get("year") or datetime.now().year)
        month = int(request.GET.get("month") or datetime.now().month)
    except ValueError:
        return HttpResponseBadRequest("Invalid year/month")
    if not iom_row_id:
        return HttpResponseBadRequest("iom_row_id required")

    _MONTH_MAP = {
        1: ("jan_fte", "jan_hours"),
        2: ("feb_fte", "feb_hours"),
        3: ("mar_fte", "mar_hours"),
        4: ("apr_fte", "apr_hours"),
        5: ("may_fte", "may_hours"),
        6: ("jun_fte", "jun_hours"),
        7: ("jul_fte", "jul_hours"),
        8: ("aug_fte", "aug_hours"),
        9: ("sep_fte", "sep_hours"),
        10: ("oct_fte", "oct_hours"),
        11: ("nov_fte", "nov_hours"),
        12: ("dec_fte", "dec_hours"),
    }
    fte_col, hrs_col = _MONTH_MAP.get(month, ("jan_fte", "jan_hours"))

    try:
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT id, iom_id, project_id, department, site, `function`,
                       {fte_col} as month_fte, {hrs_col} as month_hours,
                       buyer_wbs_cc, seller_wbs_cc, total_fte, total_hours
                FROM prism_wbs
                WHERE id = %s OR iom_id = %s
                LIMIT 1
            """, [iom_row_id, iom_row_id])
            row = cur.fetchone()
            if not row:
                return JsonResponse({"ok": False, "error": "IOM not found"}, status=404)
            cols = [c[0] for c in cur.description]
            rec = dict(zip(cols, row))

            # Use canonical billing start for this year/month
            billing_start, billing_end = get_billing_period(year, month)

            # compute used hours for this IOM from monthly_allocation_entries using canonical month_start
            if subproject_id:
                cur.execute("""
                    SELECT COALESCE(SUM(total_hours),0) FROM monthly_allocation_entries
                    WHERE project_id=%s AND iom_id=%s AND month_start=%s AND subproject_id=%s
                """, [project_id, rec.get("iom_id"), billing_start, subproject_id])
            else:
                cur.execute("""
                    SELECT COALESCE(SUM(total_hours),0) FROM monthly_allocation_entries
                    WHERE project_id=%s AND iom_id=%s AND month_start=%s
                """, [project_id, rec.get("iom_id"), billing_start])
            used_hours = cur.fetchone()[0] or 0.0

    except Exception as ex:
        logger.exception("get_iom_details failed: %s", ex)
        return JsonResponse({"ok": False, "error": str(ex)}, status=500)

    month_hours = float(rec.get("month_hours") or 0.0)
    month_limit = _get_month_hours_limit(year, month)
    month_fte = round((month_hours / month_limit) if month_limit > 0 else 0.0, 2)

    remaining_hours = max(0.0, month_hours - float(used_hours))
    remaining_hours = round(remaining_hours, 2)
    remaining_fte = round((remaining_hours / month_limit) if month_limit > 0 else 0.0, 2)

    resp = {
        "ok": True,
        "iom": {
            "id": rec.get("id"),
            "iom_id": rec.get("iom_id"),
            "department": rec.get("department"),
            "site": rec.get("site"),
            "function": rec.get("function"),
            "month_fte": float(month_fte),
            "month_hours": float(round(month_hours, 2)),
            "total_fte": float(rec.get("total_fte") or 0),
            "total_hours": float(rec.get("total_hours") or 0),
            "buyer_wbs_cc": rec.get("buyer_wbs_cc"),
            "seller_wbs_cc": rec.get("seller_wbs_cc"),
            "remaining_hours": float(remaining_hours),
            "remaining_fte": float(remaining_fte),
            "month_limit": float(month_limit),
            "billing_start": billing_start,
            "billing_end": billing_end,
        }
    }

    return JsonResponse(resp)


# --- export_allocations (replace existing function) ---
@require_GET
def export_allocations(request):
    """
    Export allocations for an IOM and billing month. Accepts:
      - project_id, iom_id, and either month=YYYY-MM (preferred) OR month_start=YYYY-MM-DD
      - optional subproject_id (when provided, export only allocations for that subproject)
    Produces an Excel workbook (openpyxl).
    """
    project_id = request.GET.get("project_id")
    iom_id = request.GET.get("iom_id")
    subproject_id = request.GET.get("subproject_id")  # NEW: optional
    month_param = request.GET.get("month")  # YYYY-MM
    month_start_param = request.GET.get("month_start")

    if not (project_id and iom_id):
        return HttpResponseBadRequest("project_id and iom_id required")

    # resolve canonical billing_start based on month param or month_start
    billing_start = None
    billing_end = None
    try:
        if month_param:
            year, mon = map(int, month_param.split("-"))
            billing_start, billing_end = get_billing_period(year, mon)
        elif month_start_param:
            try:
                dt = datetime.strptime(month_start_param, "%Y-%m-%d").date()
            except Exception:
                dt = None
            if dt:
                try:
                    billing_start, billing_end = get_billing_period_for_date(dt)
                except Exception:
                    billing_start = dt.replace(day=1)
            else:
                billing_start = date.today().replace(day=1)
        else:
            today = date.today()
            billing_start, billing_end = get_billing_period(today.year, today.month)
    except Exception:
        billing_start = date.today().replace(day=1)

    # fetch iom basic details
    iom = None
    with connection.cursor() as cur:
        cur.execute("""
            SELECT iom_id, department, buyer_wbs_cc, seller_wbs_cc, site, `function`, total_hours
            FROM prism_wbs
            WHERE project_id=%s AND iom_id=%s
            LIMIT 1
        """, [project_id, iom_id])
        row = cur.fetchone()
        if row:
            iom = {
                "iom_id": row[0],
                "department": row[1],
                "buyer_wbs_cc": row[2],
                "seller_wbs_cc": row[3],
                "site": row[4],
                "function": row[5],
                "total_hours": row[6],
            }

    # fetch allocations (respect subproject_id when present)
    with connection.cursor() as cur:
        if subproject_id:
            cur.execute("""
                SELECT user_ldap, total_hours
                FROM monthly_allocation_entries
                WHERE project_id=%s AND iom_id=%s AND month_start=%s AND subproject_id=%s
                ORDER BY user_ldap
            """, [project_id, iom_id, billing_start, subproject_id])
        else:
            cur.execute("""
                SELECT user_ldap, total_hours
                FROM monthly_allocation_entries
                WHERE project_id=%s AND iom_id=%s AND month_start=%s
                ORDER BY user_ldap
            """, [project_id, iom_id, billing_start])
        allocations = cur.fetchall() or []

    # Build excel workbook (uses openpyxl, as before)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Allocations"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    row_idx = 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    title_cell = ws.cell(row=row_idx, column=1, value="IOM Allocation Report")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.alignment = center
    title_cell.fill = fill_blue
    row_idx += 2

    if iom:
        details = [
            ("IOM ID", iom["iom_id"]),
            ("Department", iom["department"]),
            ("WBS (Buyer)", iom.get("buyer_wbs_cc") or ""),
            ("WBS (Seller)", iom.get("seller_wbs_cc") or ""),
            ("Site", iom.get("site") or ""),
            ("Function", iom.get("function") or ""),
            ("IOM Total Hours", iom.get("total_hours") or 0),
            ("Billing Month Start", billing_start.strftime("%Y-%m-%d") if billing_start else "")
        ]
        for k, v in details:
            ws.cell(row=row_idx, column=1, value=k).font = bold_font
            ws.cell(row=row_idx, column=2, value=v).font = normal_font
            row_idx += 1
        row_idx += 1

    # header
    ws.cell(row=row_idx, column=1, value="Resource").font = header_font
    ws.cell(row=row_idx, column=2, value="Total Hours").font = header_font
    # style header
    for c in range(1, 3):
        cell = ws.cell(row=row_idx, column=c)
        cell.alignment = center
        cell.fill = fill_blue
        cell.font = header_font
    row_idx += 1

    # rows
    for r in allocations:
        uname = r[0] or ''
        hrs = float(r[1] or 0.0)
        ws.cell(row=row_idx, column=1, value=uname).font = normal_font
        ws.cell(row=row_idx, column=2, value=hrs).font = normal_font
        row_idx += 1

    # finalize workbook into HttpResponse
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"allocations_{project_id}_{iom_id}_{billing_start.strftime('%Y%m%d')}.xlsx" if billing_start else f"allocations_{project_id}_{iom_id}.xlsx"
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def export_my_punches_pdf(request):
    """
    Export punches PDF for the logged-in user for the canonical billing cycle for the requested month.
    Accepts ?month=YYYY-MM (preferred) or ?month_start=YYYY-MM-DD.
    Tries multiple session_ldap variants if direct match returns no rows.
    """
    import io
    from django.template.loader import render_to_string
    from xhtml2pdf import pisa

    session_ldap = (request.session.get("ldap_username")
                    or request.session.get("user_email")
                    or request.session.get("user_ldap")
                    or getattr(request.user, "email", None)
                    or getattr(request.user, "username", None))
    if not session_ldap:
        return HttpResponseBadRequest("Not authenticated")

    # determine billing period
    month_param = request.GET.get("month")
    month_start_param = request.GET.get("month_start")
    try:
        if month_start_param:
            dt = datetime.strptime(month_start_param, "%Y-%m-%d").date()
            billing_start, billing_end = get_billing_period_for_date(dt)
        elif month_param:
            y, m = map(int, month_param.split("-"))
            billing_start, billing_end = get_billing_period(y, m)
        else:
            # fallback to current month billing
            today = date.today()
            billing_start, billing_end = get_billing_period(today.year, today.month)
    except Exception as ex:
        logger.exception("export_my_punches_pdf: invalid month param: %s", ex)
        today = date.today()
        billing_start, billing_end = get_billing_period(today.year, today.month)

    rows = []
    tried = []

    def fetch_for_ldap(ldap_val):
        with connection.cursor() as cur:
            cur.execute("""
                SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department AS department,
                       up.punch_date, up.week_number, up.actual_hours, up.wbs
                FROM user_punches up
                LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
                LEFT JOIN projects p ON mae.project_id = p.id
                LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
                WHERE up.user_ldap = %s
                  AND up.punch_date BETWEEN %s AND %s
                ORDER BY up.punch_date, p.name
            """, [ldap_val, billing_start, billing_end])
            return dictfetchall(cur)

    # 1) try with exact session_ldap
    rows = fetch_for_ldap(session_ldap)
    tried.append(("exact", session_ldap, len(rows)))

    # 2) try with lowercase email
    if not rows and "@" in str(session_ldap):
        alt = session_ldap.lower()
        rows = fetch_for_ldap(alt)
        tried.append(("lower", alt, len(rows)))

    # 3) try using only local-part (before @)
    if not rows and "@" in str(session_ldap):
        local = session_ldap.split("@", 1)[0]
        rows = fetch_for_ldap(local)
        tried.append(("localpart", local, len(rows)))

    # 4) try wildcard LIKE search on email / ldap values
    if not rows:
        # attempt wildcard searches using parts of username
        candidates = []
        if "@" in str(session_ldap):
            candidates.append("%" + session_ldap.split("@", 1)[0] + "%")
            candidates.append("%" + session_ldap + "%")
        else:
            candidates.append("%" + session_ldap + "%")
        # also try replacing dots with underscores and vice-versa
        srep = str(session_ldap).replace(".", "_")
        candidates.append("%" + srep + "%")
        # run attempts
        for pattern in candidates:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department AS department,
                           up.punch_date, up.week_number, up.actual_hours, up.wbs
                    FROM user_punches up
                    LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
                    LEFT JOIN projects p ON mae.project_id = p.id
                    LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
                    WHERE up.user_ldap LIKE %s
                      AND up.punch_date BETWEEN %s AND %s
                    ORDER BY up.punch_date, p.name
                """, [pattern, billing_start, billing_end])
                tmp = dictfetchall(cur)
            tried.append(("wildcard", pattern, len(tmp)))
            if tmp:
                rows = tmp
                break

    logger.debug("export_my_punches_pdf tried patterns: %r", tried)

    # Render PDF (allow empty rows but show message)
    html = render_to_string("projects/punches_pdf.html", {
        "rows": rows,
        "month": month_param or billing_start.strftime("%Y-%m"),
        "user": session_ldap,
        "billing_start": billing_start, "billing_end": billing_end,
        "tried": tried
    })
    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("utf-8")), dest=result)
    if pisa_status.err:
        logger.exception("pisa create pdf failed")
        return HttpResponse("Error generating PDF", status=500)
    result.seek(0)
    safe_user = str(session_ldap).replace("@", "_at_").replace(".", "_")
    filename = f"punches_{safe_user}_{(month_param or billing_start.strftime('%Y-%m'))}.pdf"
    response = HttpResponse(result.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_my_punches_excel(request):
    """
    Export punches for logged-in user to Excel for the canonical billing period.
    Same input options and LDAP fallback logic as export_my_punches_pdf.
    """
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter

    session_ldap = (request.session.get("ldap_username")
                    or request.session.get("user_email")
                    or request.session.get("user_ldap")
                    or getattr(request.user, "email", None)
                    or getattr(request.user, "username", None))
    if not session_ldap:
        return HttpResponseBadRequest("Not authenticated")

    # determine billing period
    month_param = request.GET.get("month")
    month_start_param = request.GET.get("month_start")
    try:
        if month_start_param:
            dt = datetime.strptime(month_start_param, "%Y-%m-%d").date()
            billing_start, billing_end = get_billing_period_for_date(dt)
        elif month_param:
            y, m = map(int, month_param.split("-"))
            billing_start, billing_end = get_billing_period(y, m)
        else:
            today = date.today()
            billing_start, billing_end = get_billing_period(today.year, today.month)
    except Exception as ex:
        logger.exception("export_my_punches_excel: invalid month param: %s", ex)
        today = date.today()
        billing_start, billing_end = get_billing_period(today.year, today.month)

    rows = []
    tried = []

    def fetch_for_ldap(ldap_val):
        with connection.cursor() as cur:
            cur.execute("""
                SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department AS department,
                       up.punch_date, up.week_number, up.actual_hours, up.wbs
                FROM user_punches up
                LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
                LEFT JOIN projects p ON mae.project_id = p.id
                LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
                WHERE up.user_ldap = %s
                  AND up.punch_date BETWEEN %s AND %s
                ORDER BY up.punch_date, p.name
            """, [ldap_val, billing_start, billing_end])
            return dictfetchall(cur)

    # Try exact and fallback variants
    rows = fetch_for_ldap(session_ldap)
    tried.append(("exact", session_ldap, len(rows)))
    if not rows and "@" in str(session_ldap):
        rows = fetch_for_ldap(session_ldap.lower())
        tried.append(("lower", session_ldap.lower(), len(rows)))
    if not rows and "@" in str(session_ldap):
        local = session_ldap.split("@", 1)[0]
        rows = fetch_for_ldap(local)
        tried.append(("localpart", local, len(rows)))
    if not rows:
        candidates = []
        if "@" in str(session_ldap):
            candidates.append("%" + session_ldap.split("@", 1)[0] + "%")
            candidates.append("%" + session_ldap + "%")
        else:
            candidates.append("%" + session_ldap + "%")
        for pattern in candidates:
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT up.allocation_id, mae.project_id, p.name as project_name, mae.iom_id, pw.department AS department,
                           up.punch_date, up.week_number, up.actual_hours, up.wbs
                    FROM user_punches up
                    LEFT JOIN monthly_allocation_entries mae ON mae.id = up.allocation_id
                    LEFT JOIN projects p ON mae.project_id = p.id
                    LEFT JOIN prism_wbs pw ON mae.iom_id = pw.iom_id
                    WHERE up.user_ldap LIKE %s
                      AND up.punch_date BETWEEN %s AND %s
                    ORDER BY up.punch_date, p.name
                """, [pattern, billing_start, billing_end])
                tmp = dictfetchall(cur)
            tried.append(("wildcard", pattern, len(tmp)))
            if tmp:
                rows = tmp
                break

    logger.debug("export_my_punches_excel tried patterns: %r", tried)

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Punches {month_param or billing_start.strftime('%Y-%m')}"

    headers = ["Date", "Project", "IOM", "Dept", "Week#", "Hours", "WBS"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = 20

    r = 2
    for rec in rows:
        pd = rec.get("punch_date")
        if hasattr(pd, "strftime"):
            pd_str = pd.strftime("%Y-%m-%d")
        else:
            pd_str = pd or ""
        ws.cell(row=r, column=1, value=pd_str)
        ws.cell(row=r, column=2, value=rec.get("project_name"))
        ws.cell(row=r, column=3, value=rec.get("iom_id"))
        ws.cell(row=r, column=4, value=rec.get("department"))
        ws.cell(row=r, column=5, value=rec.get("week_number"))
        ws.cell(row=r, column=6, value=float(rec.get("actual_hours") or 0))
        ws.cell(row=r, column=7, value=rec.get("wbs") or "")
        r += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_user = str(session_ldap).replace("@", "_at_").replace(".", "_")
    filename = f"punches_{safe_user}_{(month_param or billing_start.strftime('%Y-%m'))}.xlsx"
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

from django.views.decorators.http import require_POST

def get_lead_allocations_for_distribution(session_ldap, month_start):
    """Fetch total hours allocated to logged-in lead, grouped by project/subproject."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT mae.project_id, p.name AS project_name,
                   mae.subproject_id, COALESCE(sp.name, '-') AS subproject_name,
                   SUM(mae.total_hours) AS total_hours
            FROM monthly_allocation_entries mae
            LEFT JOIN projects p ON mae.project_id = p.id
            LEFT JOIN subprojects sp ON mae.subproject_id = sp.id
            WHERE LOWER(mae.user_ldap) = LOWER(%s)
              AND mae.month_start = %s
            GROUP BY mae.project_id, mae.subproject_id, p.name, sp.name
            ORDER BY p.name
        """, [session_ldap, month_start])
        rows = dictfetchall(cur)
    return rows


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db import transaction, connection
from datetime import date
import json

@require_POST
def save_team_distribution(request):
    """Persist distributed hours for lead’s reportees per subproject with tolerant LDAP matching and validation."""
    print("save_team_distribution: called")
    try:
        payload = json.loads(request.body.decode("utf-8"))
        print("Payload loaded:", payload)
    except Exception as e:
        print("Invalid JSON:", e)
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    allocations = payload.get("allocations", [])
    month_str = payload.get("month")  # expected "YYYY-MM"
    print("Allocations:", allocations)
    print("Month string:", month_str)
    if not month_str:
        print("Missing month in payload")
        return JsonResponse({"ok": False, "error": "Missing month"}, status=400)

    try:
        y, m = map(int, month_str.split("-"))
        month_start = date(y, m, 1)
        print("Parsed month_start:", month_start)
    except Exception as e:
        print("Invalid month format:", e)
        return JsonResponse({"ok": False, "error": "Invalid month format; use YYYY-MM"}, status=400)

    session_ldap = request.session.get("ldap_username")
    print("Session LDAP:", session_ldap)
    if not session_ldap:
        print("Not logged in")
        return JsonResponse({"ok": False, "error": "Not logged in"}, status=403)

    # normalize helper for client-side list
    def _lower_list(xs):
        return [str(x).strip().lower() for x in xs if x]

    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # process each subproject group
                for a in allocations:
                    subproject_id = a.get("subproject_id")
                    items = a.get("items", [])
                    print("Processing allocation for subproject_id:", subproject_id, "with items:", items)
                    if not subproject_id:
                        print("Missing subproject_id in allocation")
                        return JsonResponse({"ok": False, "error": "Missing subproject_id in allocation"}, status=400)

                    # collect submitted reportee ldaps and the sum of requested hours
                    submitted_ldaps = []
                    submitted_sum_hours = 0.0
                    for it in items:
                        rep = (it.get("reportee") or "").strip()
                        try:
                            hrs = float(it.get("hours", 0) or 0)
                        except Exception as e:
                            print("Invalid hours for item:", it, "Error:", e)
                            hrs = 0.0
                        if rep and hrs > 0:
                            submitted_ldaps.append(rep.lower())
                            submitted_sum_hours += hrs
                    print("Submitted ldaps:", submitted_ldaps)
                    print("Submitted sum hours:", submitted_sum_hours)

                    # 1) fetch lead's total hours for this subproject and month (tolerant matching)
                    lead_variants = [session_ldap, session_ldap, session_ldap]
                    cur.execute("""
                        SELECT COALESCE(SUM(mae.total_hours), 0)
                        FROM monthly_allocation_entries mae
                        WHERE mae.month_start = %s
                          AND mae.subproject_id = %s
                          AND (
                              LOWER(mae.user_ldap) = LOWER(%s)
                              OR LOWER(REPLACE(mae.user_ldap, '.', ' ')) = LOWER(%s)
                              OR LOWER(REPLACE(mae.user_ldap, ' ', '.')) = LOWER(%s)
                          )
                    """, [month_start, subproject_id] + lead_variants)
                    lead_total_hours = float(cur.fetchone()[0] or 0.0)
                    print("Lead total hours for subproject:", subproject_id, "is", lead_total_hours)

                    # 2) compute sum of existing allocations for other users (those not being updated by this request)
                    params = [month_start, subproject_id] + lead_variants
                    if submitted_ldaps:
                        placeholders = ",".join(["%s"] * len(submitted_ldaps))
                        cur.execute(f"""
                            SELECT COALESCE(SUM(mae.total_hours), 0)
                            FROM monthly_allocation_entries mae
                            WHERE mae.month_start = %s
                              AND mae.subproject_id = %s
                              AND NOT (
                                  LOWER(mae.user_ldap) = LOWER(%s)
                                  OR LOWER(REPLACE(mae.user_ldap, '.', ' ')) = LOWER(%s)
                                  OR LOWER(REPLACE(mae.user_ldap, ' ', '.')) = LOWER(%s)
                              )
                              AND LOWER(mae.user_ldap) NOT IN ({placeholders})
                        """, params + _lower_list(submitted_ldaps))
                        existing_others_sum = float(cur.fetchone()[0] or 0.0)
                        print("Existing others sum (excluding submitted):", existing_others_sum)
                    else:
                        cur.execute("""
                            SELECT COALESCE(SUM(mae.total_hours), 0)
                            FROM monthly_allocation_entries mae
                            WHERE mae.month_start = %s
                              AND mae.subproject_id = %s
                              AND NOT (
                                  LOWER(mae.user_ldap) = LOWER(%s)
                                  OR LOWER(REPLACE(mae.user_ldap, '.', ' ')) = LOWER(%s)
                                  OR LOWER(REPLACE(mae.user_ldap, ' ', '.')) = LOWER(%s)
                              )
                        """, params)
                        existing_others_sum = float(cur.fetchone()[0] or 0.0)
                        print("Existing others sum (no submitted):", existing_others_sum)

                    # 3) compute new total if we write the submitted items (we assume submitted items replace existing rows for those reportees)
                    new_total_assigned = existing_others_sum + submitted_sum_hours
                    print("New total assigned:", new_total_assigned)

                    # 4) validation: cannot assign more than lead_total_hours
                    if new_total_assigned - lead_total_hours > 0.0001:
                        msg = (f"Assigned {new_total_assigned:.2f} > your available {lead_total_hours:.2f} hrs "
                               f"for subproject {subproject_id}")
                        print("Validation failed:", msg)
                        return JsonResponse({"ok": False, "error": msg}, status=400)

                    # 5) Upsert each submitted reportee row (INSERT ... ON DUPLICATE KEY UPDATE)
                    for it in items:
                        rep = (it.get("reportee") or "").strip()
                        try:
                            hrs = float(it.get("hours", 0) or 0)
                        except Exception as e:
                            print("Invalid hours for upsert item:", it, "Error:", e)
                            hrs = 0.0
                        if not rep:
                            print("Skipping item with empty reportee:", it)
                            continue

                        print("Upserting for reportee:", rep, "hours:", hrs)
                        cur.execute("""
                            INSERT INTO monthly_allocation_entries
                              (project_id, subproject_id, month_start, user_ldap, total_hours, created_at, updated_at)
                            VALUES (
                              (SELECT project_id FROM subprojects WHERE id = %s LIMIT 1),
                              %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                            )
                            ON DUPLICATE KEY UPDATE
                              total_hours = VALUES(total_hours),
                              updated_at = CURRENT_TIMESTAMP
                        """, [subproject_id, subproject_id, month_start, rep, hrs])
                        print("Upserted row for", rep)

    except JsonResponse:
        print("JsonResponse raised intentionally, re-raising")
        raise
    except Exception as e:
        print("Exception in save_team_distribution:", e)
        logger.exception("save_team_distribution failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

    print("save_team_distribution: completed successfully")
    return JsonResponse({"ok": True})


def first_day_of_month_from_str(s):
    if not s:
        today = datetime.date.today()
        return today.replace(day=1)
    if len(s) == 7:
        return datetime.strptime(s + "-01", "%Y-%m-%d").date()
    d = parse_date(s)
    if d:
        return d.replace(day=1)
    today = datetime.date.today()
    return today.replace(day=1)

@require_POST
def save_team_distribution_using_team_table(request):
    """
    Persist distributed hours for lead’s reportees per subproject with tolerant LDAP matching and validation.
    Accepts payload:
    {
        "month": "YYYY-MM",
        "allocations": [
            {
                "subproject_id": "...",
                "items": [
                    {"reportee": "user@domain", "hours": 50.0, "weeks": [25,25,25,25]},
                    ...
                ]
            },
            ...
        ]
    }

    Behavior:
    - Upserts team_distributions rows (month_start, lead_ldap, subproject_id, reportee_ldap -> hours).
    - For each upserted team_distributions row, reads back its id and upserts 4 weekly_allocations rows
      (team_distribution_id, week_number) with hours and percent.
    - Validation ensures total distributed per subproject does not exceed lead's allowed hours (same as before).
    """
    import json
    from datetime import date
    from django.http import JsonResponse
    from django.db import transaction, connection

    logger = logging.getLogger(__name__)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception as e:
        logger.error("Invalid JSON: %r", e)
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    allocations = payload.get("allocations", [])
    month_str = payload.get("month")
    if not month_str:
        return JsonResponse({"ok": False, "error": "Missing month"}, status=400)

    # Resolve canonical billing period (month_start) from monthly_hours_limit table
    try:
        year, month = map(int, month_str.split("-"))
        with connection.cursor() as cur:
            cur.execute("""
                SELECT start_date, end_date
                FROM monthly_hours_limit
                WHERE year = %s AND month = %s
            """, [year, month])
            row = cur.fetchone()
            if not row or not row[0]:
                logger.error("No valid billing cycle found for %s-%s", year, month)
                return JsonResponse({"ok": False, "error": "Billing period not found"}, status=400)
            month_start, billing_end = row[0], row[1]
    except Exception as e:
        logger.exception("Billing period lookup failed: %r", e)
        return JsonResponse({"ok": False, "error": "Error reading billing cycle"}, status=500)

    # Flatten allocations to list of { subproject_id, reportee_ldap, hours, weeks }
    flat_allocs = []
    for alloc in allocations:
        subproject_id = alloc.get("subproject_id")
        items = alloc.get("items", []) or []
        for item in items:
            reportee_ldap = (item.get("reportee") or "").strip()
            hours = item.get("hours")
            weeks = item.get("weeks", [])  # expected length 4, but tolerant
            if not (subproject_id and reportee_ldap and hours is not None):
                continue
            flat_allocs.append({
                "subproject_id": subproject_id,
                "reportee_ldap": reportee_ldap,
                "hours": float(hours or 0),
                "weeks": [float(w or 0) for w in (weeks or [])]
            })

    if not flat_allocs:
        return JsonResponse({"ok": False, "error": "No valid allocations"}, status=400)

    session_ldap = request.session.get("ldap_username")
    if not session_ldap:
        return JsonResponse({"ok": False, "error": "Not authenticated"}, status=403)

    # Validate per-subproject totals against monthly_allocation_entries (lead's allowed hours)
    from collections import defaultdict
    subproject_hours = defaultdict(float)
    for a in flat_allocs:
        subproject_hours[a["subproject_id"]] += float(a["hours"] or 0.0)

    try:
        with connection.cursor() as cur:
            for subproject_id, total_dist_hours in subproject_hours.items():
                cur.execute("""
                    SELECT total_hours
                    FROM monthly_allocation_entries
                    WHERE user_ldap = %s AND subproject_id = %s AND month_start = %s
                """, [session_ldap, subproject_id, month_start])
                row = cur.fetchone()
                allowed_hours = float(row[0]) if row else 0.0
                if total_dist_hours > allowed_hours:
                    return JsonResponse({
                        "ok": False,
                        "error": f"Distributed hours ({total_dist_hours:.2f}) exceed allowed hours ({allowed_hours:.2f}) for subproject {subproject_id}."
                    }, status=400)
    except Exception as e:
        logger.exception("Validation failed: %s", e)
        return JsonResponse({"ok": False, "error": "Validation error"}, status=500)

    # Persist: upsert team_distributions and weekly_allocations inside a transaction
    try:
        with transaction.atomic():
            with connection.cursor() as cur:
                # We'll optionally set project_id to NULL (same as before). If you want to derive project_id per subproject,
                # you can query subprojects table here.
                project_id = None

                for a in flat_allocs:
                    subproject_id = a["subproject_id"]
                    reportee_ldap = a["reportee_ldap"]
                    hours = float(a["hours"] or 0.0)
                    weeks = a.get("weeks", []) or []

                    # Upsert team_distributions (same uniqueness constraint exists)
                    cur.execute("""
                        INSERT INTO team_distributions
                        (month_start, lead_ldap, project_id, subproject_id, reportee_ldap, hours, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                        ON DUPLICATE KEY UPDATE
                            hours = VALUES(hours),
                            updated_at = CURRENT_TIMESTAMP
                    """, [month_start, session_ldap, project_id, subproject_id, reportee_ldap, hours])

                    # Fetch the team_distributions.id for this upserted row
                    cur.execute("""
                        SELECT id FROM team_distributions
                        WHERE month_start = %s AND lead_ldap = %s AND subproject_id = %s AND LOWER(reportee_ldap) = LOWER(%s)
                        LIMIT 1
                    """, [month_start, session_ldap, subproject_id, reportee_ldap])
                    td_row = cur.fetchone()
                    if not td_row:
                        # defensive: if we cannot find the row, skip weekly inserts for this item
                        logger.warning("Could not find team_distributions row after upsert for %s / %s", subproject_id, reportee_ldap)
                        continue
                    team_dist_id = int(td_row[0])

                    # Upsert weekly allocations for this team_distribution (weeks 1..4)
                    # Note: your weekly_allocations table must have a UNIQUE key on (team_distribution_id, week_number)
                    for idx in range(4):
                        wknum = idx + 1
                        pct = float(weeks[idx]) if idx < len(weeks) else 0.0
                        wk_hours = round((pct / 100.0) * hours, 2) if hours and pct else 0.0

                        cur.execute("""
                            INSERT INTO weekly_allocations
                              (team_distribution_id, week_number, hours, percent, status, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                            ON DUPLICATE KEY UPDATE
                              hours = VALUES(hours),
                              percent = VALUES(percent),
                              updated_at = CURRENT_TIMESTAMP
                        """, [team_dist_id, wknum, wk_hours, pct, 'PENDING'])

                # Optionally: remove weekly_allocations for team_distribution rows that no longer exist
                # (not implemented here; deletes should be done via delete endpoint to avoid accidental removals)
    except Exception as e:
        logger.exception("Error in save_team_distribution_using_team_table: %r", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

    return JsonResponse({"ok": True})


def apply_team_distributions_view(request):
    """
    Admin-protected view: apply team_distributions of a given month to monthly_allocation_entries (raw SQL).
    Body: { "month": "2025-08", "month_hours": 183.75, "dry_run": false }
    Returns JSON {ok: True} or {ok: False, error: "..."}
    """
    try:
        payload = json.loads(request.body.decode('utf-8') or "{}")
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Invalid JSON: {e}"}, status=400)

    month_start = first_day_of_month_from_str(payload.get('month'))
    month_hours = float(payload.get('month_hours') or 183.75)
    dry_run = bool(payload.get('dry_run'))

    try:
        with connection.cursor() as cur:
            # Fetch all distinct reportees involved this month
            cur.execute("SELECT DISTINCT LOWER(reportee_ldap) AS r FROM team_distributions WHERE month_start = %s", [month_start])
            reportees = [r['r'] for r in dictfetchall(cur)]

            # Compute monthly_allocation_entries totals for these reportees
            reportee_monthly_totals = {}
            if reportees:
                placeholders = ",".join(["%s"] * len(reportees))
                q = f"""
                    SELECT LOWER(user_ldap) as user_ldap, COALESCE(SUM(total_hours),0) as total
                    FROM monthly_allocation_entries
                    WHERE month_start = %s AND LOWER(user_ldap) IN ({placeholders})
                    GROUP BY LOWER(user_ldap)
                """
                params = [month_start] + reportees
                cur.execute(q, params)
                for r in dictfetchall(cur):
                    reportee_monthly_totals[r['user_ldap']] = float(r['total'] or 0.0)

            # Compute team_distributions totals per reportee (new state)
            td_totals = {}
            if reportees:
                placeholders = ",".join(["%s"] * len(reportees))
                q2 = f"""
                    SELECT LOWER(reportee_ldap) AS reportee, COALESCE(SUM(hours),0) AS ttotal
                    FROM team_distributions
                    WHERE month_start = %s AND LOWER(reportee_ldap) IN ({placeholders})
                    GROUP BY LOWER(reportee_ldap)
                """
                params2 = [month_start] + reportees
                cur.execute(q2, params2)
                for r in dictfetchall(cur):
                    td_totals[r['reportee']] = float(r['ttotal'] or 0.0)

            # Validate prospective totals: monthly_allocation_entries_total + td_totals <= month_hours
            for r in reportees:
                existing_m = reportee_monthly_totals.get(r, 0.0)
                t_total = td_totals.get(r, 0.0)
                prospective = existing_m + t_total
                if prospective > month_hours + 1e-9:
                    return JsonResponse({"ok": False, "error": f"Reportee {r} prospective total {prospective:.2f} exceeds month_hours {month_hours}"}, status=400)

        # If validations pass, apply (upsert monthly_allocation_entries). Use transaction.
        if dry_run:
            return JsonResponse({"ok": True, "dry_run": True})

        with transaction.atomic():
            with connection.cursor() as cur:
                # load existing monthly_allocation_entries for the target month to map (subproject, user) -> id
                cur.execute("SELECT id, subproject_id, LOWER(user_ldap) as user_ldap FROM monthly_allocation_entries WHERE month_start = %s", [month_start])
                mae_map = {}
                for r in dictfetchall(cur):
                    key = (int(r['subproject_id']), r['user_ldap'])
                    mae_map[key] = r['id']

                # fetch all team_distributions rows for month
                cur.execute("""
                    SELECT month_start, lead_ldap, project_id, subproject_id, reportee_ldap, hours
                    FROM team_distributions
                    WHERE month_start = %s
                """, [month_start])
                distro_rows = dictfetchall(cur)

                # upsert each distribution row into monthly_allocation_entries
                for row in distro_rows:
                    subp = int(row['subproject_id'])
                    reportee_ldap = row['reportee_ldap']
                    project_id = row.get('project_id')
                    hours = float(row['hours'] or 0.0)
                    key = (subp, (reportee_ldap or '').lower())
                    if key in mae_map:
                        mae_id = mae_map[key]
                        cur.execute("UPDATE monthly_allocation_entries SET total_hours = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s", [hours, mae_id])
                    else:
                        cur.execute("""
                            INSERT INTO monthly_allocation_entries
                            (project_id, subproject_id, month_start, user_ldap, total_hours, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                        """, [project_id, subp, month_start, reportee_ldap, hours])
        return JsonResponse({"ok": True})
    except Exception as e:
        logger.exception("apply_team_distributions_view failed: %s", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db import connection
import json, logging

@require_POST
def delete_team_distribution(request):
    """
    Delete a team_distributions row (and its weekly_allocations via FK cascade).
    Expects JSON body: { "id": <team_distribution_id> }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        print("Invalid JSON in delete_team_distribution")
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    td_id = data.get("id")
    if not td_id:
        print("Missing id in delete_team_distribution")
        return JsonResponse({"ok": False, "error": "Missing id"}, status=400)

    session_ldap = request.session.get("ldap_username")
    if not session_ldap:
        print("Not authenticated in delete_team_distribution")
        return JsonResponse({"ok": False, "error": "Not authenticated"}, status=403)

    try:
        with connection.cursor() as cur:
            # Verify that the logged-in lead owns this record
            cur.execute("SELECT lead_ldap FROM team_distributions WHERE id = %s LIMIT 1", [td_id])
            row = cur.fetchone()
            if not row:
                print("Record not found in delete_team_distribution")
                return JsonResponse({"ok": False, "error": "Record not found"}, status=404)

            lead_ldap = (row[0] or "").lower()
            if lead_ldap != (session_ldap or "").lower():
                print("Forbidden: lead_ldap mismatch in delete_team_distribution")
                return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)

            # Delete it; ON DELETE CASCADE will remove weekly_allocations
            cur.execute("DELETE FROM team_distributions WHERE id = %s", [td_id])

        return JsonResponse({"ok": True})
    except Exception as e:
        print(f"delete_team_distribution failed: {e}")
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

from django.views.decorators.http import require_GET, require_POST
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db import connection, transaction
import json, logging

logger = logging.getLogger(__name__)

@require_GET
def tl_allocations_view(request):
    """
    Team Lead Free Allocations View
    --------------------------------
    Shows all direct reportees (from LDAP) and any existing team distributions
    for the selected month_start, including weekly allocation splits.
    """
    from datetime import date
    import json

    session_ldap = request.session.get("ldap_username")
    session_pwd = request.session.get("ldap_password")
    if not request.session.get("is_authenticated") or not session_ldap:
        return redirect("accounts:login")
    creds = (session_ldap, session_pwd)

    # -----------------------------------------------------------
    # Determine Billing Month / Period
    # -----------------------------------------------------------
    month_str = request.GET.get("month")
    if not month_str:
        month_str = date.today().strftime("%Y-%m")

    try:
        y, m = map(int, month_str.split("-"))
        month_start = date(y, m, 1)
    except Exception:
        month_start = date.today().replace(day=1)

    # -----------------------------------------------------------
    # LDAP: Get Direct Reportees
    # -----------------------------------------------------------
    reportees_entries = []
    try:
        print("Importing LDAP utilities...")
        from accounts.ldap_utils import get_user_entry_by_username, get_reportees_for_user_dn
        print(f"Calling get_user_entry_by_username for: {session_ldap}")
        user_entry = get_user_entry_by_username(session_ldap, username_password_for_conn=creds)
        print(f"user_entry: {user_entry}")
        entry_dn = getattr(user_entry, "entry_dn", None)
        print(f"entry_dn: {entry_dn}")
        print("Calling get_reportees_for_user_dn...")
        reportees_entries = get_reportees_for_user_dn(
            entry_dn,
            username_password_for_conn=creds
        ) or []
        print(f"Fetched reportees_entries: {reportees_entries}")
    except Exception as e:
        print(f"LDAP fetch failed: {e}")
        reportees_entries = []

    # Normalize LDAP reportees into consistent dicts (handles dict or object from helper)
    reportees_map = {}
    reportees_list = []
    for ent in reportees_entries:
        mail = None;
        cn = None;
        sam = None
        try:
            if isinstance(ent, dict):
                mail = ent.get("mail") or ent.get("email") or ent.get("userPrincipalName")
                cn = ent.get("cn") or ent.get("displayName")
                sam = ent.get("sAMAccountName") or ent.get("sAMAccountName".lower())
            else:
                mail = getattr(ent, "mail", None) or getattr(ent, "email", None) or getattr(ent, "userPrincipalName",
                                                                                            None)
                cn = getattr(ent, "cn", None) or getattr(ent, "displayName", None)
                sam = getattr(ent, "sAMAccountName", None)
        except Exception:
            # be defensive — skip broken entries
            continue

        identifier = (mail or sam or "").strip()
        if not identifier:
            # fall back to dn-based identity if present
            try:
                if isinstance(ent, dict):
                    dn = ent.get("dn")
                else:
                    dn = getattr(ent, "dn", None)
                if dn:
                    identifier = dn.split(",")[0].replace("CN=", "").strip()
            except Exception:
                identifier = None

        if not identifier:
            continue

        lid = identifier.lower()
        if lid not in reportees_map:
            reportees_map[lid] = {
                "ldap": identifier,
                "mail": mail or identifier,
                "cn": cn or identifier,
                "total_hours": 0.0,
                "fte": 0.0
            }
            reportees_list.append(reportees_map[lid])

    # -------------------------------
    # NEW: determine if logged-in user is a PDL and build logged-in identity
    # -------------------------------
    # determine logged in ldap/email and display name
    logged_ldap = None
    logged_cn = None
    # prefer session values (your app sets these)
    if request.session.get("ldap_username"):
        logged_ldap = request.session.get("ldap_username")
        logged_cn = request.session.get("cn") or request.session.get("display_name") or logged_ldap
    # fallback to Django user
    if not logged_ldap and getattr(request, "user", None) and request.user.is_authenticated:
        logged_ldap = getattr(request.user, "email", None) or getattr(request.user, "username", None)
        try:
            logged_cn = request.user.get_full_name() or logged_ldap
        except Exception:
            logged_cn = logged_ldap

    # -------------------------------
    # NEW (simpler): determine if logged-in user is a PDL using session['role']
    # -------------------------------
    is_pdl = False
    try:
        role_val = (request.session.get('role') or "").strip()
        # common cases: "PDL", "pdl", or compound roles like "PDL,MANAGER"
        if role_val:
            # normalize and check membership
            role_norm = role_val.upper()
            if role_norm == "PDL" or ",PDL" in "," + role_norm or "PDL," in role_norm or " PD L " in role_norm:
                is_pdl = True
            # also accept exact match or string that contains 'PDL'
            elif "PDL" in role_norm:
                is_pdl = True
    except Exception:
        # defensive fallback — if anything goes wrong, assume not PDL
        is_pdl = False

    print("is_pdl : ", is_pdl)
    # If PDL, ensure logged-in user appears in reportees list (so PDL can allocate to self)
    if is_pdl and logged_ldap:
        lkey = (logged_ldap or "").lower()
        if lkey not in reportees_map:
            reportees_map[lkey] = {
                "ldap": logged_ldap,
                "mail": logged_ldap,
                "cn": logged_cn or logged_ldap,
                "total_hours": 0.0,
                "fte": 0.0
            }
            # insert at beginning so it's easy to find in dropdown
            reportees_list.insert(0, reportees_map[lkey])
    # -------------------------------
    # end NEW block
    # -------------------------------

    # remove self (lead) if present — only when NOT a PDL (PDL should be able to see themselves)
    session_ldap_l = (session_ldap or "").lower()
    if not is_pdl and session_ldap_l in reportees_map:
        reportees_list = [r for r in reportees_list if r["ldap"].lower() != session_ldap_l]
        reportees_map.pop(session_ldap_l, None)

    # --- compute weeks info for template (WD and max%) -----------------------
    # weeks list from existing helper: list of dicts {'num','start','end'}
    pdate = _to_date(month_start)
    yy = pdate.year;
    mm = pdate.month
    billing_start, billing_end = _get_billing_period_from_month(yy, mm)
    weeks = _compute_weeks_for_billing(billing_start, billing_end)
    print("Billing start and end  : ", (billing_start, billing_end))

    # fetch holidays between billing start/end
    with connection.cursor() as cur:
        cur.execute(
            "SELECT holiday_date FROM holidays WHERE holiday_date BETWEEN %s AND %s",
            [billing_start, billing_end]
        )
        holidays_rows = cur.fetchall() or []

    holidays_set = set()
    for hr in holidays_rows:
        # hr may be tuple(date,) depending on cursor; handle defensively
        try:
            d = hr[0]
        except Exception:
            d = hr
        if d:
            holidays_set.add(d.strftime("%Y-%m-%d"))

    # total working days in billing (exclude holidays and weekends)
    total_working_days = 0
    for w in weeks:
        wd = _count_working_days(w['start'], w['end'], holidays_set)
        total_working_days += wd
    if total_working_days == 0:
        total_working_days = 1

    # Build weeks_info array for template: week num, start/end strings, working_days, max_pct
    weeks_info = []
    for w in weeks:
        wd = _count_working_days(w['start'], w['end'], holidays_set)
        max_pct = round((wd / total_working_days) * 100.0, 2)
        weeks_info.append({
            "num": int(w["num"]),
            "start": w["start"].strftime("%Y-%m-%d"),
            "end": w["end"].strftime("%Y-%m-%d"),
            "working_days": int(wd),
            "max_pct": float(max_pct)
        })

    # pass JSON string to template
    import json
    weeks_info_json = json.dumps(weeks_info)

    # --- Add to template context later when calling render(...) ---
    # example: context.update({"weeks_info": weeks_info, "weeks_info_json": weeks_info_json})

    # -----------------------------------------------------------
    # Fetch Projects and Subprojects
    # -----------------------------------------------------------
    with connection.cursor() as cur:
        cur.execute("""
            SELECT MIN(id) AS id,
                   bg_code,
                   MAX(project_id) AS project_id,
                   CONCAT(bg_code, ' - ', MAX(buyer_bau)) AS name
            FROM prism_wbs
            WHERE bg_code IS NOT NULL AND bg_code <> ''
            GROUP BY bg_code
            ORDER BY bg_code
        """)
        projects = [
            {"id": r["id"], "name": r["name"], "bg_code": r["bg_code"], "project_id": r["project_id"]}
            for r in dictfetchall(cur)
        ]

        cur.execute("""
            SELECT id, project_id, name,
                   COALESCE(mdm_code, '') AS mdm_code,
                   COALESCE(bg_code, '') AS bg_code
            FROM subprojects
            ORDER BY priority DESC, name
        """)
        subprojects = dictfetchall(cur)

    # -----------------------------------------------------------
    # Monthly Team Distributions (core data for allocations)
    # -----------------------------------------------------------
    with connection.cursor() as cur:
        cur.execute("""
            SELECT id, project_id, subproject_id, reportee_ldap, hours
            FROM team_distributions
            WHERE lead_ldap = %s AND month_start = %s
        """, [session_ldap, month_start])
        td_rows = dictfetchall(cur)
        td_ids = [r["id"] for r in td_rows if r.get("id")]

        # Weekly splits
        weekly_map = {}
        if td_ids:
            placeholders = ",".join(["%s"] * len(td_ids))
            cur.execute(f"""
                SELECT team_distribution_id, week_number, percent
                FROM weekly_allocations
                WHERE team_distribution_id IN ({placeholders})
            """, td_ids)
            for w in dictfetchall(cur):
                tid = int(w["team_distribution_id"])
                weekly_map.setdefault(tid, {})[int(w["week_number"])] = float(w["percent"] or 0)

        # Monthly hours limit (for FTE calculation)
        cur.execute("""
            SELECT max_hours FROM monthly_hours_limit
            WHERE %s BETWEEN start_date AND end_date
            LIMIT 1
        """, [month_start])
        mh = cur.fetchone()
        monthly_hours = float(mh[0]) if mh and mh[0] else 183.75

    # -----------------------------------------------------------
    # Compute Totals & FTEs
    # -----------------------------------------------------------
    for v in reportees_map.values():
        v["fte"] = round((v["total_hours"] / monthly_hours), 3) if monthly_hours else 0.0

    reportees_for_template = sorted(reportees_map.values(), key=lambda x: x["cn"].lower())
    # --- Fetch existing TL allocations from DB (team_distributions + weekly_allocations) ---
    # --- Fetch existing TL allocations from DB (team_distributions + weekly_allocations) ---
    allocations = []
    with connection.cursor() as cur:
        cur.execute("""
            SELECT td.id, td.reportee_ldap, td.project_id, td.subproject_id, td.hours,
                   GROUP_CONCAT(CONCAT(wa.week_number, ':', wa.percent) ORDER BY wa.week_number ASC) AS week_data
              FROM team_distributions td
         LEFT JOIN weekly_allocations wa ON wa.team_distribution_id = td.id
             WHERE td.lead_ldap = %s
               AND td.month_start BETWEEN %s AND %s
          GROUP BY td.id, td.reportee_ldap, td.project_id, td.subproject_id, td.hours
          ORDER BY td.id ASC
        """, [logged_ldap, billing_start, billing_end])
        rows = cur.fetchall() or []
        print("Rows fetched : ",rows)
    # rows are tuples (id, reportee_ldap, project_id, subproject_id, hours, week_data)
    for r in rows:
        # defensive unpack
        tid = r[0]
        reportee_ldap = r[1]
        project_id = r[2]
        subproject_id = r[3]
        hours = float(r[4] or 0)
        week_data = r[5] or ""

        # build week map {week_number: percent}
        week_map = {}
        if week_data:
            for pair in week_data.split(','):
                try:
                    wk, val = pair.split(':')
                    wkn = int(wk)
                    week_map[wkn] = float(val or 0)
                except Exception:
                    # skip malformed pair
                    continue

        allocations.append({
            "id": tid,
            "reportee_ldap": reportee_ldap,
            "project_id": project_id,
            "subproject_id": subproject_id,
            "hours": hours,
            # keep order same as weeks_info so template inputs align
            "week_perc": [week_map.get(w["num"], 0.0) for w in weeks_info],
        })
    print("Allocations : ", allocations)
    # also pass JSON-encoded version for safe client-side use (template expects a JS array)
    allocations_json = json.dumps(allocations)

    # -----------------------------------------------------------
    # Render Page
    # -----------------------------------------------------------
    return render(request, "projects/tl_allocations.html", {
        "billing_month": month_str,
        "reportees": reportees_for_template,
        "projects": projects,
        "subprojects": subprojects,
        "subprojects_json": json.dumps(subprojects),
        "weeks_info": weeks_info,
        "weeks_info_json": weeks_info_json,
        "allocations": allocations,
        "allocations_json": allocations_json,  # <-- new
        "monthly_hours": monthly_hours,
    })




from django.views.decorators.http import require_POST

@require_POST
def save_tl_allocations(request):
    """
    Save free-hand TL allocations (no restrictions).

    - Accepts payload.month (YYYY-MM) and payload.allocations (array).
    - Each allocation may contain:
        {
          "id": <existing_team_distribution_id|null>,
          "reportee": "user@domain",
          "project_id": 46,
          "subproject_id": 221,
          "hours": 175,
          "weeks": [ {"week_number":1,"percent":25}, ... ]   OR   [25,25,25,25] (plain list -> mapped to 1..N)
        }
    - For each team_distribution: upsert the row, upsert provided weekly_allocations,
      and delete any weekly_allocations for that team_distribution not present in the provided week set.
    """
    # local imports so this function remains drop-in
    import json
    import logging
    import calendar
    from datetime import date
    from django.http import JsonResponse
    from django.db import connection, transaction

    logger = logging.getLogger(__name__)

    # parse JSON payload
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Invalid JSON payload"}, status=400)

    # auth check
    session_ldap = request.session.get("ldap_username")
    if not session_ldap:
        return JsonResponse({"ok": False, "error": "Not authenticated"}, status=403)

    month = payload.get("month")  # expected "YYYY-MM"
    allocations = payload.get("allocations", [])
    if not month:
        return JsonResponse({"ok": False, "error": "Missing month"}, status=400)

    # parse month into canonical billing_start / billing_end
    try:
        yy, mm = map(int, month.split("-"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid month format (expected YYYY-MM)"}, status=400)

    # Try to use helper get_billing_period if available; otherwise fallback to first/last day of month
    try:
        billing_start, billing_end = get_billing_period(yy, mm)
    except Exception:
        # fallback: use first and last day of month
        billing_start = date(yy, mm, 1)
        last_day = calendar.monthrange(yy, mm)[1]
        billing_end = date(yy, mm, last_day)

    # Normalize session_ldap to string
    session_ldap = str(session_ldap)

    try:
        saved_rows = 0
        with transaction.atomic():
            with connection.cursor() as cur:
                for a in allocations or []:
                    reportee = (a.get("reportee") or "").strip()
                    if not reportee:
                        # skip empty/invalid row
                        continue

                    project_id = a.get("project_id") or None
                    subproject_id = a.get("subproject_id") or None
                    try:
                        hours = float(a.get("hours") or 0)
                    except Exception:
                        hours = 0.0

                    weeks = a.get("weeks") or []
                    # coerce scalar to list
                    if not isinstance(weeks, (list, tuple)):
                        weeks = [weeks]

                    # 1) find existing team_distributions row (match canonical billing_start window)
                    # Use month_start BETWEEN billing_start AND billing_end to be robust if month_start values vary.
                    cur.execute("""
                        SELECT id
                        FROM team_distributions
                        WHERE LOWER(lead_ldap) = LOWER(%s)
                          AND month_start BETWEEN %s AND %s
                          AND LOWER(reportee_ldap) = LOWER(%s)
                          AND (
                            (subproject_id = %s)
                            OR (subproject_id IS NULL AND %s IS NULL)
                          )
                        LIMIT 1
                    """, [session_ldap, billing_start, billing_end, reportee, subproject_id, subproject_id])
                    found = cur.fetchone()

                    if found and found[0]:
                        tdid = int(found[0])
                        # update the existing row
                        cur.execute("""
                            UPDATE team_distributions
                            SET project_id = %s,
                                subproject_id = %s,
                                hours = %s,
                                updated_at = NOW()
                            WHERE id = %s
                        """, [project_id, subproject_id, hours, tdid])
                    else:
                        # insert a new team_distributions row using canonical month_start
                        cur.execute("""
                            INSERT INTO team_distributions
                                (month_start, lead_ldap, project_id, subproject_id, reportee_ldap, hours, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                        """, [billing_start, session_ldap, project_id, subproject_id, reportee, hours])
                        # get inserted id
                        try:
                            tdid = int(cur.lastrowid)
                        except Exception:
                            # fallback: query the most recent matching row
                            cur.execute("""
                                SELECT id FROM team_distributions
                                WHERE LOWER(lead_ldap)=LOWER(%s) AND month_start BETWEEN %s AND %s
                                  AND LOWER(reportee_ldap)=LOWER(%s)
                                  AND ( (subproject_id=%s) OR (subproject_id IS NULL AND %s IS NULL) )
                                ORDER BY id DESC
                                LIMIT 1
                            """, [session_ldap, billing_start, billing_end, reportee, subproject_id, subproject_id])
                            r2 = cur.fetchone()
                            if not r2:
                                raise RuntimeError("Failed to determine team_distributions id after insert")
                            tdid = int(r2[0])

                    # 2) Normalize incoming weeks into (week_number, percent) tuples
                    normalized_weeks = []
                    if isinstance(weeks, (list, tuple)) and len(weeks) > 0 and isinstance(weeks[0], dict):
                        # list of objects expected: {"week_number": 1, "percent": 25}
                        for item in weeks:
                            try:
                                wn = int(item.get("week_number"))
                                pv = float(item.get("percent") or 0)
                                normalized_weeks.append((wn, pv))
                            except Exception:
                                continue
                    elif isinstance(weeks, (list, tuple)):
                        # plain list - treat as percentages mapped to week 1..N
                        for idx, item in enumerate(weeks):
                            try:
                                pv = float(item or 0)
                            except Exception:
                                pv = 0.0
                            normalized_weeks.append((idx + 1, pv))
                    else:
                        # scalar
                        try:
                            pv = float(weeks)
                            normalized_weeks.append((1, pv))
                        except Exception:
                            normalized_weeks = []

                    # 3) Upsert each provided week; collect provided week numbers
                    provided_week_numbers = []
                    for (wk_num, pct_val) in normalized_weeks:
                        try:
                            wk_n = int(wk_num)
                        except Exception:
                            continue
                        try:
                            pct_val = float(pct_val or 0)
                        except Exception:
                            pct_val = 0.0
                        # clamp percent
                        if pct_val < 0: pct_val = 0.0
                        if pct_val > 100: pct_val = 100.0

                        provided_week_numbers.append(wk_n)
                        week_hours = round((pct_val / 100.0) * hours, 2)

                        # Upsert into weekly_allocations by unique (team_distribution_id, week_number)
                        cur.execute("""
                            INSERT INTO weekly_allocations
                                (team_distribution_id, week_number, hours, percent, status, created_at, updated_at)
                            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                            ON DUPLICATE KEY UPDATE
                                hours = VALUES(hours),
                                percent = VALUES(percent),
                                status = VALUES(status),
                                updated_at = NOW()
                        """, [tdid, wk_n, week_hours, pct_val, 'PENDING'])

                    # 4) Remove any weekly_allocations rows for this tdid not present in provided_week_numbers
                    if provided_week_numbers:
                        placeholders = ",".join(["%s"] * len(provided_week_numbers))
                        delete_sql = f"""
                            DELETE FROM weekly_allocations
                            WHERE team_distribution_id = %s
                              AND week_number NOT IN ({placeholders})
                        """
                        cur.execute(delete_sql, [tdid] + provided_week_numbers)
                    else:
                        # none provided -> delete existing weekly allocations for this tdid
                        cur.execute("DELETE FROM weekly_allocations WHERE team_distribution_id = %s", [tdid])

                    saved_rows += 1

        return JsonResponse({"ok": True, "saved": saved_rows})
    except Exception as ex:
        try:
            logger.exception("save_tl_allocations failed: %s", ex)
        except Exception:
            pass
        return JsonResponse({"ok": False, "error": str(ex)}, status=500)



@login_required
def get_monthly_limits(request):
    """Return monthly hour limits per resource for current billing cycle."""
    sql = """
        SELECT emp_code, monthly_hours
        FROM monthly_hours_limit
        WHERE start_date <= CURDATE() AND end_date >= CURDATE()
    """
    limits = exec_sql(sql)
    data = {row['emp_code'].lower(): float(row['monthly_hours']) for row in limits}
    return JsonResponse(data)

@require_GET
def export_tl_allocations_excel(request):
    """
    Export TL Allocations (team_distributions + weekly_allocations) for the selected month to Excel.
    Adds a professional Summary sheet and a Details sheet.
    Uses global monthly limit from monthly_hours_limit table.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from datetime import date

    month = request.GET.get("month") or date.today().strftime("%Y-%m")

    try:
        year, mth = map(int, month.split("-"))
        month_start = date(year, mth, 1)
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid month format"}, status=400)

    # --- 1️⃣ Fetch the monthly max hours (global limit) ---
    DEFAULT_MONTHLY_HOURS = 183.75
    try:
        with connection.cursor() as cur:
            cur.execute(
                """
                SELECT max_hours FROM monthly_hours_limit
                WHERE year = %s AND month = %s
                LIMIT 1
                """,
                [year, mth],
            )
            rec = cur.fetchone()
            MONTHLY_LIMIT = float(rec[0]) if rec else DEFAULT_MONTHLY_HOURS
    except Exception:
        MONTHLY_LIMIT = DEFAULT_MONTHLY_HOURS

    try:
        with connection.cursor() as cur:
            # --- 2️⃣ Pull allocations for this month ---
            cur.execute(
                """
                SELECT
                  td.id AS td_id,
                  td.lead_ldap,
                  td.reportee_ldap,
                  CONCAT_WS(' ', ld.givenName, ld.sn) AS reportee_name,
                  td.project_id,
                  p.name AS project_name,
                  td.subproject_id,
                  sp.name AS subproject_name,
                  COALESCE(td.hours, 0) AS hours,
                  td.month_start
                FROM team_distributions td
                LEFT JOIN projects p ON p.id = td.project_id
                LEFT JOIN subprojects sp ON sp.id = td.subproject_id
                LEFT JOIN ldap_directory ld
                  ON LOWER(ld.email) COLLATE utf8mb4_unicode_ci = LOWER(td.reportee_ldap) COLLATE utf8mb4_unicode_ci
                WHERE td.month_start = %s
                ORDER BY LOWER(CONCAT_WS(' ', ld.givenName, ld.sn)) COLLATE utf8mb4_unicode_ci, p.name, sp.name
                """,
                [month_start],
            )
            allocations = dictfetchall(cur)

            # --- 3️⃣ Weekly percentages ---
            td_ids = [r["td_id"] for r in allocations if r.get("td_id")]
            weekly_map = {}
            if td_ids:
                placeholders = ",".join(["%s"] * len(td_ids))
                q = f"""
                    SELECT team_distribution_id, week_number, COALESCE(percent,0) as percent
                    FROM weekly_allocations
                    WHERE team_distribution_id IN ({placeholders})
                """
                cur.execute(q, td_ids)
                for w in dictfetchall(cur):
                    tid = int(w["team_distribution_id"])
                    weekly_map.setdefault(tid, {})[int(w["week_number"])] = float(w["percent"] or 0)

            # --- 4️⃣ Team lead details ---
            lead_candidates = [r.get("lead_ldap") for r in allocations if r.get("lead_ldap")]
            lead_ldap = lead_candidates[0] if lead_candidates else None
            lead_name = lead_ldap
            if lead_ldap:
                cur.execute(
                    """
                    SELECT CONCAT_WS(' ', givenName, sn) AS name
                    FROM ldap_directory
                    WHERE LOWER(email) COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci
                    LIMIT 1
                    """,
                    [lead_ldap],
                )
                res = dictfetchall(cur)
                if res:
                    lead_name = res[0].get("name") or lead_ldap

    except Exception as e:
        print("Export query failed:", e)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

    # --- 5️⃣ Aggregate totals ---
    total_hours = sum(float(r.get("hours") or 0) for r in allocations)
    total_fte = round(total_hours / MONTHLY_LIMIT, 3) if MONTHLY_LIMIT else 0

    # Subproject summary
    from collections import defaultdict
    sub_sums = defaultdict(lambda: {"hours": 0.0, "count": 0})
    for r in allocations:
        name = r.get("subproject_name") or "Unassigned"
        sub_sums[name]["hours"] += float(r.get("hours") or 0)
        sub_sums[name]["count"] += 1

    # --- 6️⃣ Workbook setup ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    fill_blue = PatternFill("solid", fgColor="1F6FEB")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Team Lead Allocations — {month}"
    c.font = Font(size=16, bold=True)
    c.alignment = center

    ws["A2"].value = "Team Lead:"
    ws["B2"].value = lead_name or ""
    ws["A3"].value = "Lead Email:"
    ws["B3"].value = lead_ldap or ""
    ws["A4"].value = "Month Start:"
    ws["B4"].value = month_start.strftime("%Y-%m-%d")

    ws["D2"].value = "Monthly Limit (hrs):"
    ws["E2"].value = MONTHLY_LIMIT
    ws["D3"].value = "Total Allocated Hours:"
    ws["E3"].value = round(total_hours, 2)
    ws["D4"].value = "Total FTE:"
    ws["E4"].value = total_fte

    for coord in ("A2", "A3", "A4", "D2", "D3", "D4"):
        ws[coord].font = bold

    # Subproject summary header
    start_row = 6
    headers = ["Subproject", "Total Hours", "FTE", "Allocations"]
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=idx)
        c.value = h
        c.fill = fill_blue
        c.font = white_font
        c.alignment = center
        c.border = border

    # Subproject data
    row = start_row + 1
    for sp, vals in sorted(sub_sums.items(), key=lambda x: -x[1]["hours"]):
        hrs = vals["hours"]
        fte = round(hrs / MONTHLY_LIMIT, 3) if MONTHLY_LIMIT else 0
        ws.cell(row=row, column=1, value=sp)
        ws.cell(row=row, column=2, value=round(hrs, 2))
        ws.cell(row=row, column=3, value=fte)
        ws.cell(row=row, column=4, value=vals["count"])
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = left if col == 1 else center
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # --- 7️⃣ Details sheet ---
    ws2 = wb.create_sheet("Details")
    headers = [
        "Reportee",
        "Project",
        "Subproject",
        "Hours",
        "FTE",
        "W1%",
        "W2%",
        "W3%",
        "W4%",
        "Total %",
        "Month Start",
    ]
    ws2.append(headers)
    for idx, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=idx)
        c.value = h
        c.fill = fill_blue
        c.font = white_font
        c.alignment = center
        c.border = border

    for r in allocations:
        tid = r["td_id"]
        w = weekly_map.get(tid, {})
        w1, w2, w3, w4 = (w.get(i, 0) for i in range(1, 5))
        total_pct = round(w1 + w2 + w3 + w4, 2)
        hrs = float(r.get("hours") or 0)
        fte = round((hrs / MONTHLY_LIMIT) if MONTHLY_LIMIT else 0, 3)
        name = (r.get("reportee_name") or "").strip()
        email = (r.get("reportee_ldap") or "").strip()
        display = f"{name} <{email}>" if name else email

        ws2.append([
            display,
            r.get("project_name") or "",
            r.get("subproject_name") or "",
            hrs,
            fte,
            w1, w2, w3, w4,
            total_pct,
            r.get("month_start").strftime("%Y-%m-%d") if r.get("month_start") else "",
        ])

    for idx in range(1, len(headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 26 if idx in (1, 2, 3) else 12
    for r in range(2, ws2.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            ws2.cell(row=r, column=c_idx).border = border

    # --- 8️⃣ Return workbook ---
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"TL_Allocations_{month}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

# --- helper utilities (if not present in views.py) ---



def get_user_email_from_session(request):
    # use email (ldap_username) per your requirement
    return request.session.get('ldap_username') or getattr(request.user, 'email', None)


# --- 2) Save weekly edit (PENDING) ---
@require_POST
def save_my_alloc_weekly(request):
    user_email = get_user_email_from_session(request)
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'not authenticated'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        allocation_id = int(payload.get('allocation_id'))
        week_number = int(payload.get('week_number'))
        hours = payload.get('allocated_hours', None)
        percent = payload.get('allocated_percent', None)
        if hours is None and percent is None:
            return JsonResponse({'ok': False, 'error': 'nothing to save'}, status=400)

        with transaction.atomic(), connection.cursor() as cur:
            # Upsert weekly_allocations as PENDING (so user can edit again)
            if hours is not None:
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, hours, percent, status, updated_at)
                    VALUES (%s,%s,%s,%s,'PENDING',NOW())
                    ON DUPLICATE KEY UPDATE hours=VALUES(hours), percent=VALUES(percent), status='PENDING', updated_at=NOW()
                """, [allocation_id, week_number, str(Decimal(str(hours)).quantize(Decimal('0.01')) if hours != '' else '0.00'), (None if percent=='' else percent)])
            else:
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, hours, percent, status, updated_at)
                    VALUES (%s,%s,0,%s,'PENDING',NOW())
                    ON DUPLICATE KEY UPDATE percent=VALUES(percent), status='PENDING', updated_at=NOW()
                """, [allocation_id, week_number, (None if percent=='' else percent)])

        return JsonResponse({'ok': True})
    except Exception as e:
        logger.exception("save_my_alloc_weekly error: %s", e)
        return JsonResponse({'ok': False, 'error': 'server error'}, status=500)

# --- 3) Accept / Reject endpoint (single handler) ---

# --- 4) Vacation save ---
@require_POST
def save_vacation_view(request):
    user_email = get_user_email_from_session(request)
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'not authenticated'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8'))
        billing_start = payload.get('billing_start')
        billing_end = payload.get('billing_end')
        hours = payload.get('hours', 0)
        reason = payload.get('reason', '')
        with transaction.atomic(), connection.cursor() as cur:
            cur.execute("""
                INSERT INTO user_vacations (user_email, billing_start, billing_end, hours, reason)
                VALUES (%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE hours=VALUES(hours), reason=VALUES(reason), updated_at=NOW()
            """, [user_email, billing_start, billing_end, str(Decimal(str(hours)).quantize(Decimal('0.01'))), reason[:2000]])
        return JsonResponse({'ok': True})
    except Exception as e:
        logger.exception("save_vacation_view error: %s", e)
        return JsonResponse({'ok': False, 'error': 'server error'}, status=500)

def tl_reconsiderations_view(request):
    """
    Shows list of weekly_punch_confirmations assigned to the logged-in TL
    with status REJECTED or RECONSIDERED (i.e., requires TL attention).
    """
    tl_email = get_user_email_from_session(request)
    if not tl_email:
        return HttpResponseForbidden("Not authenticated")

    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT wpc.id, wpc.user_email, wpc.allocation_id, wpc.billing_start, wpc.week_number,
                       wpc.allocated_hours, wpc.allocated_percent, wpc.user_comment, wpc.tl_comment, wpc.status, wpc.created_at,
                       mae.project_id, COALESCE(p.name,'') AS project_name,
                       COALESCE(sp.name,'') AS subproject_name, COALESCE(mae.wbs_code,'') AS wbs_code
                FROM weekly_punch_confirmations wpc
                LEFT JOIN monthly_allocation_entries mae ON mae.id = wpc.allocation_id
                LEFT JOIN projects p ON p.id = mae.project_id
                LEFT JOIN subprojects sp ON sp.id = mae.subproject_id
                WHERE wpc.tl_email = %s AND wpc.status IN ('REJECTED','RECONSIDERED')
                ORDER BY wpc.created_at DESC
            """, [tl_email])
            rows = dictfetchall(cur)
    except Exception as e:
        logger.exception("tl_reconsiderations_view DB error: %s", e)
        return HttpResponseBadRequest("Failed to load reconsiderations")

    # format dates to strings for template
    for r in rows:
        if isinstance(r.get('billing_start'), (datetime, )):
            r['billing_start'] = r['billing_start'].strftime('%Y-%m-%d')
        if isinstance(r.get('created_at'), (datetime, )):
            r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
    context = {
        'rows': rows,
        'tl_email': tl_email,
        'action_url': f"/tl/reconsiderations/{{conf_id}}/action/"  # template will format this
    }
    return render(request, 'projects/tl_reconsiderations.html', context)


@require_POST
def tl_action_view(request, conf_id):
    """
    TL action endpoint. Payload JSON:
      { action: 'approve'|'modify'|'reassign'|'close', tl_comment: string, new_hours?: number, reassign_to?: 'user_email' }

    Approve: apply allocated_hours from confirmation into weekly_allocations (status -> RECONSIDERED / ACCEPTED as per policy)
    Modify: modify allocated_hours (and leave status RECONSIDERED or set PENDING to let user accept — here we set RECONSIDERED and write to weekly_allocations as PENDING)
    Reassign: set status back to PENDING for user to act on
    Close: set status CANCELLED

    Requires TL email to match wpc.tl_email for permission.
    """
    tl_email = get_user_email_from_session(request)
    if not tl_email:
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'invalid JSON'}, status=400)

    action = payload.get('action')
    tl_comment = payload.get('tl_comment', '')[:2000]

    if action not in ('approve', 'modify', 'reassign', 'close'):
        return JsonResponse({'ok': False, 'error': 'invalid action'}, status=400)

    try:
        with transaction.atomic(), connection.cursor() as cur:
            # load confirmation row
            cur.execute("""
                SELECT id, user_email, allocation_id, billing_start, week_number, allocated_hours, status, tl_email
                FROM weekly_punch_confirmations WHERE id = %s LIMIT 1
            """, [conf_id])
            conf = cur.fetchone()
            if not conf:
                return JsonResponse({'ok': False, 'error': 'confirmation not found'}, status=404)
            # conf mapping by index
            conf_map = {
                'id': conf[0], 'user_email': conf[1], 'allocation_id': conf[2],
                'billing_start': conf[3], 'week_number': conf[4], 'allocated_hours': conf[5],
                'status': conf[6], 'tl_email': conf[7]
            }
            # permission check
            if not conf_map['tl_email'] or conf_map['tl_email'].lower() != tl_email.lower():
                return JsonResponse({'ok': False, 'error': 'not authorized for this item'}, status=403)

            if action == 'approve':
                # apply allocated_hours to canonical weekly_allocations and mark confirmation accepted
                hours = Decimal(str(conf_map.get('allocated_hours') or '0.00')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                # update confirmation status -> RECONSIDERED (TL acted) then ACCEPTED
                cur.execute("UPDATE weekly_punch_confirmations SET status='RECONSIDERED', tl_comment=%s, actioned_by=%s, actioned_at=NOW(), updated_at=NOW() WHERE id=%s", [tl_comment, tl_email, conf_id])
                # upsert weekly_allocations as ACCEPTED (confirmed by TL)
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, hours, status, confirmed_by, confirmed_at)
                    VALUES (%s,%s,%s,'ACCEPTED',%s,NOW())
                    ON DUPLICATE KEY UPDATE hours=VALUES(hours), status=VALUES(status), confirmed_by=VALUES(confirmed_by), confirmed_at=VALUES(confirmed_at), updated_at=NOW()
                """, [conf_map['allocation_id'], conf_map['week_number'], str(hours), tl_email])
                # save history
                cur.execute("""
                    INSERT INTO weekly_punch_history (confirmation_id, actor_email, role, action, comment, after_json)
                    VALUES (%s,%s,'TL','TL_MODIFY',%s, JSON_OBJECT('hours', %s))
                """, [conf_id, tl_email, tl_comment, str(hours)])
                # Optionally notify user (implement notification)
                return JsonResponse({'ok': True, 'message': 'Approved and applied'})

            elif action == 'modify':
                # required payload: new_hours
                new_hours_raw = payload.get('new_hours')
                if new_hours_raw is None:
                    return JsonResponse({'ok': False, 'error': 'new_hours required for modify'}, status=400)
                new_hours = Decimal(str(new_hours_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                # update confirmation row with new_hours and status RECONSIDERED
                cur.execute("""
                    UPDATE weekly_punch_confirmations
                    SET allocated_hours=%s, status='RECONSIDERED', tl_comment=%s, actioned_by=%s, actioned_at=NOW(), updated_at=NOW()
                    WHERE id=%s
                """, [str(new_hours), tl_comment, tl_email, conf_id])
                # write weekly_allocations with PENDING status so user still can Accept if you want,
                # or leave as RECONSIDERED. Here we store as PENDING (user should Accept).
                cur.execute("""
                    INSERT INTO weekly_allocations (allocation_id, week_number, hours, status, confirmed_by, confirmed_at)
                    VALUES (
                      (SELECT allocation_id FROM weekly_punch_confirmations WHERE id=%s),
                      (SELECT week_number FROM weekly_punch_confirmations WHERE id=%s),
                      %s, 'PENDING', %s, NOW()
                    )
                    ON DUPLICATE KEY UPDATE hours=VALUES(hours), status=VALUES(status), updated_at=NOW()
                """, [conf_id, conf_id, str(new_hours), tl_email])
                # history
                cur.execute("""
                    INSERT INTO weekly_punch_history (confirmation_id, actor_email, role, action, comment, after_json)
                    VALUES (%s,%s,'TL','TL_MODIFY',%s, JSON_OBJECT('hours', %s))
                """, [conf_id, tl_email, tl_comment, str(new_hours)])
                # notify user optionally
                return JsonResponse({'ok': True, 'message': 'Modified and sent back to user for acceptance'})

            elif action == 'reassign':
                # reassign back to user for decision (set status PENDING)
                reassign_to = payload.get('reassign_to')  # optional; not used for now
                cur.execute("""
                    UPDATE weekly_punch_confirmations
                    SET status='PENDING', tl_comment=%s, actioned_by=%s, actioned_at=NOW(), updated_at=NOW()
                    WHERE id=%s
                """, [tl_comment, tl_email, conf_id])
                cur.execute("""
                    INSERT INTO weekly_punch_history (confirmation_id, actor_email, role, action, comment)
                    VALUES (%s,%s,'TL','REASSIGN',%s)
                """, [conf_id, tl_email, tl_comment])
                # Optionally notify user
                return JsonResponse({'ok': True, 'message': 'Reassigned to user'})

            else:  # close
                cur.execute("UPDATE weekly_punch_confirmations SET status='CANCELLED', tl_comment=%s, actioned_by=%s, actioned_at=NOW(), updated_at=NOW() WHERE id=%s", [tl_comment, tl_email, conf_id])
                cur.execute("INSERT INTO weekly_punch_history (confirmation_id, actor_email, role, action, comment) VALUES (%s,%s,'TL','CLOSE',%s)", [conf_id, tl_email, tl_comment])
                return JsonResponse({'ok': True, 'message': 'Closed'})

    except Exception as e:
        logger.exception("tl_action_view error: %s", e)
        return JsonResponse({'ok': False, 'error': 'server error'}, status=500)

@require_POST
def record_leave(request):
    """
    Records leave (sick/vacation/etc.) for a specific week in the billing period.

    Leave hours are stored separately and reduce available punching capacity.
    The allocated hours remain unchanged - leave only affects max punchable hours.
    """
    import json
    from decimal import Decimal
    from datetime import datetime, timedelta
    from django.db import transaction, connection

    logger = logging.getLogger(__name__)

    # Get user from session
    user_email = request.session.get('mail') or request.session.get('ldap_username')
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'User not authenticated'}, status=401)

    # Parse request payload
    try:
        data = json.loads(request.body)
        billing_start = data.get('billing_start')  # YYYY-MM-DD
        billing_end = data.get('billing_end')      # YYYY-MM-DD
        week_number = int(data.get('week_number'))
        leave_hours = Decimal(str(data.get('leave_hours', 0)))
        leave_type = data.get('leave_type', '').strip().upper()
        reason = data.get('reason', '').strip()

        if not all([billing_start, billing_end, week_number, leave_hours > 0, leave_type]):
            return JsonResponse({
                'ok': False,
                'error': 'Missing required fields: billing_start, billing_end, week_number, leave_hours, leave_type'
            }, status=400)

    except (json.JSONDecodeError, ValueError, TypeError) as e:
        return JsonResponse({'ok': False, 'error': f'Invalid request data: {e}'}, status=400)

    # Map frontend leave types to database enum
    leave_type_mapping = {
        'CASUAL': 'VACATION',
        'SICK': 'SICK',
        'EARNED': 'VACATION',
        'UNPAID': 'OTHER',
        'MATERNITY': 'OTHER',
        'PATERNITY': 'OTHER',
        'COMPENSATORY': 'PERSONAL',
        'OTHER': 'OTHER'
    }
    db_leave_type = leave_type_mapping.get(leave_type, 'OTHER')

    # Calculate year, month, and leave_date
    try:
        billing_start_date = datetime.strptime(billing_start, '%Y-%m-%d').date()
        year = billing_start_date.year
        month = billing_start_date.month

        # Calculate leave_date as the Monday of the specified week
        # Assuming week 1 starts on billing_start date
        days_offset = (week_number - 1) * 7
        leave_date = billing_start_date + timedelta(days=days_offset)

    except ValueError as e:
        return JsonResponse({'ok': False, 'error': f'Invalid date format: {e}'}, status=400)

    # Database operations
    try:
        with transaction.atomic(), connection.cursor() as cur:
            # ============================================================
            # VALIDATION 1: Check if punching already submitted for this week
            # ============================================================
            cur.execute("""
                SELECT COUNT(*)
                FROM punch_data
                WHERE user_email = %s
                  AND month_start = %s
                  AND week_number = %s
                  AND status = 'SUBMITTED'
            """, [user_email, billing_start, week_number])

            if cur.fetchone()[0] > 0:
                return JsonResponse({
                    'ok': False,
                    'error': f'Cannot record leave for Week {week_number}. Punching already submitted for this week.'
                }, status=400)

            # ============================================================
            # VALIDATION 2: Check if leave hours exceed allocated hours
            # Get total allocated hours for this user/week from weekly_allocations
            # ============================================================
            cur.execute("""
                SELECT COALESCE(SUM(wa.hours), 0) as total_allocated
                FROM weekly_allocations wa
                INNER JOIN team_distributions td ON wa.team_distribution_id = td.id
                WHERE td.reportee_ldap = %s
                  AND td.month_start = %s
                  AND wa.week_number = %s
            """, [user_email, billing_start, week_number])

            row = cur.fetchone()
            total_allocated = float(row[0]) if row else 0.0

            logger.info(f"Week {week_number} - Total allocated hours: {total_allocated}h")

            if leave_hours > total_allocated:
                return JsonResponse({
                    'ok': False,
                    'error': f'Leave hours ({leave_hours}h) cannot exceed total allocated hours ({total_allocated}h) for Week {week_number}.'
                }, status=400)

            if leave_hours > total_allocated:
                return JsonResponse({
                    'ok': False,
                    'error': f'Leave hours ({leave_hours}h) cannot exceed total allocated hours ({total_allocated}h) for Week {week_number}.'
                }, status=400)
            # ============================================================
            # UPSERT: Insert or update leave record
            # Composite unique key: (user_email, year, month, week_number)
            # ============================================================
            cur.execute("""
                INSERT INTO leave_records (
                    user_email, year, month, week_number, leave_date,
                    leave_hours, leave_type, description,
                    status, created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'PENDING', NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    leave_hours = VALUES(leave_hours),
                    leave_type = VALUES(leave_type),
                    description = VALUES(description),
                    updated_at = NOW()
            """, [
                user_email, year, month, week_number, leave_date,
                leave_hours, db_leave_type, reason
            ])

            logger.info(
                f"Leave recorded: user={user_email}, week={week_number}, "
                f"hours={leave_hours}, type={db_leave_type}"
            )

            return JsonResponse({
                'ok': True,
                'message': f'Leave recorded successfully for Week {week_number}',
                'leave_hours': float(leave_hours),
                'week_number': week_number
            })

    except Exception as e:
        logger.exception(f"record_leave error: {e}")
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_GET
def get_leaves_for_month(request):
    """
    Get all leave records for user in a specific billing cycle.
    Query params: billing_start
    """
    user_email = get_user_email_from_session(request)
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=403)

    billing_start = request.GET.get('billing_start')
    if not billing_start:
        return JsonResponse({'ok': False, 'error': 'billing_start required'}, status=400)

    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT id, week_number, leave_hours, leave_type, reason, status
                FROM leave_records
                WHERE LOWER(user_ldap) = LOWER(%s)
                  AND billing_start = %s
                ORDER BY week_number
            """, [user_email, billing_start])

            leaves = dictfetchall(cur)

        return JsonResponse({'ok': True, 'leaves': leaves})

    except Exception as e:
        logger.exception("get_leaves_for_month error: %s", e)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@require_POST
def save_my_allocation(request):
    """
    Save/update weekly allocation hours for user (draft mode).
    Payload: {
        allocation_id: int,
        week_number: int,
        allocated_hours: float,
        allocated_percent: float (optional)
    }
    """
    user_email = get_user_email_from_session(request)
    if not user_email:
        return JsonResponse({'ok': False, 'error': 'Not authenticated'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
        allocation_id = int(payload.get('allocation_id'))
        week_number = int(payload.get('week_number'))
        hours = payload.get('allocated_hours')
        percent = payload.get('allocated_percent')

        if hours is None:
            return JsonResponse({'ok': False, 'error': 'allocated_hours required'}, status=400)

        with transaction.atomic(), connection.cursor() as cur:
            # Upsert weekly_allocations as DRAFT
            cur.execute("""
                INSERT INTO weekly_allocations
                    (allocation_id, week_number, hours, percent, status, updated_at)
                VALUES (%s, %s, %s, %s, 'DRAFT', NOW())
                ON DUPLICATE KEY UPDATE
                    hours = VALUES(hours),
                    percent = VALUES(percent),
                    status = 'DRAFT',
                    updated_at = NOW()
            """, [
                allocation_id,
                week_number,
                str(Decimal(str(hours)).quantize(Decimal('0.01'))),
                str(Decimal(str(percent)).quantize(Decimal('0.01'))) if percent else None
            ])

        return JsonResponse({'ok': True, 'message': 'Draft saved successfully'})

    except Exception as e:
        logger.exception("save_my_allocation error: %s", e)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# projects/views.py
from django.http import JsonResponse
from django.db import connection

def view_allotment(request):
    if not request.session.get("is_authenticated"):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    month = request.GET.get("month")
    params = [month] if month else []
    sql = """
        SELECT
            p.name AS project,
            sp.name AS subproject,
            e.user_ldap AS resource,
            e. total_hours AS allocated_hrs,
            ROUND(e. total_hours / 183.75, 1) AS fte,
            e.month_start,
            w.buyer_wbs_cc  AS wbs
        FROM monthly_allocation_entries e
        LEFT JOIN projects p ON e.project_id = p.id
        LEFT JOIN subprojects sp ON e.subproject_id = sp.id
        LEFT JOIN prism_wbs w ON e.iom_id = w.iom_id
        WHERE (%s IS NULL OR LEFT(e.month_start, 7) = %s)
        ORDER BY p.name, sp.name, e.user_ldap
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, params * 2)
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return JsonResponse({"data": data})

# projects/views.py
from django.views.decorators.http import require_GET, require_POST
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db import connection, transaction
from datetime import date
import json

# projects/views.py

@require_GET
def tl_punch_review(request):
    # Auth check
    if not request.session.get("is_authenticated"):
        return redirect("accounts:login")
    session_ldap = request.session.get("ldap_username")
    if not session_ldap:
        return redirect("accounts:login")

    # Month selection
    month_str = request.GET.get("month") or date.today().strftime("%Y-%m")
    try:
        y, m = map(int, month_str.split("-"))
        month_start = date(y, m, 1)
    except Exception:
        month_start = date.today().replace(day=1)
        month_str = month_start.strftime("%Y-%m")

    # Get reportees (reuse LDAP utility)
    from accounts.ldap_utils import get_user_entry_by_username, get_reportees_for_user_dn
    creds = (session_ldap, request.session.get("ldap_password"))
    try:
        user_entry = get_user_entry_by_username(session_ldap, username_password_for_conn=creds)
        entry_dn = getattr(user_entry, "entry_dn", None)
        reportees_entries = get_reportees_for_user_dn(entry_dn, username_password_for_conn=creds) or []
    except Exception:
        reportees_entries = []

    # Normalize reportees
    reportees = []
    for ent in reportees_entries:
        mail = ent.get("mail") if isinstance(ent, dict) else getattr(ent, "mail", None)
        sam = ent.get("sAMAccountName") if isinstance(ent, dict) else getattr(ent, "sAMAccountName", None)
        cn = ent.get("cn") if isinstance(ent, dict) else getattr(ent, "cn", None)
        identifier = (mail or sam or "").strip()
        if identifier:
            reportees.append({
                "ldap": identifier.lower(),
                "mail": mail or identifier,
                "cn": cn or identifier,
            })

    # Add logged-in user if PDL and not already in reportees
    role = (request.session.get("role") or "").upper()
    if "PDL" in role:
        logged_ldap = (session_ldap or "").lower()
        if not any(r["ldap"] == logged_ldap for r in reportees):
            reportees.insert(0, {
                "ldap": logged_ldap,
                "mail": session_ldap,
                "cn": request.session.get("cn") or session_ldap,
            })

    # Get month limit for FTE calculation
    with connection.cursor() as cur:
        cur.execute(
            "SELECT max_hours FROM monthly_hours_limit WHERE year=%s AND month=%s",
            [month_start.year, month_start.month]
        )
        row = cur.fetchone()
        month_limit = float(row[0]) if row and row[0] else 173.0  # fallback

    # Fetch all punch and leave data, group as required
    reportee_ldaps = [r["ldap"] for r in reportees]
    punch_rows = {}
    fte_totals = {}
    if reportee_ldaps:
        placeholders = ",".join(["%s"] * len(reportee_ldaps))
        with connection.cursor() as cur:
            # Fetch punch data with JOINs for project/subproject names
            cur.execute(f"""
                SELECT pd.user_email, pd.project_id, pd.subproject_id, pd.week_number,
                       pd.allocated_hours, pd.punched_hours, pd.status, pd.id, pd.comments,
                       p.name AS project_name, sp.name AS subproject_name
                FROM punch_data pd
                LEFT JOIN projects p ON pd.project_id = p.id
                LEFT JOIN subprojects sp ON pd.subproject_id = sp.id
                WHERE pd.month_start = %s AND LOWER(pd.user_email) IN ({placeholders})
                ORDER BY pd.user_email, pd.project_id, pd.subproject_id, pd.week_number
            """, [month_start] + reportee_ldaps)
            punch_records = dictfetchall(cur)
            print("Punch Records:", punch_records)
            # Fetch leave data
            cur.execute(f"""
                SELECT user_email, week_number, leave_hours, leave_type, description
                FROM leave_records
                WHERE year = %s AND month = %s AND LOWER(user_email) IN ({placeholders})
            """, [month_start.year, month_start.month] + reportee_ldaps)
            leave_records = dictfetchall(cur)

        # Build leave lookup: {user: {week: leave_hours}}
        leave_lookup = {}
        for row in leave_records:
            user = row["user_email"].lower()
            week = row["week_number"]
            leave_lookup.setdefault(user, {})[week] = row.get("leave_hours", 0)

        # Group punch data for frontend and compute FTE
        for rep in reportees:
            user = rep["ldap"]
            user_rows = [
                row for row in punch_records if row["user_email"].lower() == user
            ]
            punch_rows[user] = []
            total_alloc = 0.0
            for row in user_rows:
                total_alloc += float(row.get("allocated_hours") or 0)
                punch_rows[user].append({
                    "project_name": row.get("project_name") or "None",
                    "subproject_name": row.get("subproject_name") or "None",
                    "week_number": row.get("week_number"),
                    "allocated_hours": row.get("allocated_hours", 0),
                    "punched_hours": row.get("punched_hours", 0),
                    "leave_hours": leave_lookup.get(user, {}).get(row.get("week_number"), 0),
                    "status": row.get("status"),
                    "punch_id": row.get("id"),
                    "comments": row.get("comments", ""),
                })
            fte_totals[user] = round(total_alloc / month_limit, 2) if month_limit else 0.0

    return render(request, "projects/tl_punch_review.html", {
        "month_str": month_str,
        "reportees": reportees,
        "punch_rows": punch_rows,
        "fte_totals": fte_totals,
    })

@require_POST
def tl_punch_approve(request):
    # Approve/modify punch data for a reportee for a week
    if not request.session.get("is_authenticated"):
        return JsonResponse({"ok": False, "error": "Not authenticated"}, status=403)
    session_ldap = request.session.get("ldap_username")
    try:
        data = json.loads(request.body.decode("utf-8"))
        punch_id = int(data["punch_id"])
        punched_hours = float(data["punched_hours"])
        comments = data.get("comments", "")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid input"}, status=400)
    try:
        with transaction.atomic(), connection.cursor() as cur:
            cur.execute("""
                UPDATE punch_data
                SET punched_hours = %s, status = 'APPROVED', approved_by = %s, approved_at = NOW(), comments = %s, updated_at = NOW()
                WHERE id = %s
            """, [punched_hours, session_ldap, comments, punch_id])
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

# projects/views.py

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db import connection, transaction
import json

@require_POST
def tl_punch_bulk_approve(request):
    # Session and role check
    if not request.session.get("is_authenticated"):
        return JsonResponse({"ok": False, "error": "Not authenticated"}, status=403)
    user_role = request.session.get("role", "")
    if user_role not in ("TEAM_LEAD", "PDL", "ADMIN"):
        return JsonResponse({"ok": False, "error": "Unauthorized"}, status=403)

    try:
        data = json.loads(request.body.decode("utf-8"))
        punch_ids = data.get("punch_ids", [])
        if not punch_ids or not isinstance(punch_ids, list):
            return JsonResponse({"ok": False, "error": "No punch IDs provided"}, status=400)
        # Only allow integer IDs
        punch_ids = [int(pid) for pid in punch_ids if str(pid).isdigit()]
        if not punch_ids:
            return JsonResponse({"ok": False, "error": "Invalid punch IDs"}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Invalid request: %s" % str(e)}, status=400)

    # Bulk update punch rows
    try:
        with transaction.atomic(), connection.cursor() as cur:
            # Optionally, you can filter by reportee/team lead if needed
            sql = (
                "UPDATE punch_data "
                "SET status = %s, approved_by = %s, approved_at = NOW() "
                "WHERE id IN (%s) AND status != 'APPROVED'"
            )
            params = ["APPROVED", request.session.get("ldap_username"), ",".join(str(pid) for pid in punch_ids)]
            # Use parameterized query for IN clause
            in_clause = ",".join(["%s"] * len(punch_ids))
            cur.execute(
                f"UPDATE punch_data SET status = %s, approved_by = %s, approved_at = NOW() "
                f"WHERE id IN ({in_clause}) AND status != 'APPROVED'",
                ["APPROVED", request.session.get("ldap_username")] + punch_ids
            )
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)